"""
excel_writer.py — Write StatementTable objects to Data_* sheets in an Excel file.

Rules:
- Only writes/deletes sheets with the Data_* prefix
- Never modifies sheets with other prefixes (My_*, Analysis, etc.)
- Full rewrite on every call — ensures restatements are captured

Layout per sheet:
    Col A  = Std Name (concept/label)
    Col B  = Original Item (company's XBRL label; empty if none)
    Col C+ = quarterly data, oldest → newest

    Row 1: A=ticker, B=empty, C..=quarter labels
    Row 2: A=empty,  B=empty, C..=filing dates
    Row 3+: A=concept, B=original_item, C..=values
"""

import os
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from fetcher_gaap import StatementTable
from excel_formatter import format_workbook
from fiscal_input import apply_fiscal_year_input
from zh_labels import zh_label, axis_label, ratio_label, meta_label

# 版面：A 英文標準名 / B 中文說明 / C 公司原始 XBRL 標籤 / D 起各期數據
# 中文說明表在 zh_labels.py，改那裡不影響任何計算邏輯（程式一律用 A 欄英文名比對）。
_DATA_START_COL = 4

# ── 覆蓋防護（TODO 第 8 項，2026-08-01）────────────────────────────────────
#
# 覆蓋既有檔前先留一份備份。整批替換 Data_* 沒有版本控制，第二次抓的年份範圍
# 較窄時舊季度會直接消失（GAAP 沒有快取，重抓要再等一分鐘）。
# 只保留一份滾動備份，不堆檔案。不想要就設 False。
KEEP_BACKUP   = True
BACKUP_SUFFIX = ".bak.xlsx"


def check_output_writable(output_path: str | Path) -> str | None:
    """檢查輸出檔現在能不能寫。可寫回 None，否則回中文錯誤訊息。

    要在**抓取開始前**呼叫。原本的失敗點在 `wb.save()`——那是所有下載與 AI
    呼叫都跑完的最後一步，檔案被 Excel 開著時使用者白等一分多鐘才看到錯誤。
    """
    path = Path(output_path)
    if not path.exists():
        return None      # 檔案不存在，write_statements 會自己建（含父資料夾）

    try:
        handle = open(path, "r+b")
    except PermissionError:
        return (f"輸出檔無法寫入（可能正被 Excel 開啟）：{path.name}\n"
                f"請先關閉該檔案再重試。")
    except OSError as exc:
        return f"輸出檔無法寫入：{path.name}（{type(exc).__name__}）"

    try:
        # Excel 開檔時多半在 open() 就擋下來了，但不同版本行為不一，
        # 再用非阻塞鎖確認一次。
        if os.name == "nt":
            try:
                import msvcrt
                msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
                msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
            except OSError:
                return (f"輸出檔正被其他程式鎖定（可能是 Excel）：{path.name}\n"
                        f"請先關閉該檔案再重試。")
    finally:
        handle.close()
    return None


def _backup_path(output_path: Path) -> Path:
    return output_path.with_name(output_path.stem + BACKUP_SUFFIX)


def _meta_fy_end_month(tables: list[StatementTable]) -> int:
    """從 Data_Meta 取財年結束月。找不到退回 12（日曆年）。"""
    meta = next((t for t in tables if t.sheet_name == "Data_Meta"), None)
    if meta is None or "Fiscal Year End Month" not in meta.concepts:
        return 12
    vals = meta.values[meta.concepts.index("Fiscal Year End Month")]
    try:
        return int(vals[0])
    except (IndexError, TypeError, ValueError):
        return 12


def write_statements(tables: list[StatementTable], output_path: str | Path,
                     template_path: str | Path | None = None) -> None:
    """Write StatementTable list to an Excel file.

    Replaces all existing Data_* sheets. Preserves all other sheets.
    Creates the file (and parent directory) if absent.

    If template_path is provided and the file exists, the template's cell
    formatting is preserved; only values are written. New quarter columns beyond
    the template width copy the format of the last template column.

    Args:
        tables:        List of StatementTable objects to write.
        output_path:   Path to .xlsx file.
        template_path: Optional path to template.xlsx with pre-set formatting.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    use_template = bool(template_path) and Path(template_path).exists()

    if use_template:
        wb = load_workbook(template_path)
        # Remove Data_* sheets that are NOT in the template (stale sheets)
        tbl_names = {t.sheet_name for t in tables}
        for name in list(wb.sheetnames):
            if name.startswith("Data_") and name not in tbl_names:
                del wb[name]
        for tbl in tables:
            if tbl.sheet_name in wb.sheetnames:
                ws = wb[tbl.sheet_name]
                _write_sheet_template(ws, tbl)
            else:
                # Sheet not in template: create and apply standard formatting
                ws = wb.create_sheet(tbl.sheet_name)
                _write_sheet(ws, tbl)
        # Still build the Index sheet
        from excel_formatter import _build_index_sheet, _apply_column_widths, _set_freeze_panes
        _build_index_sheet(wb, tables)
    else:
        if output_path.exists():
            wb = load_workbook(output_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
        for name in list(wb.sheetnames):
            if name.startswith("Data_"):
                del wb[name]
        for tbl in tables:
            ws = wb.create_sheet(tbl.sheet_name)
            _write_sheet(ws, tbl)
        format_workbook(wb, tables)

    # 期間表頭改成公式，由 Index 上那格「財年起始月」驅動。放在最後——
    # format_workbook 會重建 Index，先寫輸入格會被蓋掉。
    apply_fiscal_year_input(wb, _meta_fy_end_month(tables))

    # 先寫暫存檔再置換：save() 中途失敗（磁碟滿、被鎖）時原檔完好，
    # 不會留下一個寫到一半、開不起來的 xlsx。
    tmp_path = output_path.with_name(output_path.name + ".tmp")
    try:
        try:
            wb.save(tmp_path)
        finally:
            wb.close()

        if KEEP_BACKUP and output_path.exists():
            import shutil
            try:
                shutil.copy2(output_path, _backup_path(output_path))
            except OSError as exc:
                # 備份失敗不該擋住主要輸出，但要讓使用者知道這次沒有後路
                print(f"[excel_writer] 備份失敗（{type(exc).__name__}），"
                      f"仍會寫入 {output_path.name}", file=sys.stderr)

        os.replace(tmp_path, output_path)
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def _copy_cell_format(src, dst) -> None:
    """Copy formatting (fill, font, border, alignment, number_format) from src to dst."""
    import copy
    if src.has_style:
        dst.font        = copy.copy(src.font)
        dst.fill        = copy.copy(src.fill)
        dst.border      = copy.copy(src.border)
        dst.alignment   = copy.copy(src.alignment)
        dst.number_format = src.number_format


def _write_sheet_template(ws: Worksheet, tbl: StatementTable) -> None:
    """Write StatementTable into an existing template worksheet.

    Preserves all existing cell formatting. Clears old data values in C+
    columns, then writes new values. For columns beyond the template width,
    copies the format from the last pre-existing data column.
    """
    # Determine how many data columns the template already has
    template_max_col = ws.max_column  # last column with any content/format

    # Clear existing data values in C+ columns (rows 1+), keep format
    for col in range(_DATA_START_COL, template_max_col + 1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=col).value = None

    n_quarters = len(tbl.quarter_labels)
    last_tpl_data_col = template_max_col  # column to copy format from for new cols

    # Write row 1: ticker (A1), quarter labels (C1+)
    ws.cell(row=1, column=1).value = tbl.ticker or None
    ws.cell(row=1, column=2).value = None
    ws.cell(row=1, column=3).value = None
    for i, label in enumerate(tbl.quarter_labels):
        col = _DATA_START_COL + i
        cell = ws.cell(row=1, column=col)
        cell.value = label
        if col > template_max_col:
            _copy_cell_format(ws.cell(row=1, column=last_tpl_data_col), cell)

    # Write row 2: filing dates (C2+)
    ws.cell(row=2, column=1).value = None
    ws.cell(row=2, column=2).value = None
    ws.cell(row=2, column=3).value = None
    for i, dt in enumerate(tbl.filing_dates):
        col = _DATA_START_COL + i
        cell = ws.cell(row=2, column=col)
        cell.value = dt
        if col > template_max_col:
            _copy_cell_format(ws.cell(row=2, column=last_tpl_data_col), cell)

    # Write row 3+: concept (A), label (B), values (C+)
    has_labels = bool(tbl.labels)
    for row_offset, (concept, row_values) in enumerate(zip(tbl.concepts, tbl.values)):
        row = 3 + row_offset
        ws.cell(row=row, column=1).value = concept
        ws.cell(row=row, column=2).value = zh_label(concept) or None
        ws.cell(row=row, column=3).value = tbl.labels[row_offset] if has_labels else None
        for i, val in enumerate(row_values):
            col = _DATA_START_COL + i
            cell = ws.cell(row=row, column=col)
            cell.value = val
            if col > template_max_col:
                _copy_cell_format(ws.cell(row=row, column=last_tpl_data_col), cell)


def _col_b(tbl: StatementTable, concept: str, row_offset: int) -> str:
    """B 欄內容（跟著介面語言走）。A 欄的英文機器鍵是查表的 key。

    每張 sheet 的 A 欄住在不同的命名空間，所以要分流：三表是科目
    （`acct.*`）、Data_Segments 是維度軸（`axis.*`）、Data_Ratios 是比率名
    （`ratio.*`）、Data_Meta 是欄位名（`meta.*`）。查不到一律留白。
    """
    if tbl.sheet_name == "Data_Segments":
        raw_axis = tbl.labels[row_offset] if tbl.labels and row_offset < len(tbl.labels) else ""
        return axis_label(raw_axis)
    if tbl.sheet_name == "Data_Ratios":
        return ratio_label(concept)
    if tbl.sheet_name == "Data_Meta":
        return meta_label(concept)
    return zh_label(concept)


def _write_sheet(ws: Worksheet, tbl: StatementTable) -> None:
    """Write one StatementTable into a worksheet.

    Layout:
        Row 1: A=ticker, B=empty, C..=quarter labels
        Row 2: A=empty,  B=empty, C..=filing dates
        Row 3+: A=concept, B=original_item, C..=values
    """
    # Row 1: ticker in A1; quarter labels from C1
    ws.cell(row=1, column=1, value=tbl.ticker or None)
    ws.cell(row=1, column=2, value=None)
    ws.cell(row=1, column=3, value=None)
    for col_idx, label in enumerate(tbl.quarter_labels, start=_DATA_START_COL):
        ws.cell(row=1, column=col_idx, value=label)

    # Row 2: filing dates from C2
    ws.cell(row=2, column=1, value=None)
    ws.cell(row=2, column=2, value=None)
    ws.cell(row=2, column=3, value=None)
    for col_idx, date_str in enumerate(tbl.filing_dates, start=_DATA_START_COL):
        ws.cell(row=2, column=col_idx, value=date_str)

    # Row 3+: concept (A), original item (B), values (C+)
    has_labels = bool(tbl.labels)
    for row_offset, (concept, row_values) in enumerate(zip(tbl.concepts, tbl.values)):
        row = 3 + row_offset
        ws.cell(row=row, column=1, value=concept)
        ws.cell(row=row, column=2, value=_col_b(tbl, concept, row_offset) or None)
        ws.cell(row=row, column=3,
                value=tbl.labels[row_offset] if has_labels else None)
        for col_idx, val in enumerate(row_values, start=_DATA_START_COL):
            ws.cell(row=row, column=col_idx, value=val)
