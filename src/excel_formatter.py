"""
excel_formatter.py — Apply professional formatting to Data_* sheets and generate Index sheet.

Public API:
    format_workbook(wb, tables) -> None

Called by excel_writer.write_statements() before wb.save(). Modifies cell values
(÷1M unit conversion) and applies openpyxl styles. Does not change sheet structure.
"""

from __future__ import annotations
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fetcher_gaap import StatementTable
from datetime import date
from override_engine import check_key_rows
from i18n import t, excel_font
import metric_rules

# ── 字型（2026-08-08）──────────────────────────────────────────────────────
#
# 以前所有 Font() 都沒帶 name=，Excel 各自套預設（英文 Calibri、中文新細明體），
# 同一張表混兩種字體。統一吃這個常數，日後換字型只改這一行。
# openpyxl 的 Font 是不可變的，不能事後設 .name——所以一律走 _font()。
# CTH 選的是「全部含數字欄」：整份一致優先於數字等寬對齊（2026-08-08 決定）。
# 繁中的字型。實際採用哪個由 i18n.excel_font() 依語言決定——微軟正黑體
# 缺日文假名字形，日文要 Yu Gothic。字面只住在 i18n.LANGUAGES 一處，
# 這裡推導出來，避免兩邊各寫一次然後漂移。
FONT_NAME = excel_font("zh_tw")

# Index 字級（2026-08-08 CTH 指定，整體放大一級）。fiscal_input 也吃這裡。
INDEX_TITLE_SIZE = 16    # A1 公司抬頭
INDEX_META_SIZE  = 10    # A2 抓取日期／最新期間那一列
INDEX_TABLE_SIZE = 11    # sheet 清單、品質明細、財年起始月的標籤
INDEX_INPUT_SIZE = 12    # B4 黃底輸入格
INDEX_NOTE_SIZE  = 10    # 財年核對提醒


# 數字用等寬字型（2026-08-08 CTH 追加）。微軟正黑體的數字不等寬，同一欄的
# 1,234,567 與 89,012 位數對不齊，翻財報時很難掃。判斷依據是**儲存格的值是不是
# 數字**，不是欄號——Data_Meta／Data_Segments 的 D 欄起是文字，用欄號會誤傷。
NUMBER_FONT_NAME = "Consolas"


def _font(numeric: bool = False, **kwargs) -> Font:
    """所有字型一律走這裡。numeric=True 給數字用等寬字型。"""
    return Font(name=NUMBER_FONT_NAME if numeric else excel_font(), **kwargs)


def _is_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


# Index 上留給「財年起始月」輸入格 + 提醒的列數（第 4、5 列）。
# 實際內容由 fiscal_input 寫，這裡只負責讓位——見該模組的說明。
FY_INPUT_ROWS = 5

# 與 excel_writer 一致：A 英文名 / B 中文 / C 原始標籤 / D 起數據
_DATA_START_COL = 4

# ── Colours (ARGB) ────────────────────────────────────────────────────────
NAVY_DARK = "FF1F3864"
NAVY_MID  = "FF2D4A82"
BLUE_MID  = "FF2E75B6"
GREY_SEP  = "FFEEEEEE"
ROW_ALT   = "FFF5F8FF"
ROW_WHITE = "FFFFFFFF"
BLUE_HDR  = "FFDDE8F5"

# ── Row classification ────────────────────────────────────────────────────
import nongaap_layout as _ng_layout

# 分隔標題列：值全空、整列上色。Data_NonGAAP 的四個分區也算在內，
# 否則它們會被當成一般指標列去套 ÷1M。
# 三表併在同一個 sheet，捲動時容易分不清現在看的是哪一張，所以三個 section
# 標題各給一種底色，overflow 區用灰色。顏色是唯一的視覺線索，不要拿掉。
SECTION_COLOURS = {
    "Income Statement":    "FF1F3864",   # 深藍
    "Balance Sheet":       "FF31859B",   # 藍綠
    "Cash Flow":           "FF7B4F9D",   # 紫
    "Other (as reported)": "FF808080",   # 灰（公司特有科目，非模板）
    "Ratios":              "FF2E75B6",
}

SECTION_HEADERS = ({"Income Statement", "Balance Sheet", "Cash Flow"}
                   | set(_ng_layout.ALL_SECTIONS) | set(SECTION_COLOURS))

SUBTOTAL_CONCEPTS = {
    "Gross Profit", "Total Operating Expense", "Operating Income",
    "Pre-tax Income", "Net Income",
    "Total Current Assets", "Total Non-current Assets", "Total Assets",
    "Total Current Liabilities", "Total Non-current Liabilities", "Total Liabilities",
    "Total Equity — Parent", "Total Equity incl. NCI", "Total Liabilities & Equity",
    "Operating Cash Flow", "Investing Cash Flow", "Financing Cash Flow", "Free Cash Flow",
}

# Index 上每張 sheet 的一句話說明。key 是 sheet 名稱（機器鍵，不翻譯），
# 值走 i18n。
SHEET_DESCRIPTION_KEYS = (
    "Data_Financials(Q)", "Data_Financials(Y)", "Data_EPS_Recon", "Data_NonGAAP",
    "Data_Std", "Data_Segments", "Data_Ratios", "Data_Meta",
)


# ── 數值分類 ──────────────────────────────────────────────────────────────
#
# 關鍵字表在 metric_rules.py（唯一可調整處）。判斷順序：每股 → 百分比 → 股數 → 金額，
# 先命中先套用。順序有意義：「Non-GAAP 每股盈餘」同時含「每股」與（若加了寬鬆詞）
# 可能的百分比詞，每股必須優先。
#
# 決策（2026-08-01，使用者選定）：百分比存成 **Excel 原生百分比**——37.5 存為
# 0.375，格式 "0.0%"，儲存格顯示 37.5%。好處是在 Excel 裡拉公式、畫圖、算加權
# 平均都直接可用，不必手動 ÷100。
# 改回存原始數字（37.5 + '#,##0.0"%"'）把這個開關設 False 即可，兩種都有測試覆蓋。
PERCENT_AS_EXCEL_RATIO = True


def _keyword_matcher(keywords: tuple[str, ...]):
    """把關鍵字表編成一條 regex。

    ASCII 關鍵字要求詞界，中日韓字元用裸比對（中文沒有詞界的概念）。
    詞界是必要的，不是潔癖：'Operations' 含 'ratio'、'Corporate' 含 'rate'、
    'Steps' 含 'eps'，裸子字串比對會讓 XBRL overflow 的金額行不再除以 1M。
    """
    parts = []
    for kw in keywords:
        esc = re.escape(kw.casefold())
        if re.search(r"[a-z0-9]", kw.casefold()):
            parts.append(rf"(?<![a-z0-9]){esc}(?![a-z0-9])")
        else:
            parts.append(esc)
    pattern = re.compile("|".join(parts))
    return lambda c: bool(pattern.search((c or "").casefold()))


_is_eps_concept    = _keyword_matcher(metric_rules.EPS_KEYWORDS)
_is_percent_concept = _keyword_matcher(metric_rules.PERCENT_KEYWORDS)
_is_shares_concept = _keyword_matcher(metric_rules.SHARES_KEYWORDS)


def _sheet_description(name: str) -> str:
    if name in SHEET_DESCRIPTION_KEYS:
        return t(f"xls.sheet_desc.{name}")
    if name.startswith("Data_Seg_"):
        return t("xls.sheet_desc.segment_detail", axis=name[9:])
    return name


def _fill(hex_argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_argb)


# ── Column widths ─────────────────────────────────────────────────────────

def _apply_column_widths(ws) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 34    # 中文說明
    ws.column_dimensions["C"].width = 30    # 公司原始 XBRL 標籤
    for col in range(4, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13


# ── Freeze panes ──────────────────────────────────────────────────────────

def _set_freeze_panes(ws) -> None:
    ws.freeze_panes = "D3"


# ── Number formatting and unit conversion ────────────────────────────────

FMT_FINANCIAL = "#,##0.0_ ;[Red](#,##0.0)"
FMT_EPS       = "#,##0.00_ ;[Red](#,##0.00)"
FMT_SHARES    = "#,##0"
FMT_PERCENT   = "0.0%" if PERCENT_AS_EXCEL_RATIO else '#,##0.0"%"'
FMT_MULTIPLE  = '#,##0.00"x"'
FMT_DAYS      = '#,##0.0"d"'

# Data_Ratios 的列名一定帶單位後綴，格式照後綴走、且一律不 ÷1,000,000。
# 後綴判斷必須**優先於**關鍵字判斷：「流動比率 (x)」含「率」會被關鍵字當成
# 百分比而 ÷100，「DSO (days)」不含任何關鍵字會被當金額 ÷1,000,000。
_UNIT_SUFFIX_FORMATS = {
    "(%)":     (FMT_PERCENT,  100 if PERCENT_AS_EXCEL_RATIO else 1),
    "(x)":     (FMT_MULTIPLE, 1),
    "(days)":  (FMT_DAYS,     1),
    "($)":     (FMT_EPS,      1),
    "($mm)":   (FMT_FINANCIAL, 1_000_000),
}


def _unit_suffix_rule(concept: str) -> tuple[str, int] | None:
    """列名帶單位後綴時回 (格式, 除數)，否則回 None 交給關鍵字判斷。"""
    for suffix, rule in _UNIT_SUFFIX_FORMATS.items():
        if concept.endswith(suffix):
            return rule
    return None


def unit_format_for(concept: str) -> tuple[str, int]:
    """回傳 (Excel 數字格式, 除數)。判斷順序：單位後綴 → 每股 → 百分比 →
    股數 → 預設金額（÷1,000,000）。跟 format_workbook() 內的判斷順序完全一致，
    給 Data_Ratios 以外的表（如跨公司比較）共用同一套規則，不要各寫一份。
    """
    rule = _unit_suffix_rule(concept)
    if rule is not None:
        return rule
    if _is_eps_concept(concept):
        return FMT_EPS, 1
    if _is_percent_concept(concept):
        return FMT_PERCENT, 100 if PERCENT_AS_EXCEL_RATIO else 1
    if _is_shares_concept(concept):
        return FMT_SHARES, 1_000_000
    return FMT_FINANCIAL, 1_000_000


def _apply_row_styles(ws) -> None:
    """Apply fill and font styles to all rows."""
    def _paint(cells, fill, **font_kw) -> None:
        """整列上色，數字格換等寬字型（其餘樣式完全相同）。"""
        text_font = _font(**font_kw)
        num_font  = _font(numeric=True, **font_kw)
        for cell in cells:
            cell.fill = fill
            cell.font = num_font if _is_number(cell.value) else text_font

    # Row 1: ticker / quarter labels — dark navy
    _paint(ws[1], _fill(NAVY_DARK), color="FFFFFFFF", bold=True, size=11)

    # Row 2: filing dates — medium navy
    _paint(ws[2], _fill(NAVY_MID), color="FFAABBCC", size=9)

    # Row 3+: classify by col A value
    for row_idx in range(3, ws.max_row + 1):
        concept = ws.cell(row=row_idx, column=1).value or ""
        concept = str(concept).strip()

        if concept in SECTION_HEADERS:
            row_fill  = _fill(SECTION_COLOURS.get(concept, BLUE_MID))
            font_kw   = dict(color="FFFFFFFF", bold=True, size=10)
            row_height = 16
        elif concept == "":
            row_fill  = _fill(GREY_SEP)
            font_kw   = dict(size=9)
            row_height = 6
        else:
            row_fill  = _fill(ROW_WHITE) if row_idx % 2 == 0 else _fill(ROW_ALT)
            font_kw   = dict(bold=True) if concept in SUBTOTAL_CONCEPTS else {}
            row_height = None

        _paint(ws[row_idx], row_fill, **font_kw)
        if row_height is not None:
            ws.row_dimensions[row_idx].height = row_height


def _apply_number_formats(ws) -> None:
    """Convert values to millions and apply number formats. Skips section/blank rows."""
    for row_idx in range(3, ws.max_row + 1):
        concept = str(ws.cell(row=row_idx, column=1).value or "").strip()

        # Section headers and blank separators have no numeric data
        if concept in SECTION_HEADERS or concept == "":
            continue

        # 單位後綴優先於一切關鍵字判斷（Data_Ratios 用）；判斷順序見 unit_format_for()
        fmt, divisor = unit_format_for(concept)

        for col_idx in range(_DATA_START_COL, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value / divisor
                cell.number_format = fmt


# ── Quality check colours ─────────────────────────────────────────────────
QUALITY_GREEN   = "FF1A7A34"
QUALITY_ORANGE  = "FFC25C00"
QUALITY_MISS_BG = "FFFFF0E0"
QUALITY_MISS_FG = "FFC00000"

# 完整度區塊最多列幾列問題。全部列出來會把 Index 撐到上百列——實測 TSLA 有
# 19 列有洞。超過的用一行「還有 N 列」帶過，細節在 Data_Financials 上看得到。
_DQ_MAX_LISTED = 10

ALL_KEY_ROWS = [
    "Revenue", "Operating Income", "Net Income", "Diluted EPS",
    "Total Assets", "Total Liabilities", "Total Equity — Parent",
    "Operating Cash Flow", "Capex",
]


def _compute_quality(tables: list):
    """Data_Financials(Q) 的缺漏判斷。沒有 Q 表回 None。

    2026-08-22 換掉舊的「9 個關鍵列、最近 4 期全空才算缺」——那個判準寬到
    形同虛設（NVDA 顯示 `9/9 ✓`，實際上 95 個欄位有 27 個幾乎全空、3 季完全
    沒資料）。改用 `data_quality.assess()` 的四個判斷，細節見該模組說明。
    """
    q_tbl = next((t for t in tables if t.sheet_name == "Data_Financials(Q)"), None)
    if q_tbl is None:
        return None
    from data_quality import assess          # 延後 import：避免匯入期循環
    return assess(q_tbl)


def _issue_count(report) -> int:
    return (sum(g.count for g in report.missing_quarters)
            + len(report.holed) + len(report.contradictions))


# ── Index sheet ───────────────────────────────────────────────────────────

def _build_index_sheet(wb: Workbook, tables: list) -> None:
    """Insert or replace the Index sheet at position 0."""
    if "Index" in wb.sheetnames:
        del wb["Index"]

    ticker       = tables[0].ticker if tables else ""
    company_name = ""
    meta = next((t for t in tables if t.sheet_name == "Data_Meta"), None)
    if meta and len(meta.concepts) > 1 and len(meta.values) > 1 and meta.values[1]:
        company_name = meta.values[1][0] or ""

    header_text = f"{ticker} — {company_name}" if company_name else ticker

    quality = _compute_quality(tables)

    ws = wb.create_sheet("Index", 0)

    # Row 1: company header
    ws["A1"] = header_text
    ws["A1"].fill = _fill(NAVY_DARK)
    ws["A1"].font = _font(color="FFFFFFFF", bold=True, size=INDEX_TITLE_SIZE)
    ws.merge_cells("A1:E1")

    # Row 2: metadata。使用者要在第一頁就看到「資料抓到哪一季、那季何時結束、
    # 公司財年怎麼算」——不必再翻到 Data_Meta。
    def _meta_value(name: str) -> str:
        if meta is None or name not in meta.concepts:
            return ""
        vals = meta.values[meta.concepts.index(name)]
        return str(vals[0]) if vals else ""

    latest_period = _meta_value("Latest Period")
    latest_end    = _meta_value("Latest Period End")
    fy_span       = _meta_value("Fiscal Year Span")

    bits = [t("xls.index.fetched_on", date=date.today())]
    if latest_period:
        bits.append(t("xls.index.data_through", period=latest_period)
                    + (t("xls.index.period_end_paren", end=latest_end) if latest_end else ""))
    if fy_span:
        bits.append(t("xls.index.fy_span", span=fy_span))
    bits.append(t("xls.index.source"))

    ws["A2"] = "　　".join(bits)
    ws["A2"].fill = _fill(NAVY_MID)
    ws["A2"].font = _font(color="FFAABBCC", size=INDEX_META_SIZE)
    ws.merge_cells("A2:E2")

    # Row 3: 抓取缺漏警告，沒缺漏時留白當間隔列。
    #
    # 放這裡不放下面的「品質明細」區塊：那一區講的是「這個科目 SEC 沒報」，
    # 是資料本身的性質；缺漏講的是「這次抓取沒拿到」，是這一份檔案的狀態，
    # 重抓可能就有了。兩件事混在一起使用者會分不出哪個該重抓。
    # 擺在第一頁最上方是因為 GUI 的 log 關掉就沒了，而使用者真正會搞混的
    # 時點是三天後重新打開這份 Excel 的時候。
    gap_note = _meta_value("Fetch Gaps")
    if gap_note and gap_note != t("xls.meta.none"):
        ws["A3"] = gap_note
        ws["A3"].fill = _fill(QUALITY_MISS_BG)
        ws["A3"].font = _font(color=QUALITY_MISS_FG, bold=True, size=INDEX_META_SIZE)
        ws["A3"].alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells("A3:E3")
        # 一行約 90 字元；換行後高度要跟著長，否則長訊息會被切掉只看到開頭
        ws.row_dimensions[3].height = 15 * max(1, (len(gap_note) // 60) + 1)
    else:
        ws.row_dimensions[3].height = 6

    # Row 4-5 由 fiscal_input.apply_fiscal_year_input() 填「財年起始月」輸入格
    # 與提醒（在 write_statements 最後才寫，因為這裡每次都會重建整張 Index）。
    # 這裡只把表格往下讓位，不要在中間插列——插列會讓那格的位址跟著動。
    _TABLE_HDR_ROW = FY_INPUT_ROWS + 1

    # 表格標題列
    hdr_font = _font(bold=True, size=INDEX_TABLE_SIZE)
    hdr_fill = _fill(BLUE_HDR)
    headers = ["Sheet", t("xls.index.col_desc"), t("xls.index.col_earliest"),
               t("xls.index.col_latest"), t("xls.index.col_complete")]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=_TABLE_HDR_ROW, column=col, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill

    # 之後每張 Data_* sheet 一列
    data_sheets = [t for t in tables if t.sheet_name.startswith("Data_")]
    for i, tbl in enumerate(data_sheets):
        row = _TABLE_HDR_ROW + 1 + i
        earliest = tbl.quarter_labels[0]  if tbl.quarter_labels else "—"
        latest   = tbl.quarter_labels[-1] if tbl.quarter_labels else "—"

        is_primary = tbl.sheet_name in ("Data_Financials(Q)", "Data_Financials(Y)")
        name_font  = _font(color="FF1F3864" if is_primary else "FF666666",
                          bold=is_primary, size=INDEX_TABLE_SIZE)
        row_fill   = _fill(ROW_WHITE) if row % 2 == 0 else _fill(ROW_ALT)

        for col, val in enumerate([tbl.sheet_name,
                                    _sheet_description(tbl.sheet_name),
                                    earliest, latest], start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            # B 欄是中文說明——不指定字型的話會掉回新細明體，
            # 整張 Index 就會混兩種字體（這正是 D0b-2 要修的）。
            cell.font = name_font if col == 1 else _font(size=INDEX_TABLE_SIZE)

        # ── E 欄：完成度 ──
        e_cell = ws.cell(row=row, column=5)
        e_cell.fill = row_fill
        if tbl.sheet_name == "Data_Financials(Q)" and quality is not None:
            if quality.template_mismatch:
                e_cell.value = t("xls.index.dq_mismatch_short")
                e_cell.font = _font(color=QUALITY_MISS_FG, size=INDEX_TABLE_SIZE)
            elif _issue_count(quality) == 0:
                e_cell.value = "✓"
                e_cell.font = _font(color=QUALITY_GREEN, size=INDEX_TABLE_SIZE)
            else:
                e_cell.value = t("xls.index.dq_issues_short", n=_issue_count(quality))
                e_cell.font = _font(color=QUALITY_ORANGE, size=INDEX_TABLE_SIZE)
        else:
            e_cell.value = "—"
            e_cell.font = _font(color="FF999999", size=INDEX_TABLE_SIZE)

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10

    # ── 資料完整度區塊 ─────────────────────────────────────────────────────
    if quality is None:
        return

    next_row = _TABLE_HDR_ROW + 1 + len(data_sheets) + 2   # blank row gap

    hdr_cell = ws.cell(row=next_row, column=1, value=t("xls.index.quality_detail"))
    hdr_cell.fill = _fill(NAVY_DARK)
    hdr_cell.font = _font(color="FFFFFFFF", bold=True, size=INDEX_TABLE_SIZE)
    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=2)
    next_row += 1

    def _line(label: str, value: str, color: str = "FF333333") -> None:
        nonlocal next_row
        a = ws.cell(row=next_row, column=1, value=label)
        b = ws.cell(row=next_row, column=2, value=value)
        for c in (a, b):
            c.fill = _fill(ROW_WHITE)
            c.font = _font(color=color, size=INDEX_TABLE_SIZE)
        next_row += 1

    # 模板不適用時，逐列的判斷沒有意義——先講這件事，其餘不列
    if quality.template_mismatch:
        _line(t("xls.index.dq_template_mismatch"),
              t("xls.index.dq_sparse_n", n=len(quality.sparse_periods),
                total=quality.total_periods),
              QUALITY_MISS_FG)
        return

    gaps = sum(g.count for g in quality.missing_quarters)
    _line(t("xls.index.dq_periods"),
          t("xls.index.dq_periods_v", n=quality.total_periods, missing=gaps),
          QUALITY_MISS_FG if gaps else QUALITY_GREEN)
    if quality.sparse_periods:
        _line(t("xls.index.dq_sparse"),
              t("xls.index.dq_sparse_n", n=len(quality.sparse_periods),
                total=quality.total_periods), QUALITY_ORANGE)

    # 矛盾：整列全空、但同一家公司的相關欄位顯示它應該要有值。可信度最高，放最前面
    for c in quality.contradictions[:_DQ_MAX_LISTED]:
        _line(f"⚠ {c.row}", t("xls.index.dq_contradiction_v", evidence=c.evidence),
              QUALITY_MISS_FG)
    # 中間有洞：同一列有些期有、有些沒有。一定是漏抓
    for h in quality.holed[:_DQ_MAX_LISTED]:
        _line(f"⚠ {h.row}", t("xls.index.dq_holed_v", have=h.have, span=h.span),
              QUALITY_ORANGE)
    rest = max(0, len(quality.holed) - _DQ_MAX_LISTED)
    if rest:
        _line("", t("xls.index.dq_more", n=rest), "FF999999")
    # 零星有值：填滿率太低，多半是公司本來就沒有這項活動（併購、借款、買回）。
    # **刻意不用 ⚠、也不算進問題數**——它是參考資訊，不是抓漏（見 data_quality
    # 的 `_SPORADIC_FILL_RATIO`，52 家實測那些洞只有 8~22% 是真的漏抓）。
    for sp in quality.sporadic[:_DQ_MAX_LISTED]:
        _line(sp.row, t("xls.index.dq_sporadic_v", have=sp.have, span=sp.span),
              "FF999999")
    if not quality.contradictions and not quality.holed and not gaps:
        _line("", t("xls.index.dq_clean"), QUALITY_GREEN)


# ── Public API ────────────────────────────────────────────────────────────

def format_workbook(wb: Workbook, tables: list[StatementTable]) -> None:
    """Apply formatting to all Data_* sheets."""
    _build_index_sheet(wb, tables)
    for ws in wb.worksheets:
        if not ws.title.startswith("Data_"):
            continue
        _apply_column_widths(ws)
        _set_freeze_panes(ws)
        _apply_row_styles(ws)
        if ws.title != "Data_Meta":
            _apply_number_formats(ws)
