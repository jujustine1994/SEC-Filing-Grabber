"""fiscal_input.py — 讓使用者自己改財年起始月，其他期間標籤用公式帶。

**為什麼要做這個**：財年結束月是程式從 10-K 的 XBRL 欄名偵測出來的
（`fetcher_gaap._detect_fy_end_month()`）。偵測錯了，`Data_Financials` 的財季
標籤就整排錯，而使用者除了重跑程式沒有別的辦法，看到的還只是一堆寫死的文字。

改法：Index 上放**一格可以改的財年起始月**，其他地方全部用 Excel 公式從它
推算。程式猜錯時，使用者改那一格，整本活頁簿的財季立刻跟著更新。

    Index!B4        使用者輸入（黃底），定義名稱 FY_START_MONTH
    Data_*  第 5 列  期末結算日 —— XBRL 的真實日期，**公式的錨，永遠是靜態值**
    Data_*  第 1 列  期間標籤   ── 公式
    Data_*  第 3 列  財季       ── 公式
    Data_*  第 4 列  日曆季     ── 公式（只看期末日，與財年無關）

## 內縮 15 天

所有換算都先把期末日往前推 15 天再取年月。美股多用 52/53 週制，期末日會在
月底前後浮動最多 6 天——WDC 的 FY2026 Q2 結束在 **2026-01-02**，直接看月份會
算成 Q3，整整差一季。`docs/8k-period-off-by-one.md` 量化過這個誤差：COST /
WDC / PANW 有 7 份就是栽在這裡。往前推 15 天一定會落回該季最後一個月。

## 沒有連動的地方（刻意的）

`Data_Ratios` 與 `Data_Meta` 是 Python 算好寫死的，改起始月不會重算。比率的
YoY 是按欄位相對位置取的，不受標籤文字影響；`Data_Meta` 的「財年起訖」會與
使用者改過的值不一致——這是已知取捨，全部改成公式的複雜度不划算。Index 上的
說明有寫。
"""
from __future__ import annotations

import math
import re
from calendar import monthrange
from datetime import date, timedelta

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# 輸入格寫在 Index 上，字型／字級必須跟 excel_formatter 建的表格一致，
# 否則同一頁會混兩種字體。常數只有一份，在 excel_formatter。
from i18n import t, excel_font
from excel_formatter import (
    INDEX_TABLE_SIZE, INDEX_INPUT_SIZE, INDEX_NOTE_SIZE,
)


def _font(**kwargs) -> Font:
    # 每次呼叫重查，不在 import 時綁死：語言是 import 之後才設定的，
    # 綁死的話日文 Index 會留在微軟正黑體，同一頁混兩種字體。
    return Font(name=excel_font(), **kwargs)

# 使用者輸入格與它的定義名稱。公式一律引用定義名稱而不是 Index!$B$4——
# 日後 Index 版面調整，公式不必跟著改。
FY_START_CELL = "B4"
FY_START_DEFINED_NAME = "FY_START_MONTH"

# 期間表頭在各 Data_* sheet 的固定列號
ROW_PERIOD_LABEL = 1
ROW_FISCAL_QUARTER = 3
ROW_CALENDAR_QUARTER = 4
ROW_PERIOD_END = 5

_DATA_START_COL = 4

# 52/53 週制的期末日浮動不會超過 ±6 天，往前推 15 天必定落在該季最後一個月。
_SHRINK_DAYS = 15

_ISO_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")

_INPUT_FILL = PatternFill("solid", fgColor="FFFFF2CC")   # 淡黃＝可編輯
_INPUT_BORDER = Border(*[Side(style="thin", color="FFBF8F00")] * 4)


# ── Python 參考實作 ─────────────────────────────────────────────────────────
#
# 這幾個函式是 Excel 公式的規格：同一套邏輯寫兩次（Python 一次、公式一次），
# 測試釘住 Python 這份，改公式時對照著改。

def fy_start_month(fy_end_month: int) -> int:
    """財年結束月 → 起始月。AAPL 9 月結束 → 10 月開始。"""
    return fy_end_month % 12 + 1


def _anchor(period_end: str | None) -> date | None:
    """期末日字串 → 內縮 15 天後的日期。不是 ISO 日期回 None。"""
    m = _ISO_RE.match((period_end or "").strip())
    if m is None:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3))) - timedelta(days=_SHRINK_DAYS)
    except ValueError:
        return None


def _fiscal_year(d: date, start_month: int) -> int:
    """財年 N 結束於西元 N 年（SEC 慣例），所以起始月之後的月份屬於下一個財年。"""
    return d.year + 1 if start_month > 1 and d.month >= start_month else d.year


def fiscal_quarter_of(period_end: str | None, start_month: int) -> str:
    """期末日 + 財年起始月 → `FY2026Q2`。算不出來回空字串。"""
    d = _anchor(period_end)
    if d is None:
        return ""
    quarter = (d.month - start_month) % 12 // 3 + 1
    return f"FY{_fiscal_year(d, start_month)}Q{quarter}"


# ── 零下載規則：發布日 + EDGAR fiscal_year_end → 財季（B5）──────────────────
#
# Item 2.02 8-K 的 `period_of_report` 放的是**發布日**不是財期結束日，直接換算
# 有系統性 off-by-one（實測 119 份只有 16 份對，偏 -3 到 +1 季）。這條規則改成
# 從 EDGAR 的 `Company.fiscal_year_end`（完整 MMDD，例 WDC "0703"）往前推
# 3/6/9 個月得到「名目季末」，再取不晚於發布日的最新一個。
#
#   200 份 Item 2.02 8-K 實測：in-sample 113/113、out-of-sample 44/44，**皆 100%**
#   （基準是下載新聞稿抓期末日算出的 `fiscal_label`）。完整驗證見
#   `docs/superpowers/report-2026-08-25-8k-years-zero-download-rule.md`。
#
# **傳完整 MMDD 不要只傳月份**：報告實測的規則 B（名目季末一律取財季結束月的
# 月底）只有 79.8%，52/53 週制的真實季末離月底最多 20 天（WDC 季末 1/2、
# COST Q3 季末 5/10）。⚠ 在本實作的錨定方式下（候選 = MMDD 往前推 + tol=21），
# 200 份實測「用月底」與「用 MMDD 那天」給出的 label 完全相同——差別只在需要
# 多大的 tol 才吃得下（COST 那份 MMDD 要 2 天、月底要 3 天）。所以這句是照
# 報告的結論保留的保險做法，不是本實作量出來的差異。

# 名目季末可以比發布日晚幾天仍然採用。COST Q3 真實 5/10 結束、名目算出來 5/30，
# 而它 5/28 就發了。實測 tol 在 3~30 之間命中率完全相同（高原 27 天寬），
# tol=0 與 tol>=35 都掉到 95.0%，所以 21 不是硬調出來的魔術數字。
_ANNOUNCE_TOL_DAYS = 21

# sanity check：選中的名目季末離發布日超過這麼多天就不採用（回空字串）。
# 實測最大發布延遲 58 天（200 份，範圍 -2~58），70 留了緩衝。
#
# ⚠ 這道檢查**擋不住「公司改過財年」**那個風險：候選季末永遠相隔 89~92 天，
# 所以選中的那個必然落在 [-tol, 70] 內，tol=21 時門檻算術上碰不到。它擋的是
# 參數被改壞（tol 被放大）與畸形輸入，不是 fiscal_year_end 漂移。真的要偵測
# 漂移，靠的是下載後 `fiscal_label` 與這個 label 對不對得起來（cli.py 有比對）。
_MAX_ANNOUNCE_LAG_DAYS = 70

_MMDD_RE = re.compile(r"^(\d{2})(\d{2})$")


def _month_shifted(year: int, month: int, day: int) -> date | None:
    """`year-month-day` 往前推月份後的日期。該月沒有那一天就退到當月最後一天。"""
    while month < 1:
        year, month = year - 1, month + 12
    try:
        return date(year, month, min(day, monthrange(year, month)[1]))
    except ValueError:
        return None


def _to_date(value: str | date | None) -> date | None:
    """`"20260129"`／`"2026-01-29"`／`date` → `date`。認不得回 None。"""
    if isinstance(value, date):
        return value
    raw = str(value or "").strip().replace("-", "")
    if len(raw) != 8 or not raw.isdigit():
        return None
    try:
        return date(int(raw[:4]), int(raw[4:6]), int(raw[6:8]))
    except ValueError:
        return None


def quarter_label_from_announcement(
    announce_date: str | date | None,
    fiscal_year_end_mmdd: str | None,
    tol: int = _ANNOUNCE_TOL_DAYS,
    max_lag_days: int = _MAX_ANNOUNCE_LAG_DAYS,
) -> str:
    """發布日 + EDGAR `fiscal_year_end`（MMDD）→ `FY2026Q2`。零 I/O、純日期運算。

    算不出來一律回**空字串**，呼叫端要自己退回舊算法——EDGAR 少一個欄位不該
    讓整批列清單失敗。回空字串的情況：MMDD 不合法、發布日不合法、或選中的
    名目季末離發布日超過 `max_lag_days`（sanity check，見上方常數註解）。
    """
    m = _MMDD_RE.match(str(fiscal_year_end_mmdd or "").strip())
    announced = _to_date(announce_date)
    if m is None or announced is None:
        return ""
    fye_month, fye_day = int(m.group(1)), int(m.group(2))
    if not (1 <= fye_month <= 12) or not (1 <= fye_day <= 31):
        return ""
    if _month_shifted(2000, fye_month, fye_day) is None:
        return ""

    limit = announced + timedelta(days=tol)
    best: date | None = None
    for year in (limit.year + 1, limit.year, limit.year - 1):
        for shift in (0, 3, 6, 9):
            cand = _month_shifted(year, fye_month - shift, fye_day)
            if cand is not None and cand <= limit and (best is None or cand > best):
                best = cand
    if best is None:
        return ""
    if not -tol <= (announced - best).days <= max_lag_days:
        return ""
    return fiscal_quarter_of(best.isoformat(), fy_start_month(fye_month))


def fiscal_year_of(period_end: str | None, start_month: int) -> str:
    """年報用：期末日 + 財年起始月 → `FY2025`。"""
    d = _anchor(period_end)
    return f"FY{_fiscal_year(d, start_month)}" if d else ""


# ── 日曆季：一份實作、兩個具名基準點 ────────────────────────────────────────
#
# 「這一季算哪個日曆季」有兩種問法，而且**兩種都要**：
#
#   basis="end"   結算季 —— 這一季**結束**在哪個日曆季。單一公司 Data_* 第 4 列
#                 用這個，因為它正下方第 5 列就是期末日，兩列必須自洽。
#   basis="span"  對齊季 —— 這一季的**多數天數**落在哪個日曆季。跨公司比較用
#                 這個：NVDA 7 月底結束那季要跟 AMD/INTC 6 月底那季擺同一欄
#                 （同一波財報，分析師就是這樣比），不是跟 AMD 9 月那季。
#
# 2026-08-22 之前這兩件事有三套算法散在三個檔案（其中一套還忘了內縮，
# 直接把 INTC 結束在 2023-04-01 的 Q1 算成 2023Q2）。合併成一份實作之後，
# **`basis` 刻意不給預設值**——強迫每個呼叫端表態要哪種語意，這就是「統一」
# 的實質保障：以後只有一個地方會算錯，而且每個使用點都看得出它要什麼。
#
# 兩個基準點都先把期末日往前推再取月份，差別只在推多遠：
#   15 天 → 回到該季**最後一個月**（吃掉 52/53 週制的月底漂移）
#   45 天 → 回到該季**中點**（13 週季的一半）
# 期中點是離日曆季邊界最遠的位置，所以最穩；期初日反而最不穩（13 週季的
# 起訖日都剛好落在邊界附近，AMD 6 月季的期初日在 3/30，會翻到前一季）。

_BASIS_DAYS = {"end": 15, "span": 45}
_HALF_YEAR_END_MONTH = 5   # 財年結束在 1-5 月 → 主要落在前一個日曆年


def _shifted(period_end: str | None, days: int) -> date | None:
    """期末日往前推 `days` 天。不是 ISO 日期回 None。"""
    m = _ISO_RE.match((period_end or "").strip())
    if m is None:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3))) - timedelta(days=days)
    except ValueError:
        return None


def calendar_quarter_of(period_end: str | None, *, basis: str) -> str:
    """期末日 → 日曆季 `2026Q2`。與財年無關。算不出來回空字串。

    `basis` 沒有預設值，一定要明講要哪一種（理由見上方區塊註解）：
        "end"   結算季——這一季結束在哪個日曆季
        "span"  對齊季——這一季的多數天數落在哪個日曆季
    """
    if basis not in _BASIS_DAYS:
        raise ValueError(f"basis must be one of {sorted(_BASIS_DAYS)}, got {basis!r}")
    d = _shifted(period_end, _BASIS_DAYS[basis])
    return f"{d.year}Q{(d.month - 1) // 3 + 1}" if d else ""


# ── 跨公司對齊（calendarization）────────────────────────────────────────────
#
# 這兩個是 basis="span" 的薄包裝。存在的理由是讓 comparison.py 的呼叫端讀起來
# 是「取對齊季」而不是「取日曆季，基準點傳 span」——語意寫在名字裡，不要讓
# 呼叫端自己記字串。不可以在這裡加任何額外邏輯，測試釘住它們與 span 等價。


def calendarized_quarter_of(period_end: str | None) -> str:
    """期末日 → 跨公司對齊用的日曆季 `2025Q2`。算不出來回空字串。"""
    return calendar_quarter_of(period_end, basis="span")


def calendarized_year_of(period_end: str | None) -> str:
    """財年結束日 → 跨公司對齊用的日曆年 `2025`。算不出來回空字串。

    慣例：財年結束在 1-5 月掛前一年（NVDA FY2026 結束在 2026-01，內容其實
    是日曆 2025 年），6-12 月掛當年（MSFT 6 月結束的 FY2025 就叫 2025）。
    """
    d = _anchor(period_end)
    if d is None:
        return ""
    return str(d.year - 1 if d.month <= _HALF_YEAR_END_MONTH else d.year)


# ── Excel 公式 ──────────────────────────────────────────────────────────────

def _date_expr(col: str) -> str:
    """第 5 列的 ISO 日期文字 → 內縮 15 天的 Excel 日期序列值。

    不用 `DATEVALUE`：它認不認得 `2026-04-26` 要看使用者的地區設定，
    自己拆 LEFT/MID 再組 `DATE()` 在哪台電腦都一樣。
    """
    c = f"{col}{ROW_PERIOD_END}"
    return f"DATE(VALUE(LEFT({c},4)),VALUE(MID({c},6,2)),VALUE(MID({c},9,2)))-{_SHRINK_DAYS}"


def _fiscal_year_expr(d: str) -> str:
    n = FY_START_DEFINED_NAME
    return f"(YEAR({d})+IF(AND({n}>1,MONTH({d})>={n}),1,0))"


def _quarter_expr(d: str) -> str:
    return f"(INT(MOD(MONTH({d})-{FY_START_DEFINED_NAME},12)/3)+1)"


def _guard(col: str, body: str) -> str:
    """沒有期末日就留空，不要讓使用者看到 #VALUE!。"""
    return f'=IF({col}{ROW_PERIOD_END}="","",{body})'


def period_label_formula(col: str, annual: bool = False) -> str:
    """第 1 列的期間標籤：季報 `FY2026Q2`、年報 `FY2025`。"""
    d = _date_expr(col)
    body = f'"FY"&{_fiscal_year_expr(d)}'
    if not annual:
        body += f'&"Q"&{_quarter_expr(d)}'
    return _guard(col, body)


def fiscal_quarter_formula(col: str) -> str:
    """第 3 列的財季 `FY2026FQ2`。

    財季用 `FQ`、日曆季用純數字，視覺上就分得開——非 12 月結算的公司同一欄
    可能是 FY2026FQ1 但日曆 2025Q4，看錯就是整整一季。
    """
    d = _date_expr(col)
    return _guard(col, f'"FY"&{_fiscal_year_expr(d)}&"FQ"&{_quarter_expr(d)}')


def calendar_quarter_formula(col: str) -> str:
    """第 4 列的日曆季 `2026Q2`。只看期末日，不引用財年起始月。"""
    d = _date_expr(col)
    return _guard(col, f'YEAR({d})&"Q"&(INT((MONTH({d})-1)/3)+1)')


# ── 套用 ────────────────────────────────────────────────────────────────────

def _note() -> str:
    """B4 輸入格底下那段提醒。每次呼叫重查——語言可能在這之間換過。"""
    return t("xls.fy_input.note")

# 財年區間：唯一「改 1 個月就會變」的即時回饋。沒有它，使用者把 2 改成 3
# 看到標籤沒動，會以為公式壞了（2026-08-08 CTH 實際回報）。
# DATE(2000, m+11, 1) 讓月份自己進位，不必寫 MOD。
def _xl_str(s: str) -> str:
    """把 Python 字串包成 Excel 公式裡的字串常值。

    不可以用 `repr()` 再把單引號換成雙引號：`repr()` 遇到字串內含單引號時
    會自己改用雙引號包，那時 replace 會把公式切碎（`="Fisc" l "&TEXT(...)`），
    Excel 開起來是 #NAME? 或乾脆拒絕開檔。英文譯文出現撇號（`Company's`）
    完全是可預期的事。Excel 的逸出規則是雙引號寫兩次。
    """
    return '"' + s.replace('"', '""') + '"'


def _fy_span_formula() -> str:
    """財年區間的 Excel 公式。

    前綴／分隔／後綴／月份格式都走 i18n。月份格式必須可換：中文與日文是
    `10 月 – 9 月`（`"m"` 出數字，後面自己接「月」），英文要的是
    `FY Oct – Sep`（`"mmm"` 才會出月份簡稱，`"m"` 只會給你 `FY 10 – 9`）。
    """
    n = FY_START_DEFINED_NAME
    fmt = _xl_str(t("xls.fy_input.span_month_format"))
    return (f'=IF({n}="","",{_xl_str(t("xls.fy_input.span_prefix"))}'
            f'&TEXT(DATE(2000,{n},1),{fmt})&{_xl_str(t("xls.fy_input.span_sep"))}'
            f'&TEXT(DATE(2000,{n}+11,1),{fmt})&{_xl_str(t("xls.fy_input.span_suffix"))})')


# 一行的高度（含行距）。10pt 字約 13.5pt，其他字級按比例。
_LINE_HEIGHT_AT_10PT = 13.5


def _wrapped_row_height(ws, text: str, first_col: int = 1, last_col: int = 5,
                        size: float = INDEX_NOTE_SIZE) -> float:
    """算 wrap 過的合併儲存格要多高。

    **合併儲存格不會自動調整列高**——這是 Excel 的行為，不是 openpyxl 的限制，
    所以只能自己算。原本寫死 28，CTH 2026-08-08 驗收時回報文字被切掉：實測要
    6 行、28 只夠 2.4 行。寫死一個大一點的數字治標，提醒文字一改又會壞，所以
    改成依文字長度與實際欄寬推算。

    中文是全形（一個字佔兩個半形當量），而 Excel 的欄寬單位就是半形字寬。
    `column_dimensions` 沒設過寬度時回 None，用 Excel 預設的 8.43。
    多抓一行餘裕——換行不會剛好斷在邊界，寧可留白也不要切到字。
    """
    width = sum(ws.column_dimensions[get_column_letter(c)].width or 8.43
                for c in range(first_col, last_col + 1))
    display = sum(2 if ord(ch) > 0x2000 else 1 for ch in text)
    lines = math.ceil(display / max(width, 1)) + 1
    return round(lines * _LINE_HEIGHT_AT_10PT * (size / 10), 1)


def _write_input_block(ws, start_month: int, row: int = 4) -> None:
    """在 Index 寫出可編輯的輸入格與提醒。"""
    label = ws.cell(row=row, column=1, value=t("xls.fy_input.label"))
    label.font = _font(bold=True, size=INDEX_TABLE_SIZE)

    cell = ws.cell(row=row, column=2, value=start_month)
    cell.fill = _INPUT_FILL
    cell.border = _INPUT_BORDER
    cell.font = _font(bold=True, size=INDEX_INPUT_SIZE, color="FFBF8F00")
    cell.alignment = Alignment(horizontal="center")
    cell.number_format = "0"

    span = ws.cell(row=row, column=3, value=_fy_span_formula())
    span.font = _font(size=INDEX_TABLE_SIZE, color="FF666666")

    note = ws.cell(row=row + 1, column=1, value=_note())
    note.font = _font(size=INDEX_NOTE_SIZE, color="FFBF8F00")
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=5)
    ws.row_dimensions[row + 1].height = _wrapped_row_height(ws, _note())


def _is_annual(ws) -> bool:
    """年報 sheet 的期間標籤沒有季別（`FY2025` 而不是 `FY2025Q1`）。"""
    for col in range(_DATA_START_COL, ws.max_column + 1):
        value = str(ws.cell(ROW_PERIOD_LABEL, col).value or "")
        if value:
            return not re.search(r"Q[1-4]$", value)
    return False


def _apply_to_sheet(ws) -> None:
    annual = _is_annual(ws)
    for col_idx in range(_DATA_START_COL, ws.max_column + 1):
        period_end = str(ws.cell(ROW_PERIOD_END, col_idx).value or "")
        if not _ISO_RE.match(period_end.strip()):
            # 舊申報沒帶期末日，公式沒有錨可用——保留原本寫死的標籤。
            continue
        col = get_column_letter(col_idx)
        ws.cell(ROW_PERIOD_LABEL, col_idx).value = period_label_formula(col, annual)
        ws.cell(ROW_CALENDAR_QUARTER, col_idx).value = calendar_quarter_formula(col)
        if not annual:
            ws.cell(ROW_FISCAL_QUARTER, col_idx).value = fiscal_quarter_formula(col)


def apply_fiscal_year_input(wb, fy_end_month: int) -> None:
    """在 Index 放輸入格，並把 Data_Financials 的期間表頭改成公式。

    沒有 Index 或沒有 Data_Financials 都安靜跳過——欄位不全的活頁簿
    （只抓年報、只抓 segment）也要能正常寫出來。
    """
    if "Index" not in wb.sheetnames:
        return

    start_month = fy_start_month(fy_end_month or 12)
    _write_input_block(wb["Index"], start_month)

    ref = f"'Index'!${FY_START_CELL[0]}${FY_START_CELL[1:]}"
    if FY_START_DEFINED_NAME in wb.defined_names:
        del wb.defined_names[FY_START_DEFINED_NAME]
    wb.defined_names.add(DefinedName(FY_START_DEFINED_NAME, attr_text=ref))

    for name in ("Data_Financials(Q)", "Data_Financials(Y)"):
        if name in wb.sheetnames:
            _apply_to_sheet(wb[name])

    # openpyxl 不算公式，寫出去的儲存格沒有快取值。不強制重算的話，Excel 有機會
    # 直接顯示空白（看起來像整排標籤不見了）。
    wb.calculation.fullCalcOnLoad = True
