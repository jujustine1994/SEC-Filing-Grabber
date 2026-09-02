"""comparison_writer.py — 把 comparison.py 的資料結構寫成跨公司比較 Excel。

Sheet 結構（見 docs/superpowers/specs/2026-08-20-cross-company-comparison-design.md）：
  Compare_Data    — 唯一一張原始資料表，每個指標一個區塊往下疊
  Snapshot        — 活的，公式驅動的單一時間點快照
  Snapshot_Manual — 空白，供人工貼值凍結存檔
  Chart_<指標>     — 每個指標各一張，只放圖表
"""

from __future__ import annotations

import datetime
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.layout import Layout
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from comparison import ComparisonResult
from data_quality import missing_quarters
from excel_formatter import FMT_FINANCIAL, unit_format_for
from fiscal_input import calendarized_quarter_of
from i18n import t

_HEADER_FONT = Font(bold=True)
_BLOCK_GAP = 1  # 區塊之間空幾列
_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# ── 說明 sheet（G7，2026-08-25）────────────────────────────────────────────
#
# 「這份檔案用了哪些定義」一次講清楚，不要讓使用者自己猜。CTH 明講這張表
# **未來會擴充**（開發中發現新的定義問題就往裡加），所以做成資料驅動：新增
# 一條只要在 NOTE_ITEMS 加一行 + 四個 locale 各加兩條，不可以把文字寫死在
# 版面程式裡。
#
# 每一條都帶「這份檔案是否真的踩到」的勾選——使用者不必讀完十二條去猜哪條
# 跟他有關。判定一律由實際資料算出來，結構性的條目（單位、符號慣例）永遠勾。

_CHECK_MARK = "✓"
# 13 週 × 7 天。用來由某一季的期末日往後推算財年結束月，跟
# `data_quality._QUARTER_DAYS` 是同一個數字、同一個理由（不是隨手取的數字）。
_QUARTER_DAYS = 91
_CALENDAR_QUARTER = re.compile(r"^\d{4}Q[1-4]$")
# 單一缺口最多補幾欄。實測 52 家沒有任何 >210 天（>2 季）的缺口，真的出現就是
# 資料異常，不該讓程式生出一長串假期間。跟 `data_quality._MAX_GAP_QUARTERS`
# 同一個理由、同一個值（那邊是判定上限，這邊是版面上限）。
_MAX_GAP_COLUMNS = 4
_WRAP = Alignment(vertical="top", wrap_text=True)


@dataclass(frozen=True)
class NoteItem:
    """說明 sheet 的一條。`applies` 拿算好的事實回傳 (要不要勾, 實際情況文字)。"""
    title_key: str
    body_key: str
    applies: Callable[["NotesFacts"], tuple[bool, str]]
    # True 時「沒踩到就整列不出現」。目前只有「本檔案缺少的公司」用得上——
    # 沒有公司失敗時列一條空的警告只會讓人以為出了事。
    hide_when_clear: bool = False


@dataclass(frozen=True)
class NotesFacts:
    """說明 sheet 每一條的判定材料，從 ComparisonResult 一次算完。"""
    companies: list[str] = field(default_factory=list)
    periods: list[str] = field(default_factory=list)
    fy_end_months: dict[str, int] = field(default_factory=dict)
    end_spread: tuple[int, str] | None = None      # (最大天數差, 差最多的那一欄)
    synthetic_periods: list[str] = field(default_factory=list)
    blank_cells: int = 0
    total_cells: int = 0
    failures: list = field(default_factory=list)


def _parse_iso(value: str) -> datetime.date | None:
    try:
        return datetime.date.fromisoformat((value or "").strip())
    except ValueError:
        return None


def _fy_end_month(fiscal_map: dict[str, str], ends: dict[str, str]) -> int | None:
    """這家公司的財年結束月。算不出來回 None。

    有 Q4（或年報）那一期就直接取它的期末日月份；沒有就拿任一季往後推
    13 週 × 剩餘季數。**不走 `fiscal_input` 那套「內縮 15 天」**：那是為了
    判斷「結束在哪一個日曆季」，用在這裡會把 AVGO（財年結束在 11-03 這種
    月初）算成 10 月，跟公司自己講的月份對不上。
    """
    dated = [(fiscal, ends.get(period, ""))
             for period, fiscal in fiscal_map.items()]
    for fiscal, end in dated:
        d = _parse_iso(end)
        if d is not None and (fiscal.endswith("Q4") or "Q" not in fiscal):
            return d.month
    for fiscal, end in dated:
        d = _parse_iso(end)
        if d is None or "Q" not in fiscal or not fiscal[-1].isdigit():
            continue
        return (d + datetime.timedelta(days=_QUARTER_DAYS * (4 - int(fiscal[-1])))).month
    return None


def collect_notes_facts(
    result: ComparisonResult, metric_names: list[str],
    companies: list[str], periods: list[str],
) -> NotesFacts:
    """把說明 sheet 需要的判定材料一次算完——版面程式只負責畫，不做判斷。"""
    fy_end_months: dict[str, int] = {}
    for company in companies:
        month = _fy_end_month(result.fiscal_labels.get(company, {}),
                              result.period_ends.get(company, {}))
        if month is not None:
            fy_end_months[company] = month

    # 同一欄各公司期末日差幾天。取全表差最多的那一欄當代表——要講的是「這件事
    # 在這份檔案裡有多嚴重」，不是逐欄列表。
    end_spread: tuple[int, str] | None = None
    for period in periods:
        dates = [d for d in (_parse_iso(result.period_ends.get(c, {}).get(period, ""))
                             for c in companies) if d is not None]
        if len(dates) < 2:
            continue
        days = (max(dates) - min(dates)).days
        if days and (end_spread is None or days > end_spread[0]):
            end_spread = (days, period)

    synthetic = sorted({
        period
        for company in companies
        for period in result.synthetic_q4.get(company, set())
        if period in periods
    })

    blank = total = 0
    for metric_name in metric_names:
        metric_data = result.metrics.get(metric_name, {})
        for company in companies:
            company_data = metric_data.get(company, {})
            for period in periods:
                total += 1
                if company_data.get(period) is None:
                    blank += 1

    return NotesFacts(
        companies=list(companies),
        periods=list(periods),
        fy_end_months=fy_end_months,
        end_spread=end_spread,
        synthetic_periods=synthetic,
        blank_cells=blank,
        total_cells=total,
        failures=list(result.failures),
    )


def _always(_facts: NotesFacts) -> tuple[bool, str]:
    return True, ""


def _fiscal_years_differ(facts: NotesFacts) -> tuple[bool, str]:
    months = facts.fy_end_months
    if len(set(months.values())) <= 1:
        return False, t("compare.xls.notes.detail_same_fy_month")
    detail = t("compare.xls.notes.list_sep").join(
        t("compare.xls.notes.detail_fy_month", ticker=c, month=months[c])
        for c in sorted(months, key=lambda c: (months[c], c))
    )
    return True, detail


def _fiscal_years_differ_quietly(facts: NotesFacts) -> tuple[bool, str]:
    """跟上一條同一個判定，實際情況那欄不重複寫一次。"""
    return _fiscal_years_differ(facts)[0], ""


def _period_ends_differ(facts: NotesFacts) -> tuple[bool, str]:
    if facts.end_spread is None:
        return False, t("compare.xls.notes.detail_same_period_end")
    days, period = facts.end_spread
    return True, t("compare.xls.notes.detail_end_spread", days=days, period=period)


def _has_synthetic_q4(facts: NotesFacts) -> tuple[bool, str]:
    if not facts.synthetic_periods:
        return False, t("compare.xls.notes.detail_no_synth_q4")
    return True, t("compare.xls.notes.detail_synth_q4",
                   periods=t("compare.xls.notes.list_sep").join(facts.synthetic_periods))


def _has_blanks(facts: NotesFacts) -> tuple[bool, str]:
    if not facts.blank_cells:
        return False, t("compare.xls.notes.detail_no_blank")
    return True, t("compare.xls.notes.detail_blank",
                   blank=facts.blank_cells, total=facts.total_cells)


def _period_span(facts: NotesFacts) -> tuple[bool, str]:
    if not facts.periods:
        return True, ""
    return True, t("compare.xls.notes.detail_period_span",
                   first=facts.periods[0], last=facts.periods[-1])


def _missing_companies(facts: NotesFacts) -> tuple[bool, str]:
    """整家抓不到的公司。這條講的**就是「本份 Excel 有沒有缺東西」**——
    實例：`INTC_NVDA_AMD_TSM_v3.xlsx` 檔名有 TSM、使用者也選了 TSM，但檔案裡
    只有三家（TSM 報 20-F 不是 10-K），而從檔案完全看不出來：沒有錯誤訊息、
    沒有空欄位、圖上就是三條線。失敗紀錄原本只寫進 GUI log。"""
    if not facts.failures:
        return False, ""
    return True, t("compare.xls.notes.list_sep").join(
        t("compare.xls.notes.detail_failure", ticker=f.ticker, error=f.error_type)
        for f in facts.failures
    )


# 順序就是 Excel 上的順序。新增一條：這裡加一行，四個 locale 各加標題與內文兩條。
NOTE_ITEMS: tuple[NoteItem, ...] = (
    NoteItem("compare.xls.notes.timeline", "compare.xls.notes.timeline_body", _always),
    NoteItem("compare.xls.notes.not_fiscal", "compare.xls.notes.not_fiscal_body",
             _fiscal_years_differ),
    NoteItem("compare.xls.notes.not_period_end", "compare.xls.notes.not_period_end_body",
             _fiscal_years_differ_quietly),
    NoteItem("compare.xls.notes.period_end_row", "compare.xls.notes.period_end_row_body",
             _period_ends_differ),
    NoteItem("compare.xls.notes.synth_q4", "compare.xls.notes.synth_q4_body",
             _has_synthetic_q4),
    NoteItem("compare.xls.notes.blanks", "compare.xls.notes.blanks_body", _has_blanks),
    NoteItem("compare.xls.notes.source", "compare.xls.notes.source_body", _period_span),
    NoteItem("compare.xls.notes.units", "compare.xls.notes.units_body", _always),
    NoteItem("compare.xls.notes.scope", "compare.xls.notes.scope_body", _always),
    NoteItem("compare.xls.notes.missing_companies",
             "compare.xls.notes.missing_companies_body", _missing_companies,
             hide_when_clear=True),
    NoteItem("compare.xls.notes.sign", "compare.xls.notes.sign_body", _always),
    NoteItem("compare.xls.notes.as_reported", "compare.xls.notes.as_reported_body",
             _always),
)

_NOTES_COL_WIDTHS = (6, 26, 78, 40)


def write_notes_sheet(
    wb: Workbook, result: ComparisonResult, metric_names: list[str]
) -> None:
    """說明 sheet。擺在 Compare_Data 之後——定義要看得到，但不佔掉開檔的第一眼。

    sheet 名稱固定英文 `Notes`（跟其他 sheet 名稱一樣是機器鍵，不隨語言變），
    表內文字才跟著語言走。
    """
    companies, periods = visible_layout(result, metric_names)
    facts = collect_notes_facts(result, metric_names, companies, periods)

    ws = wb.create_sheet("Notes", 1)
    for i, width in enumerate(_NOTES_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.cell(row=1, column=1, value=t("compare.xls.notes.title")).font = _HEADER_FONT

    headers = (t("compare.xls.notes.col_check"), t("compare.xls.notes.col_item"),
               t("compare.xls.notes.col_body"), t("compare.xls.notes.col_detail"))
    for col, text in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=text).font = _HEADER_FONT

    row = 3
    for item in NOTE_ITEMS:
        applies, detail = item.applies(facts)
        if item.hide_when_clear and not applies:
            continue
        ws.cell(row=row, column=1, value=_CHECK_MARK if applies else None)
        ws.cell(row=row, column=2, value=t(item.title_key))
        ws.cell(row=row, column=3, value=t(item.body_key)).alignment = _WRAP
        ws.cell(row=row, column=4, value=detail or None).alignment = _WRAP
        row += 1


# ── G6：抓不到的季度留一整欄空白（2026-08-25）────────────────────────────
#
# 現況欄位清單是「成功抓到什麼就放什麼」，某一季掛掉整欄消失，畫面上 2025Q1
# 直接跳到 2025Q3，使用者與 AI 都看不出中間漏了一季。改成保留欄位、內容全空，
# 讓「有漏」這件事看得見。
#
# **判定不能用固定門檻**，要用 `round(天數差 / 91) - 1`（`_QUARTER_DAYS`）。
# 52 家 1,482 對相鄰期間實測：111~150 天那 16 筆全部是 COSTCO 的 16 週第四季
# （112~119 天），固定門檻（例如「>120 天算缺」）會把它們全部誤判成缺一季，
# 而 `round(112/91) = 1` 正確判為沒缺。這條規則跟單一公司那條線共用同一份
# 實作（`data_quality.missing_quarters()`），不要在這裡另外寫一份。

_ANNUAL_LABEL = re.compile(r"^\d{4}$")


def _fill_annual_gaps(periods: list[str]) -> list[str]:
    """年度輸出：欄位是純年份，缺的那一年補一欄。"""
    years = [int(p) for p in periods]
    filled = set(years)
    for a, b in zip(years, years[1:]):
        gap = b - a - 1
        if 0 < gap <= _MAX_GAP_COLUMNS:
            filled.update(range(a + 1, b))
    return [str(y) for y in sorted(filled)]


def _fill_quarter_gaps(result: ComparisonResult, periods: list[str]) -> list[str]:
    """季度輸出：由各公司自己的期末日序列算缺口，缺的那一季補一欄。

    為什麼要逐公司算、不能直接看「日曆季標籤有沒有連號」：COSTCO 的 16 週
    第四季在對齊季上本來就會跳過一格（沒有任何一季的中點落在那一個日曆季），
    照標籤連號補會補出一堆假缺口。天數差才是判準。

    只補在最早與最新之間（CTH 已定），而且只補**沒有任何公司抓到**的那一季
    ——別家有抓到的話那一欄本來就在。
    """
    first, last = periods[0], periods[-1]
    extra: set[str] = set()
    for company_ends in result.period_ends.values():
        ends = [e for e in company_ends.values() if e]
        for gap in missing_quarters(ends):
            after = datetime.date.fromisoformat(gap.after)
            for k in range(1, gap.count + 1):
                label = calendarized_quarter_of(
                    (after + datetime.timedelta(days=_QUARTER_DAYS * k)).isoformat())
                if label and first < label < last:
                    extra.add(label)
    return sorted(set(periods) | extra)


# F6（2026-09-03，CTH 選方案 B）：圖表 X 軸的真日期序列值，只給
# write_chart_sheets() 用，不動 Compare_Data 既有的期末結算日文字列（Snapshot
# 的 SUMPRODUCT/MATCH 靠那一列是文字才對得起來，見 write_snapshot_sheets()）。
_QUARTER_MID_MONTH_DAY = {1: (2, 15), 2: (5, 15), 3: (8, 15), 4: (11, 15)}


def _approx_period_date(period: str) -> datetime.date | None:
    """G6 補出來的空白欄（沒有任何公司抓到那一期）沒有真正的期末日可用。
    日期軸的類別留空會被 Excel 塌成 1899-12-30，把整條時間軸拉爆——give it a
    粗略的季／年中點代表日就好，反正這種欄位本來就沒有任何一家公司的資料，
    折線在這裡是斷點（`display_blanks="gap"`），近似值只影響 X 軸上的位置，
    不影響任何看得到的數字。算不出日曆季／年格式（退回財季標籤的殘骸欄）
    就回 None，交給呼叫端整格留空。"""
    if _ANNUAL_LABEL.match(period):
        return datetime.date(int(period), 6, 30)
    if _CALENDAR_QUARTER.match(period):
        year = int(period[:4])
        month, day = _QUARTER_MID_MONTH_DAY[int(period[-1])]
        return datetime.date(year, month, day)
    return None


def _fill_period_gaps(result: ComparisonResult, periods: list[str]) -> list[str]:
    """期間欄位清單 → 補上缺口欄之後的清單。"""
    if len(periods) < 2:
        return periods
    if all(_ANNUAL_LABEL.match(p) for p in periods):
        return _fill_annual_gaps(periods)
    if not all(_CALENDAR_QUARTER.match(p) for p in periods):
        # 期末日抓不到而退回財季標籤（`FY2009Q4`）的殘骸欄混在裡面時不補洞——
        # 那種標籤算不出日曆位置，補出來的欄位擺哪裡都是錯的。
        return periods
    return _fill_quarter_gaps(result, periods)


def visible_layout(
    result: ComparisonResult, metric_names: list[str]
) -> tuple[list[str], list[str]]:
    """這份檔案實際會出現在表上的 (公司清單, 期間欄位清單)。

    Compare_Data 的對應表、各指標區塊與說明 sheet 都吃這一份——三邊各算一次
    的話，哪天篩選規則改了就會有一邊沒跟上（而且是靜默不一致）。

    所有公司、所有指標都沒值的期間整欄拿掉。合成 Q4 時年報沒有期末日的那幾欄
    （`comparison._aligned_labels()` 算不出日曆季，退回 `FY2009Q4` 這種財季
    標籤）值全是空的，排序時又會被排到日曆季後面，圖表 X 軸就多出兩格最新一期
    之後的空白，看起來像資料抓錯。**判斷要跨所有指標一起做**：每個區塊各自篩
    會讓區塊之間欄數不同，而 write_chart_sheets() 用的是全表 `max_column`，
    窄的區塊會被讀到別人的欄位。
    """
    companies = sorted({
        company
        for metric_data in result.metrics.values()
        for company in metric_data
    })
    seen = {
        label
        for metric_data in result.metrics.values()
        for company_data in metric_data.values()
        for label in company_data
    }
    with_values = {
        label
        for metric_data in result.metrics.values()
        for company_data in metric_data.values()
        for label, value in company_data.items()
        if value is not None
    }
    # 對應表與說明 sheet 只看 metric_names 選到的指標，跟指標區塊同一個範圍
    selected = {
        label
        for metric_name in metric_names
        for company_data in result.metrics.get(metric_name, {}).values()
        for label in company_data
    }
    periods = sorted((seen & selected) - (seen - with_values))
    return companies, _fill_period_gaps(result, periods)


def _fiscal_map_cell(fiscal_label: str, period_end: str) -> str | None:
    """對應表的一格：`FY2026Q2 (0727)`。財季標籤缺了就整格留空——寧可空白，
    也不要寫一個看不出是哪一季的日期。"""
    if not fiscal_label:
        return None
    mmdd = period_end.replace("-", "")[4:8] if period_end else ""
    return f"{fiscal_label} ({mmdd})" if mmdd else fiscal_label


def _write_fiscal_map_block(
    ws, result: ComparisonResult, all_companies: list[str], all_periods: list[str]
) -> int:
    """Compare_Data 最上方的「日曆季 ↔ 財季」對應表。回傳下一個可用的列號。

    為什麼是對應表而不是「每家公司的財年開始月份」一行帶過：對應表的每一格
    都是**逐期從實際期末日算出來的**，公司哪一年改過財年，那一欄自己就會反映
    出來，不需要任何例外處理。財年開始月份只寫得下一個值，公司改過財年就直接
    失效（G2，設計書 2026-08-22）。

    下面的財務指標區塊只給日曆季、不重複財季——財季這件事在這裡一次講完。
    """
    title = ws.cell(row=1, column=1, value=t("compare.xls.fiscal_map_title"))
    title.font = _HEADER_FONT

    header_row = 2
    ws.cell(row=header_row, column=1, value=t("compare.xls.company"))
    for col, period in enumerate(all_periods, start=2):
        ws.cell(row=header_row, column=col, value=period)

    for offset, company in enumerate(all_companies):
        r = header_row + 1 + offset
        ws.cell(row=r, column=1, value=company)
        fiscal_map = result.fiscal_labels.get(company, {})
        ends = result.period_ends.get(company, {})
        for col, period in enumerate(all_periods, start=2):
            value = _fiscal_map_cell(fiscal_map.get(period, ""), ends.get(period, ""))
            if value is not None:
                ws.cell(row=r, column=col, value=value)

    return header_row + len(all_companies) + 1 + _BLOCK_GAP


def write_compare_data_sheet(
    wb: Workbook, result: ComparisonResult, metric_names: list[str]
) -> dict[str, tuple[int, int]]:
    """寫 Compare_Data。回傳 {指標名: (資料列起, 資料列迄)}（不含標題/期末結算日列），
    給 Snapshot 的 MATCH 公式與 Chart 的資料來源 range 用。"""
    ws = wb.active
    ws.title = "Compare_Data"

    all_companies, all_periods = visible_layout(result, metric_names)

    block_ranges: dict[str, tuple[int, int]] = {}
    row = _write_fiscal_map_block(ws, result, all_companies, all_periods)
    for metric_name in metric_names:
        metric_data = result.metrics.get(metric_name, {})
        fmt, divisor = unit_format_for(metric_name)

        # 每個指標區塊的欄位跟最上方的對應表對齊（同一份 visible_layout），
        # 區塊之間欄數不同的話 write_chart_sheets() 用的全表 `max_column`
        # 會讓窄的區塊讀到別人的欄位。
        periods: list[str] = list(all_periods)

        # 標題列
        title_cell = ws.cell(row=row, column=1, value=metric_name)
        title_cell.font = _HEADER_FONT
        header_row = row + 1
        ws.cell(row=header_row, column=1, value=t("compare.xls.company"))
        for col, period in enumerate(periods, start=2):
            ws.cell(row=header_row, column=col, value=period)

        # 期間鍵是日曆季（跨公司對齊，見 comparison._aligned_labels()），同一欄
        # 各公司的實際期末日不會一樣——NVDA 那一季結束在 7/27，AMD 是 6/28。
        # 這一格只放得下一個日期，取**最晚**的：Snapshot 拿它做「不晚於 B1」
        # 的判斷，取早的那個會讓 B1 設在 7/1 就顯示 NVDA 還沒結算完的數字。
        end_dates = [
            max(
                (result.period_ends.get(company, {}).get(period, "")
                 for company in all_companies),
                default="",
            )
            for period in periods
        ]

        # F6（2026-09-03，方案 B）：另外加一列真日期序列值，只給圖表當 X 軸類別
        # 用，不動下面那列文字——Snapshot 的 SUMPRODUCT/MATCH 靠那列是文字才
        # 對得起來（見 write_snapshot_sheets()），兩件事分開比同時改風險低。
        # 隱藏起來，不佔使用者眼球，但仍在檔案裡（不是不存在）。
        chart_date_row = header_row + 1
        ws.cell(row=chart_date_row, column=1, value=t("compare.xls.period_end_chart_date"))
        ws.row_dimensions[chart_date_row].hidden = True
        for col, (period, end_date) in enumerate(zip(periods, end_dates), start=2):
            date_value = (_parse_iso(end_date) if end_date
                          else _approx_period_date(period))
            if date_value is not None:
                cell = ws.cell(row=chart_date_row, column=col, value=date_value)
                cell.number_format = "yyyy-mm-dd"

        # 期末結算日列（靜態文字，供 Snapshot 用）。fetcher_gaap 給的原始格式是
        # "YYYY-MM-DD"，這裡去掉分隔符轉成 "YYYYMMDD"——跟 Snapshot 黃底輸入格
        # 要求使用者打的格式一致，MATCH 才對得起來，不用在公式裡另外做轉換。
        end_date_row = chart_date_row + 1
        ws.cell(row=end_date_row, column=1, value=t("compare.xls.period_end"))
        for col, end_date in enumerate(end_dates, start=2):
            # 沒有任何公司抓到那一期（G6 補出來的空白欄）就整格留空——不編一個
            # 不存在的結算日出來，那會讓 Snapshot 把假日期當成真的期間
            ws.cell(row=end_date_row, column=col,
                    value=end_date.replace("-", "") or None)

        # 公司資料列
        data_start = end_date_row + 1
        for offset, company in enumerate(all_companies):
            r = data_start + offset
            ws.cell(row=r, column=1, value=company)
            company_data = metric_data.get(company, {})
            for col, period in enumerate(periods, start=2):
                value = company_data.get(period)
                cell = ws.cell(row=r, column=col, value=value)
                if isinstance(value, (int, float)):
                    cell.value = value / divisor
                    cell.number_format = fmt
        data_end = data_start + len(all_companies) - 1

        block_ranges[metric_name] = (data_start, data_end)
        row = data_end + 1 + _BLOCK_GAP

    return block_ranges


def write_snapshot_sheets(
    wb: Workbook,
    result: ComparisonResult,
    metric_names: list[str],
    block_ranges: dict[str, tuple[int, int]],
    default_date: str = "",
) -> None:
    """寫 Snapshot（活公式）與 Snapshot_Manual（空白，供人工貼值）。

    Snapshot 對 Compare_Data 每個指標區塊的「期末結算日」列（每個區塊的資料
    起始列往上一列，見 write_compare_data_sheet 的排版）取值，改 B1 的日期
    Excel 會自動重算——這是刻意選擇用真公式，不是寫死算好的值，因為這份只給
    人在 Excel 裡看，沒有下游腳本要讀它（讀 Snapshot_Manual）。

    B1 是**真正的 Excel 日期**（不是文字），使用者可以打任何日期（如
    2024/7/15），不需要剛好對到某一期的期末結算日——公式抓的是「不晚於這天
    的最近一期」，符合分析師「這個時間點看得到的最新數字是什麼」的直覺
    （不能用未來才公布的數字回填過去的時間點）。2026-08-21 CTH 回報：原本
    B1 是純文字要求剛好打中 `YYYYMMDD`，使用者打數字會被 Excel 自動轉成
    數值型別，跟 Compare_Data 期末結算日列的文字型別對不上，MATCH 直接
    抓不到值；順便換掉這個 exact-match 限制。
    """
    all_companies = sorted({
        company
        for metric_data in result.metrics.values()
        for company in metric_data
    })
    data_ws = wb["Compare_Data"]

    snap = wb.create_sheet("Snapshot")
    snap["A1"] = t("compare.xls.timepoint")
    if default_date:
        try:
            snap["B1"] = datetime.datetime.strptime(default_date, "%Y%m%d").date()
        except ValueError:
            snap["B1"] = default_date  # 給不了合法日期就照原樣寫，不讓整個產檔失敗
    snap["B1"].number_format = "yyyy/mm/dd"
    snap["B1"].fill = _YELLOW_FILL

    # CTH 回報過黃格不知道要填什麼——A1 標籤只寫「時間點」看不出格式，旁邊
    # 補一句話講清楚格式，再把 Compare_Data 裡實際存在的期末結算日列出來，
    # 使用者不用自己去翻 Compare_Data 猜有哪些日期可以填。
    available_dates = sorted({
        end.replace("-", "")
        for company_map in result.period_ends.values()
        for end in company_map.values()
        if end
    })
    snap.cell(row=1, column=3,
              value=t("compare.xls.snapshot_format_hint"))
    snap.cell(row=1, column=4,
              value=t("compare.xls.snapshot_available_dates",
                      dates=t("compare.xls.notes.list_sep").join(available_dates)))

    header_row = 2
    snap.cell(row=header_row, column=1, value=t("compare.xls.company"))
    for col, metric_name in enumerate(metric_names, start=2):
        snap.cell(row=header_row, column=col, value=metric_name)

    for r_offset, company in enumerate(all_companies):
        r = header_row + 1 + r_offset
        snap.cell(row=r, column=1, value=company)
        for col, metric_name in enumerate(metric_names, start=2):
            if metric_name not in block_ranges:
                continue
            data_start, data_end = block_ranges[metric_name]
            end_date_row = data_start - 1   # 期末結算日列緊接在資料列上方

            company_row = None
            for rr in range(data_start, data_end + 1):
                if data_ws.cell(row=rr, column=1).value == company:
                    company_row = rr
                    break
            if company_row is None:
                continue

            last_col = data_ws.max_column
            end_date_range = (
                f"Compare_Data!$B${end_date_row}:${get_column_letter(last_col)}${end_date_row}"
            )
            data_range = (
                f"Compare_Data!$B${company_row}:${get_column_letter(last_col)}${company_row}"
            )
            # 「不晚於 B1 的最近一期」：期末結算日是固定 8 碼、左補零的
            # YYYYMMDD 文字，字串比較順序跟數值順序一致，可以直接用 <=。
            # SUMPRODUCT(MAX(...)) 取代 MATCH(...,0) 的精確比對——對每一格算
            # 「這家公司這期有資料 且 不晚於目標日」則給它的欄位序號、否則
            # 給 0，取最大值就是最近（最晚但不超過目標日）那一格的位置。
            #
            # 注意：「有沒有資料」要看 data_range（這家公司自己那一列），
            # 不能看 end_date_range（期末結算日列）——後者是**所有公司期間
            # 標籤的聯集**，A 公司有 Q2、B 公司沒有時，聯集的日期列在 Q2
            # 那格仍然有值，若拿它判斷空白會誤判「B 這期有資料」，實際
            # INDEX 到 B 真正空白的儲存格，Excel 把空格當 0 處理，算出錯誤
            # 的 0（這裡是實測 Excel COM 抓出來的真的會發生，不是理論推測）。
            #
            # 目標日期比任何一期都早時 MAX 會是 0——不能只靠 IFERROR 接住，
            # 因為 INDEX(range,0) 在 Excel 不會報錯，而是回傳整個範圍的隱含
            # 第一格（同樣是實測發現，會誤顯示成 Q1 的值），要用 IF 明確擋
            # 掉 0 的情況才會顯示空白。
            offset_expr = (
                f'SUMPRODUCT(MAX(({data_range}<>"")'
                f'*({end_date_range}<=TEXT($B$1,"yyyymmdd"))'
                f'*(COLUMN({end_date_range})-COLUMN(INDEX({end_date_range},1,1))+1)))'
            )
            formula = f'=IFERROR(IF({offset_expr}=0,"",INDEX({data_range},{offset_expr})),"")'
            snap.cell(row=r, column=col, value=formula)

    # Snapshot_Manual：同樣的表頭，資料格留空供人工貼值
    manual = wb.create_sheet("Snapshot_Manual")
    manual.cell(row=1, column=1, value=t("compare.xls.company"))
    for col, metric_name in enumerate(metric_names, start=2):
        manual.cell(row=1, column=col, value=metric_name)
    for r_offset, company in enumerate(all_companies):
        manual.cell(row=2 + r_offset, column=1, value=company)


def _chart_sheet_name(metric_name: str) -> str:
    """Chart_<指標> 但要塞進 Excel 的 31 字元 sheet 名稱上限。"""
    prefix = "Chart_"
    max_metric_len = 31 - len(prefix)
    safe_name = "".join(ch for ch in metric_name if ch not in '[]:*?/\\')
    return prefix + safe_name[:max_metric_len]


def _pin_title_layout(title) -> None:
    """明講標題「用自動版面、不要疊在別的東西上面」。

    2026-08-22 CTH 截圖回報 Y 軸標題壓在「50,000.0」刻度數字上、X 軸標題
    「期間」掉進日期標籤那一排裡面。跟前一輪圖例那個 bug 同一類：openpyxl 的
    `Title` 沒有 `overlay` / `layout` 屬性時**完全不寫這兩個元素**，Excel 拿到
    「沒寫」的標題會直接畫在既有內容上面，而不是另外撥一條專屬空間給它。
    原生 Excel 輸出的每個標題一定帶 `<c:layout/>`（空元素＝明講用自動版面）
    加 `<c:overlay val="0"/>`。圖表標題與兩個軸標題都要補。
    """
    if title is None:
        return
    title.overlay = False
    title.layout = Layout()


def write_chart_sheets(
    wb: Workbook, metric_names: list[str], block_ranges: dict[str, tuple[int, int]]
) -> None:
    """每個指標各一張 sheet，只放一張折線圖（歷史趨勢，一條線一家公司）。
    使用者要看長條圖版本，在 Excel 裡對圖表右鍵「變更圖表類型」自己切，
    這裡不用同一指標產兩份圖表物件。"""
    data_ws = wb["Compare_Data"]

    for metric_name in metric_names:
        if metric_name not in block_ranges:
            continue
        data_start, data_end = block_ranges[metric_name]
        # chart_date_row 緊接在 end_date_row（Snapshot 用的文字列）上方一列，
        # 排版順序見 write_compare_data_sheet()：header → chart_date_row →
        # end_date_row → data_start。
        chart_date_row = data_start - 2
        last_col = data_ws.max_column

        chart = LineChart()
        # F6（2026-09-03，方案 B）：X 軸從文字類別軸換成真正的日期軸——原本
        # `<c:catAx>` 的類別標籤是等距排列的文字（"20200126"），COSTCO 16 週
        # 第四季跟一般 13 週季在軸上畫出來一樣寬，時間軸是失真的。DateAxis
        # 繼承自 TextAxis（同一組 delete/axPos/tickLblPos/crosses/tickLblSkip
        # 屬性都通用），這裡直接整個換掉、下面沿用原本的設定方式。
        chart.x_axis = DateAxis()
        chart.title = metric_name
        chart.style = 2
        chart.x_axis.title = t("gui.compare.period")
        # 2026-08-22 CTH 截圖回報「中間斷線＋圖例被吃＋沒有單位」，實測抓出
        # 真正根因：openpyxl 產生的 <c:catAx>／<c:valAx> 完全沒有寫
        # <c:delete> 元素。拿 Excel COM 原生建立的圖表（Shapes/ChartObjects
        # 直接建，不經 openpyxl）比對 XML，發現原生輸出**一定**帶
        # `<c:delete val="0"/>`。OOXML 規格上 delete 沒寫預設就是 false，
        # 但 Excel 實際渲染時對「沒寫」跟「明講 false」待遇不同——沒寫會
        # 保守地不畫刻度標籤，連帶把圖例／座標軸空間計算搞壞、擠壓變形。
        # 這是用 PowerShell 呼叫 Excel COM 實測、比對原生輸出 XML 才抓到的，
        # 純看 openpyxl 文件或程式碼猜不到。兩軸都要明講 delete=False。
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        # openpyxl 兩個軸預設都是 axPos="l"（見 openpyxl.chart.axis._BaseAxis
        # 的預設值）——對 Y 軸（valAx）剛好是對的，對 X 軸（catAx）是錯的，
        # 該在底部卻被標成左側。原生 Excel 輸出兩軸位置都明講，這裡跟進。
        chart.x_axis.axPos = "b"
        chart.y_axis.axPos = "l"
        # 同理明講兩軸的刻度標籤要顯示、跟軸線交叉方式，都是原生 Excel
        # 輸出必有的欄位，openpyxl 預設不會寫，一起補齊不留模糊地帶。
        chart.x_axis.tickLblPos = "nextTo"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.crosses = "autoZero"
        chart.y_axis.crosses = "autoZero"
        # baseTimeUnit 明講「以天為底」——不寫的話 Excel 對缺值/不規則間距的
        # 資料容易自己判斷成月或年為底，跟同一套「openpyxl 不寫、Excel 就當
        # 不確定狀態」的坑同一類，這裡照 TODO F6 設計書明講掉。majorUnit 留
        # 給 Excel 自動決定（跨公司比較的時間跨度差很多，寫死一個值對短區間
        # 太密、對長區間又太疏）。
        chart.x_axis.baseTimeUnit = "days"
        chart.x_axis.numFmt = "yyyy-mm-dd"
        # 不同公司財年結束月不同，財季標籤（FY2024Q3）字串排序無法反映真實時間，
        # 缺值期間 Excel 預設會直接連到下一個有值的點造成誤導折線，兩者一起處理：
        # X 軸類別改用期末結算日（絕對日期）取代財季標籤，缺值處顯示為斷點不連線
        chart.display_blanks = "gap"

        # F3（2026-08-20 CTH 截圖回報 5 項，2026-08-21 確認最終方案）：
        # 尺寸拉一倍（openpyxl 預設 15cm×7.5cm 偏小）、圖例移到下方橫排
        # （右側直排在公司數一多時會被擠進繪圖區，下方橫排可隨公司數換行）
        chart.width = 30
        chart.height = 15
        chart.legend.position = "b"
        # 同一類「openpyxl 沒寫、Excel 就當不確定狀態」的坑：legend 沒有
        # overlay 屬性時，圖例會跟 X 軸標題/刻度標籤擠在同一條窄帶、疊在
        # 一起看不清楚（實測畫面：圖例文字直接蓋在 X 軸日期上）。原生 Excel
        # 輸出一定帶 `overlay="0"`（不要疊加，另外保留專屬空間），這裡明講。
        chart.legend.overlay = False

        # Y 軸數字格式跟 Compare_Data 儲存格同一套規則（不要另外定義一套會
        # 漂移的格式）；金額類指標軸標題帶單位，百分比類指標格式本身已經
        # 看得出是 %，標題維持指標名稱就好，不重複講
        fmt, _ = unit_format_for(metric_name)
        chart.y_axis.numFmt = fmt
        chart.y_axis.title = f"{metric_name} ($mm)" if fmt == FMT_FINANCIAL else metric_name

        # 三個標題都要明講 layout/overlay，理由見 _pin_title_layout()。
        # 一定要放在三個 title 都設完之後——openpyxl 每次指派 title 都會
        # 重新造一個 Title 物件，先設屬性會被後面的指派蓋掉。
        _pin_title_layout(chart.title)
        _pin_title_layout(chart.x_axis.title)
        _pin_title_layout(chart.y_axis.title)

        data_ref = Reference(
            data_ws, min_col=1, max_col=last_col, min_row=data_start, max_row=data_end
        )
        chart.add_data(data_ref, titles_from_data=True, from_rows=True)

        # 指到 chart_date_row（真日期數值），不是 end_date_row（Snapshot 用的
        # 文字列）。openpyxl 的 set_categories() 一律寫成 <cat><numRef>，這次
        # 指向的儲存格本來就是數值型別，不需要再手動換成 strRef 那套 workaround
        # （2026-08-22 那次是因為指到文字儲存格才需要換，這裡已經不是那個情況）。
        categories_ref = Reference(
            data_ws, min_col=2, max_col=last_col, min_row=chart_date_row, max_row=chart_date_row
        )
        chart.set_categories(categories_ref)

        # 接上 D0-1 Q4 合成後跨公司比較的時間跨度可以拉到 60-70 欄，全部日期
        # 標籤硬擠在 X 軸上會疊字看不清楚（這是修復帶出的新問題，原本 F3 5
        # 項清單沒有涵蓋）。目標大約 15 個可視標籤，跳著顯示。
        n_periods = last_col - 1  # 扣掉 A 欄（公司名），B 欄開始才是期間
        tick_skip = max(1, n_periods // 15)
        chart.x_axis.tickLblSkip = tick_skip
        chart.x_axis.tickMarkSkip = tick_skip

        sheet_name = _chart_sheet_name(metric_name)
        chart_ws = wb.create_sheet(sheet_name)
        chart_ws.add_chart(chart, "B2")


def write_comparison_workbook(
    result: ComparisonResult,
    metric_names: list[str],
    output_path: Path,
    snapshot_date: str = "",
) -> None:
    """組出完整跨公司比較 Excel 並存檔。"""
    wb = Workbook()
    block_ranges = write_compare_data_sheet(wb, result, metric_names)
    write_notes_sheet(wb, result, metric_names)
    write_snapshot_sheets(wb, result, metric_names, block_ranges, default_date=snapshot_date)
    write_chart_sheets(wb, metric_names, block_ranges)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
