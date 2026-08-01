"""
excel_formatter.py — Apply professional formatting to Data_* sheets and generate Index sheet.

Public API:
    format_workbook(wb, tables) -> None

Called by excel_writer.write_statements() before wb.save(). Modifies cell values
(÷1M unit conversion) and applies openpyxl styles. Does not change sheet structure.
"""

from __future__ import annotations
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fetcher_gaap import StatementTable
from datetime import date
from override_engine import check_key_rows
import metric_rules

# ── Colours (ARGB) ────────────────────────────────────────────────────────
NAVY_DARK = "FF1F3864"
NAVY_MID  = "FF2D4A82"
BLUE_MID  = "FF2E75B6"
GREY_SEP  = "FFEEEEEE"
ROW_ALT   = "FFF5F8FF"
ROW_WHITE = "FFFFFFFF"
BLUE_HDR  = "FFDDE8F5"

# ── Row classification ────────────────────────────────────────────────────
SECTION_HEADERS = {"Income Statement", "Balance Sheet", "Cash Flow"}

SUBTOTAL_CONCEPTS = {
    "Gross Profit", "Total Operating Expense", "Operating Income",
    "Pre-tax Income", "Net Income",
    "Total Current Assets", "Total Assets",
    "Total Current Liabilities", "Total Liabilities", "Total Equity",
    "Operating Cash Flow", "Free Cash Flow",
}

SHEET_DESCRIPTIONS = {
    "Data_Financials(Q)": "季報三表合一（IS + BS + CF，from 10-Q）",
    "Data_Financials(Y)": "年報三表合一（IS + BS + CF，from 10-K）",
    "Data_EPS_Recon":     "Non-GAAP EPS 調節表（from 8-K）",
    "Data_NonGAAP":       "Non-GAAP 指標（AI 提取）",
    "Data_Ratios":        "常見財務比率（自 Data_Financials(Q) 計算，B 欄為算法）",
    "Data_Meta":          "申報資訊（Ticker、公司名、抓取日期）",
}


# ── 數值分類 ──────────────────────────────────────────────────────────────
#
# 關鍵字表在 metric_rules.py（唯一可調整處）。判斷順序：每股 → 百分比 → 股數 → 金額，
# 先命中先套用。順序有意義：「Non-GAAP 每股盈餘」同時含「每股」與（若加了寬鬆詞）
# 可能的百分比詞，每股必須優先。
#
# 決策（2026-08-01，使用者選定）：百分比存成 **Excel 原生百分比**——37.5 存為
# 0.375，格式 "0.0%"，儲存格顯示 37.5%。好處是在 Excel 裡拉公式、畫圖、算加權
# 平均都直接可用，不必手動 ÷100。
# 改回存原始數字（37.5 + '#,##0.0"%"'）把這個開關設 False 即可，兩種都有測試覆蓋。
PERCENT_AS_EXCEL_RATIO = True


def _keyword_matcher(keywords: tuple[str, ...]):
    """把關鍵字表編成一條 regex。

    ASCII 關鍵字要求詞界，中日韓字元用裸比對（中文沒有詞界的概念）。
    詞界是必要的，不是潔癖：'Operations' 含 'ratio'、'Corporate' 含 'rate'、
    'Steps' 含 'eps'，裸子字串比對會讓 XBRL overflow 的金額行不再除以 1M。
    """
    parts = []
    for kw in keywords:
        esc = re.escape(kw.casefold())
        if re.search(r"[a-z0-9]", kw.casefold()):
            parts.append(rf"(?<![a-z0-9]){esc}(?![a-z0-9])")
        else:
            parts.append(esc)
    pattern = re.compile("|".join(parts))
    return lambda c: bool(pattern.search((c or "").casefold()))


_is_eps_concept    = _keyword_matcher(metric_rules.EPS_KEYWORDS)
_is_percent_concept = _keyword_matcher(metric_rules.PERCENT_KEYWORDS)
_is_shares_concept = _keyword_matcher(metric_rules.SHARES_KEYWORDS)


def _sheet_description(name: str) -> str:
    if name in SHEET_DESCRIPTIONS:
        return SHEET_DESCRIPTIONS[name]
    if name.startswith("Data_Seg_"):
        return f"Segment 細項：{name[9:]}"
    return name


def _fill(hex_argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_argb)


# ── Column widths ─────────────────────────────────────────────────────────

def _apply_column_widths(ws) -> None:
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    for col in range(3, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 13


# ── Freeze panes ──────────────────────────────────────────────────────────

def _set_freeze_panes(ws) -> None:
    ws.freeze_panes = "C3"


# ── Number formatting and unit conversion ────────────────────────────────

FMT_FINANCIAL = "#,##0.0_ ;[Red](#,##0.0)"
FMT_EPS       = "#,##0.00_ ;[Red](#,##0.00)"
FMT_SHARES    = "#,##0"
FMT_PERCENT   = "0.0%" if PERCENT_AS_EXCEL_RATIO else '#,##0.0"%"'
FMT_MULTIPLE  = '#,##0.00"x"'
FMT_DAYS      = '#,##0.0"d"'

# Data_Ratios 的列名一定帶單位後綴，格式照後綴走、且一律不 ÷1,000,000。
# 後綴判斷必須**優先於**關鍵字判斷：「流動比率 (x)」含「率」會被關鍵字當成
# 百分比而 ÷100，「DSO (days)」不含任何關鍵字會被當金額 ÷1,000,000。
_UNIT_SUFFIX_FORMATS = {
    "(%)":     (FMT_PERCENT,  100 if PERCENT_AS_EXCEL_RATIO else 1),
    "(x)":     (FMT_MULTIPLE, 1),
    "(days)":  (FMT_DAYS,     1),
    "($)":     (FMT_EPS,      1),
}


def _unit_suffix_rule(concept: str) -> tuple[str, int] | None:
    """列名帶單位後綴時回 (格式, 除數)，否則回 None 交給關鍵字判斷。"""
    for suffix, rule in _UNIT_SUFFIX_FORMATS.items():
        if concept.endswith(suffix):
            return rule
    return None


def _apply_row_styles(ws) -> None:
    """Apply fill and font styles to all rows."""
    white_font  = Font(color="FFFFFFFF", bold=True, size=11)
    small_font  = Font(color="FFAABBCC", size=9)

    # Row 1: ticker / quarter labels — dark navy
    for cell in ws[1]:
        cell.fill = _fill(NAVY_DARK)
        cell.font = white_font

    # Row 2: filing dates — medium navy
    for cell in ws[2]:
        cell.fill = _fill(NAVY_MID)
        cell.font = small_font

    # Row 3+: classify by col A value
    for row_idx in range(3, ws.max_row + 1):
        concept = ws.cell(row=row_idx, column=1).value or ""
        concept = str(concept).strip()

        if concept in SECTION_HEADERS:
            row_fill  = _fill(BLUE_MID)
            row_font  = Font(color="FFFFFFFF", bold=True, size=10)
            row_height = 16
        elif concept == "":
            row_fill  = _fill(GREY_SEP)
            row_font  = Font(size=9)
            row_height = 6
        else:
            row_fill  = _fill(ROW_WHITE) if row_idx % 2 == 0 else _fill(ROW_ALT)
            bold      = concept in SUBTOTAL_CONCEPTS
            row_font  = Font(bold=bold) if bold else Font()
            row_height = None

        for cell in ws[row_idx]:
            cell.fill = row_fill
            cell.font = row_font
        if row_height is not None:
            ws.row_dimensions[row_idx].height = row_height


def _apply_number_formats(ws) -> None:
    """Convert values to millions and apply number formats. Skips section/blank rows."""
    for row_idx in range(3, ws.max_row + 1):
        concept = str(ws.cell(row=row_idx, column=1).value or "").strip()

        # Section headers and blank separators have no numeric data
        if concept in SECTION_HEADERS or concept == "":
            continue

        # 單位後綴優先於一切關鍵字判斷（Data_Ratios 用）
        suffix_rule = _unit_suffix_rule(concept)
        if suffix_rule is not None:
            fmt, divisor = suffix_rule
        # 順序即優先級：每股 → 百分比 → 股數 → 金額（見 metric_rules.py 第 5 節）
        elif _is_eps_concept(concept):
            fmt = FMT_EPS
            divisor = 1
        elif _is_percent_concept(concept):
            fmt = FMT_PERCENT
            divisor = 100 if PERCENT_AS_EXCEL_RATIO else 1
        elif _is_shares_concept(concept):
            fmt = FMT_SHARES
            divisor = 1_000_000
        else:
            fmt = FMT_FINANCIAL
            divisor = 1_000_000

        for col_idx in range(3, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value / divisor
                cell.number_format = fmt


# ── Quality check colours ─────────────────────────────────────────────────
QUALITY_GREEN   = "FF1A7A34"
QUALITY_ORANGE  = "FFC25C00"
QUALITY_MISS_BG = "FFFFF0E0"
QUALITY_MISS_FG = "FFC00000"

ALL_KEY_ROWS = [
    "Revenue", "Operating Income", "Net Income", "Diluted EPS",
    "Total Assets", "Total Liabilities", "Total Equity — Parent",
    "Operating Cash Flow", "Capex",
]


def _compute_quality(tables: list) -> tuple[int, int, set] | None:
    """Compute quality score for Data_Financials(Q).

    Returns (score, total, missing_set) or None if no Q table found.
    """
    q_tbl = next((t for t in tables if t.sheet_name == "Data_Financials(Q)"), None)
    if q_tbl is None:
        return None
    missing = set(
        check_key_rows(q_tbl.concepts, q_tbl.values, "IS") +
        check_key_rows(q_tbl.concepts, q_tbl.values, "BS") +
        check_key_rows(q_tbl.concepts, q_tbl.values, "CF")
    )
    total = len(ALL_KEY_ROWS)
    return total - len(missing), total, missing


# ── Index sheet ───────────────────────────────────────────────────────────

def _build_index_sheet(wb: Workbook, tables: list) -> None:
    """Insert or replace the Index sheet at position 0."""
    if "Index" in wb.sheetnames:
        del wb["Index"]

    ticker       = tables[0].ticker if tables else ""
    company_name = ""
    meta = next((t for t in tables if t.sheet_name == "Data_Meta"), None)
    if meta and len(meta.concepts) > 1 and len(meta.values) > 1 and meta.values[1]:
        company_name = meta.values[1][0] or ""

    header_text = f"{ticker} — {company_name}" if company_name else ticker

    quality = _compute_quality(tables)

    ws = wb.create_sheet("Index", 0)

    # Row 1: company header
    ws["A1"] = header_text
    ws["A1"].fill = _fill(NAVY_DARK)
    ws["A1"].font = Font(color="FFFFFFFF", bold=True, size=14)
    ws.merge_cells("A1:E1")

    # Row 2: metadata
    ws["A2"] = f"抓取日期：{date.today()}　　資料來源：SEC EDGAR"
    ws["A2"].fill = _fill(NAVY_MID)
    ws["A2"].font = Font(color="FFAABBCC", size=9)
    ws.merge_cells("A2:E2")

    # Row 3: blank
    ws.row_dimensions[3].height = 6

    # Row 4: column headers
    hdr_font = Font(bold=True, size=10)
    hdr_fill = _fill(BLUE_HDR)
    for col, label in enumerate(["Sheet", "說明", "最早期間", "最新期間", "完成度"], start=1):
        cell = ws.cell(row=4, column=col, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill

    # Row 5+: one row per Data_* sheet
    data_sheets = [t for t in tables if t.sheet_name.startswith("Data_")]
    for i, tbl in enumerate(data_sheets):
        row = 5 + i
        earliest = tbl.quarter_labels[0]  if tbl.quarter_labels else "—"
        latest   = tbl.quarter_labels[-1] if tbl.quarter_labels else "—"

        is_primary = tbl.sheet_name in ("Data_Financials(Q)", "Data_Financials(Y)")
        name_font  = Font(color="FF1F3864" if is_primary else "FF666666",
                          bold=is_primary, size=10)
        row_fill   = _fill(ROW_WHITE) if row % 2 == 0 else _fill(ROW_ALT)

        for col, val in enumerate([tbl.sheet_name,
                                    _sheet_description(tbl.sheet_name),
                                    earliest, latest], start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            if col == 1:
                cell.font = name_font

        # ── E 欄：完成度 ──
        e_cell = ws.cell(row=row, column=5)
        e_cell.fill = row_fill
        if tbl.sheet_name == "Data_Financials(Q)" and quality is not None:
            score, total, _ = quality
            if score == total:
                e_cell.value = f"{score}/{total} ✓"
                e_cell.font = Font(color=QUALITY_GREEN, size=10)
            else:
                e_cell.value = f"{score}/{total} ⚠"
                e_cell.font = Font(color=QUALITY_ORANGE, size=10)
        else:
            e_cell.value = "—"
            e_cell.font = Font(color="FF999999", size=10)

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10

    # ── 品質明細區塊 ───────────────────────────────────────────────────────
    if quality is None:
        return

    _, _, missing = quality
    next_row = 5 + len(data_sheets) + 2   # blank row gap

    # Section header
    hdr_cell = ws.cell(row=next_row, column=1, value="品質明細 — Data_Financials(Q)")
    hdr_cell.fill = _fill(NAVY_DARK)
    hdr_cell.font = Font(color="FFFFFFFF", bold=True, size=10)
    ws.merge_cells(
        start_row=next_row, start_column=1,
        end_row=next_row, end_column=2
    )
    next_row += 1

    for row_name in ALL_KEY_ROWS:
        is_missing = row_name in missing
        bg = _fill(QUALITY_MISS_BG) if is_missing else _fill(ROW_WHITE)
        status = "✗  缺失" if is_missing else "✓"
        fg = QUALITY_MISS_FG if is_missing else QUALITY_GREEN

        a = ws.cell(row=next_row, column=1, value=row_name)
        b = ws.cell(row=next_row, column=2, value=status)
        a.fill = bg
        b.fill = bg
        a.font = Font(color=fg, size=10)
        b.font = Font(color=fg, size=10)
        next_row += 1


# ── Public API ────────────────────────────────────────────────────────────

def format_workbook(wb: Workbook, tables: list[StatementTable]) -> None:
    """Apply formatting to all Data_* sheets."""
    _build_index_sheet(wb, tables)
    for ws in wb.worksheets:
        if not ws.title.startswith("Data_"):
            continue
        _apply_column_widths(ws)
        _set_freeze_panes(ws)
        _apply_row_styles(ws)
        if ws.title != "Data_Meta":
            _apply_number_formats(ws)
