"""fiscal_input.py — 讓使用者自己改財年起始月，其他期間標籤用公式帶。

**為什麼要做這個**：財年結束月是程式從 10-K 的 XBRL 欄名偵測出來的
（`fetcher_gaap._detect_fy_end_month()`）。偵測錯了，`Data_Financials` 的財季
標籤就整排錯，而使用者除了重跑程式沒有別的辦法，看到的還只是一堆寫死的文字。

改法：Index 上放**一格可以改的財年起始月**，其他地方全部用 Excel 公式從它
推算。程式猜錯時，使用者改那一格，整本活頁簿的財季立刻跟著更新。

    Index!B4        使用者輸入（黃底），定義名稱 FY_START_MONTH
    Data_*  第 5 列  期末結算日 —— XBRL 的真實日期，**公式的錨，永遠是靜態值**
    Data_*  第 1 列  期間標籤   ── 公式
    Data_*  第 3 列  財季       ── 公式
    Data_*  第 4 列  日曆季     ── 公式（只看期末日，與財年無關）

## 內縮 15 天

所有換算都先把期末日往前推 15 天再取年月。美股多用 52/53 週制，期末日會在
月底前後浮動最多 6 天——WDC 的 FY2026 Q2 結束在 **2026-01-02**，直接看月份會
算成 Q3，整整差一季。`docs/8k-period-off-by-one.md` 量化過這個誤差：COST /
WDC / PANW 有 7 份就是栽在這裡。往前推 15 天一定會落回該季最後一個月。

## 沒有連動的地方（刻意的）

`Data_Ratios` 與 `Data_Meta` 是 Python 算好寫死的，改起始月不會重算。比率的
YoY 是按欄位相對位置取的，不受標籤文字影響；`Data_Meta` 的「財年起訖」會與
使用者改過的值不一致——這是已知取捨，全部改成公式的複雜度不划算。Index 上的
說明有寫。
"""
from __future__ import annotations

import re
from datetime import date, timedelta

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# 使用者輸入格與它的定義名稱。公式一律引用定義名稱而不是 Index!$B$4——
# 日後 Index 版面調整，公式不必跟著改。
FY_START_CELL = "B4"
FY_START_DEFINED_NAME = "FY_START_MONTH"

# 期間表頭在各 Data_* sheet 的固定列號
ROW_PERIOD_LABEL = 1
ROW_FISCAL_QUARTER = 3
ROW_CALENDAR_QUARTER = 4
ROW_PERIOD_END = 5

_DATA_START_COL = 4

# 52/53 週制的期末日浮動不會超過 ±6 天，往前推 15 天必定落在該季最後一個月。
_SHRINK_DAYS = 15

_ISO_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")

_INPUT_FILL = PatternFill("solid", fgColor="FFFFF2CC")   # 淡黃＝可編輯
_INPUT_BORDER = Border(*[Side(style="thin", color="FFBF8F00")] * 4)


# ── Python 參考實作 ─────────────────────────────────────────────────────────
#
# 這幾個函式是 Excel 公式的規格：同一套邏輯寫兩次（Python 一次、公式一次），
# 測試釘住 Python 這份，改公式時對照著改。

def fy_start_month(fy_end_month: int) -> int:
    """財年結束月 → 起始月。AAPL 9 月結束 → 10 月開始。"""
    return fy_end_month % 12 + 1


def _anchor(period_end: str | None) -> date | None:
    """期末日字串 → 內縮 15 天後的日期。不是 ISO 日期回 None。"""
    m = _ISO_RE.match((period_end or "").strip())
    if m is None:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3))) - timedelta(days=_SHRINK_DAYS)
    except ValueError:
        return None


def _fiscal_year(d: date, start_month: int) -> int:
    """財年 N 結束於西元 N 年（SEC 慣例），所以起始月之後的月份屬於下一個財年。"""
    return d.year + 1 if start_month > 1 and d.month >= start_month else d.year


def fiscal_quarter_of(period_end: str | None, start_month: int) -> str:
    """期末日 + 財年起始月 → `FY2026Q2`。算不出來回空字串。"""
    d = _anchor(period_end)
    if d is None:
        return ""
    quarter = (d.month - start_month) % 12 // 3 + 1
    return f"FY{_fiscal_year(d, start_month)}Q{quarter}"


def fiscal_year_of(period_end: str | None, start_month: int) -> str:
    """年報用：期末日 + 財年起始月 → `FY2025`。"""
    d = _anchor(period_end)
    return f"FY{_fiscal_year(d, start_month)}" if d else ""


def calendar_quarter_of(period_end: str | None) -> str:
    """期末日 → 日曆季 `2026Q2`。與財年無關。"""
    d = _anchor(period_end)
    return f"{d.year}Q{(d.month - 1) // 3 + 1}" if d else ""


# ── Excel 公式 ──────────────────────────────────────────────────────────────

def _date_expr(col: str) -> str:
    """第 5 列的 ISO 日期文字 → 內縮 15 天的 Excel 日期序列值。

    不用 `DATEVALUE`：它認不認得 `2026-04-26` 要看使用者的地區設定，
    自己拆 LEFT/MID 再組 `DATE()` 在哪台電腦都一樣。
    """
    c = f"{col}{ROW_PERIOD_END}"
    return f"DATE(VALUE(LEFT({c},4)),VALUE(MID({c},6,2)),VALUE(MID({c},9,2)))-{_SHRINK_DAYS}"


def _fiscal_year_expr(d: str) -> str:
    n = FY_START_DEFINED_NAME
    return f"(YEAR({d})+IF(AND({n}>1,MONTH({d})>={n}),1,0))"


def _quarter_expr(d: str) -> str:
    return f"(INT(MOD(MONTH({d})-{FY_START_DEFINED_NAME},12)/3)+1)"


def _guard(col: str, body: str) -> str:
    """沒有期末日就留空，不要讓使用者看到 #VALUE!。"""
    return f'=IF({col}{ROW_PERIOD_END}="","",{body})'


def period_label_formula(col: str, annual: bool = False) -> str:
    """第 1 列的期間標籤：季報 `FY2026Q2`、年報 `FY2025`。"""
    d = _date_expr(col)
    body = f'"FY"&{_fiscal_year_expr(d)}'
    if not annual:
        body += f'&"Q"&{_quarter_expr(d)}'
    return _guard(col, body)


def fiscal_quarter_formula(col: str) -> str:
    """第 3 列的財季 `FY2026FQ2`。

    財季用 `FQ`、日曆季用純數字，視覺上就分得開——非 12 月結算的公司同一欄
    可能是 FY2026FQ1 但日曆 2025Q4，看錯就是整整一季。
    """
    d = _date_expr(col)
    return _guard(col, f'"FY"&{_fiscal_year_expr(d)}&"FQ"&{_quarter_expr(d)}')


def calendar_quarter_formula(col: str) -> str:
    """第 4 列的日曆季 `2026Q2`。只看期末日，不引用財年起始月。"""
    d = _date_expr(col)
    return _guard(col, f'YEAR({d})&"Q"&(INT((MONTH({d})-1)/3)+1)')


# ── 套用 ────────────────────────────────────────────────────────────────────

_NOTE = ("⚠ 請核對：財年起始月是程式從 10-K 自動判讀的，可能出錯。把上面 B4 改成"
         "正確的月份（1-12），Data_Financials(Q)/(Y) 第 1、3、4 列的期間標籤會自動更新。"
         "核對方法：看 Data_Financials 第 5 列的期末結算日（來自 XBRL，一定正確），"
         "對照公司財報上寫的財季。"
         "　※ 財季是 3 個月一段，所以 B4 改 1~2 個月常常看不出變化（例如 2、3、4 月"
         "開始，4 月底結束的那季都算 Q1），這是正確的不是沒生效——右邊的財年區間"
         "會即時反映你改的月份。"
         "　※ 本頁表格的「最早/最新期間」、Data_Ratios、Data_Meta 是程式算好的靜態值，"
         "改這格不會跟著變。")

# 財年區間：唯一「改 1 個月就會變」的即時回饋。沒有它，使用者把 2 改成 3
# 看到標籤沒動，會以為公式壞了（2026-08-08 CTH 實際回報）。
# DATE(2000, m+11, 1) 讓月份自己進位，不必寫 MOD。
_FY_SPAN_FORMULA = (
    f'=IF({FY_START_DEFINED_NAME}="","","財年 "&TEXT(DATE(2000,{FY_START_DEFINED_NAME},1),"m")'
    f'&" 月 – "&TEXT(DATE(2000,{FY_START_DEFINED_NAME}+11,1),"m")&" 月")'
)


def _write_input_block(ws, start_month: int, row: int = 4) -> None:
    """在 Index 寫出可編輯的輸入格與提醒。"""
    label = ws.cell(row=row, column=1, value="財年起始月（可修改）")
    label.font = Font(bold=True, size=10)

    cell = ws.cell(row=row, column=2, value=start_month)
    cell.fill = _INPUT_FILL
    cell.border = _INPUT_BORDER
    cell.font = Font(bold=True, size=11, color="FFBF8F00")
    cell.alignment = Alignment(horizontal="center")
    cell.number_format = "0"

    span = ws.cell(row=row, column=3, value=_FY_SPAN_FORMULA)
    span.font = Font(size=10, color="FF666666")

    note = ws.cell(row=row + 1, column=1, value=_NOTE)
    note.font = Font(size=9, color="FFBF8F00")
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=5)
    ws.row_dimensions[row + 1].height = 28


def _is_annual(ws) -> bool:
    """年報 sheet 的期間標籤沒有季別（`FY2025` 而不是 `FY2025Q1`）。"""
    for col in range(_DATA_START_COL, ws.max_column + 1):
        value = str(ws.cell(ROW_PERIOD_LABEL, col).value or "")
        if value:
            return not re.search(r"Q[1-4]$", value)
    return False


def _apply_to_sheet(ws) -> None:
    annual = _is_annual(ws)
    for col_idx in range(_DATA_START_COL, ws.max_column + 1):
        period_end = str(ws.cell(ROW_PERIOD_END, col_idx).value or "")
        if not _ISO_RE.match(period_end.strip()):
            # 舊申報沒帶期末日，公式沒有錨可用——保留原本寫死的標籤。
            continue
        col = get_column_letter(col_idx)
        ws.cell(ROW_PERIOD_LABEL, col_idx).value = period_label_formula(col, annual)
        ws.cell(ROW_CALENDAR_QUARTER, col_idx).value = calendar_quarter_formula(col)
        if not annual:
            ws.cell(ROW_FISCAL_QUARTER, col_idx).value = fiscal_quarter_formula(col)


def apply_fiscal_year_input(wb, fy_end_month: int) -> None:
    """在 Index 放輸入格，並把 Data_Financials 的期間表頭改成公式。

    沒有 Index 或沒有 Data_Financials 都安靜跳過——欄位不全的活頁簿
    （只抓年報、只抓 segment）也要能正常寫出來。
    """
    if "Index" not in wb.sheetnames:
        return

    start_month = fy_start_month(fy_end_month or 12)
    _write_input_block(wb["Index"], start_month)

    ref = f"'Index'!${FY_START_CELL[0]}${FY_START_CELL[1:]}"
    if FY_START_DEFINED_NAME in wb.defined_names:
        del wb.defined_names[FY_START_DEFINED_NAME]
    wb.defined_names.add(DefinedName(FY_START_DEFINED_NAME, attr_text=ref))

    for name in ("Data_Financials(Q)", "Data_Financials(Y)"):
        if name in wb.sheetnames:
            _apply_to_sheet(wb[name])

    # openpyxl 不算公式，寫出去的儲存格沒有快取值。不強制重算的話，Excel 有機會
    # 直接顯示空白（看起來像整排標籤不見了）。
    wb.calculation.fullCalcOnLoad = True
