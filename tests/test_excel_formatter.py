"""Tests for excel_formatter.py."""
import pytest
from openpyxl import Workbook
from fetcher_gaap import StatementTable
from excel_formatter import format_workbook, FMT_FINANCIAL, FMT_EPS, FMT_SHARES, FMT_PERCENT, _compute_quality, ALL_KEY_ROWS as _ALL_KEY_ROWS, QUALITY_GREEN, QUALITY_ORANGE, QUALITY_MISS_BG


def _make_wb(sheet_name="Data_Financials(Q)"):
    """Minimal workbook with one Data_* sheet and two data columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = "AAPL"
    ws["C1"] = "FY2023Q1"
    ws["D1"] = "FY2023Q2"
    ws["C2"] = "2023-02-03"
    ws["D2"] = "2023-05-05"
    ws["A3"] = "Income Statement"   # section header
    ws["A4"] = "Revenue"
    ws["B4"] = "Revenues"
    ws["C4"] = 117154000000.0
    ws["D4"] = 94836000000.0
    ws["A5"] = ""                   # blank separator
    ws["A6"] = "Basic EPS"
    ws["C6"] = 1.52
    ws["D6"] = 1.20
    ws["A7"] = "Basic Shares"
    ws["C7"] = 15787000000.0
    ws["D7"] = 15813000000.0
    return wb


# ── column widths ──────────────────────────────────────────────────────────

def test_col_a_width():
    wb = _make_wb()
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"].column_dimensions["A"].width == 22

def test_col_b_width():
    wb = _make_wb()
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"].column_dimensions["B"].width == 24

def test_data_col_width():
    wb = _make_wb()
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"].column_dimensions["C"].width == 13

def test_data_col_d_width():
    wb = _make_wb()
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"].column_dimensions["D"].width == 13


# ── freeze panes ───────────────────────────────────────────────────────────

def test_freeze_panes():
    wb = _make_wb()
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"].freeze_panes == "C3"

def test_freeze_panes_seg_sheet():
    wb = _make_wb(sheet_name="Data_Seg_Revenue")
    format_workbook(wb, [])
    assert wb["Data_Seg_Revenue"].freeze_panes == "C3"


# ── row styles ─────────────────────────────────────────────────────────────

def _rgb(ws, cell_ref: str) -> str:
    """Return fgColor ARGB string of a cell's fill."""
    return ws[cell_ref].fill.fgColor.rgb


def test_row1_fill_navy_dark():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    assert _rgb(ws, "A1") == "FF1F3864"

def test_row1_font_bold_white():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.color.rgb == "FFFFFFFF"

def test_row2_fill_navy_mid():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    assert _rgb(ws, "A2") == "FF2D4A82"

def test_section_header_fill_blue_mid():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    # A3 = "Income Statement"
    assert _rgb(ws, "A3") == "FF2E75B6"

def test_section_header_font_bold_white():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    assert ws["A3"].font.bold is True
    assert ws["A3"].font.color.rgb == "FFFFFFFF"

def test_blank_separator_fill_grey():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    # A5 = ""
    assert _rgb(ws, "A5") == "FFEEEEEE"

def test_section_header_row_height():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    assert ws.row_dimensions[3].height == 16

def test_blank_separator_row_height():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    assert ws.row_dimensions[5].height == 6

def test_data_row_alternating_white():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    # row_idx=4, even → ROW_WHITE
    assert _rgb(ws, "A4") == "FFFFFFFF"

def test_data_row_alternating_blue():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    # row_idx=7, odd → ROW_ALT
    assert _rgb(ws, "A7") == "FFF5F8FF"

def test_subtotal_row_bold():
    wb = _make_wb()
    # Add a Gross Profit row
    wb["Data_Financials(Q)"]["A8"] = "Gross Profit"
    wb["Data_Financials(Q)"]["C8"] = 5000000000.0
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    assert ws["A8"].font.bold is True

def test_normal_row_not_bold():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    # A4 = "Revenue" (not a subtotal)
    assert ws["A4"].font.bold is not True


# ── number formatting + unit conversion ────────────────────────────────────

def test_revenue_divided_by_million():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    # C4 = "Revenue" row, raw = 117154000000
    assert ws["C4"].value == pytest.approx(117154.0)

def test_revenue_second_col_divided():
    wb = _make_wb()
    format_workbook(wb, [])
    ws = wb["Data_Financials(Q)"]
    assert ws["D4"].value == pytest.approx(94836.0)

def test_revenue_number_format():
    wb = _make_wb()
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"]["C4"].number_format == FMT_FINANCIAL

def test_eps_not_divided():
    wb = _make_wb()
    format_workbook(wb, [])
    # C6 = "Basic EPS" = 1.52 → must stay 1.52
    assert wb["Data_Financials(Q)"]["C6"].value == pytest.approx(1.52)

def test_eps_number_format():
    wb = _make_wb()
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"]["C6"].number_format == FMT_EPS

def test_shares_divided_by_million():
    wb = _make_wb()
    format_workbook(wb, [])
    # C7 = "Basic Shares" = 15787000000 → 15787.0
    assert wb["Data_Financials(Q)"]["C7"].value == pytest.approx(15787.0)

def test_shares_number_format():
    wb = _make_wb()
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"]["C7"].number_format == FMT_SHARES

def test_section_header_values_unchanged():
    wb = _make_wb()
    format_workbook(wb, [])
    # C3 = "Income Statement" row — all None
    assert wb["Data_Financials(Q)"]["C3"].value is None

def test_data_meta_values_not_converted():
    """Data_Meta contains strings — must not be divided."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data_Meta"
    ws["A1"] = "AAPL"
    ws["C1"] = "FY2023Q1"
    ws["A3"] = "Ticker"
    ws["C3"] = "AAPL"
    format_workbook(wb, [])
    assert wb["Data_Meta"]["C3"].value == "AAPL"

def test_seg_sheet_financial_converted():
    wb = _make_wb(sheet_name="Data_Seg_Revenue")
    wb["Data_Seg_Revenue"]["A4"] = "Americas"
    wb["Data_Seg_Revenue"]["C4"] = 50000000000.0
    format_workbook(wb, [])
    assert wb["Data_Seg_Revenue"]["C4"].value == pytest.approx(50000.0)


# ── Index sheet ────────────────────────────────────────────────────────────

def _make_tables(sheet_name="Data_Financials(Q)", ticker="AAPL",
                 qs=None, dates=None):
    qs    = qs    or ["FY2020Q1", "FY2024Q4"]
    dates = dates or ["2020-04-30", "2025-01-30"]
    return [StatementTable(
        sheet_name=sheet_name, ticker=ticker,
        quarter_labels=qs, filing_dates=dates,
        concepts=["Revenue"], values=[[100.0, 200.0]], labels=["Revenues"],
    )]


def test_index_sheet_created():
    wb = _make_wb()
    format_workbook(wb, _make_tables())
    assert "Index" in wb.sheetnames

def test_index_sheet_is_first():
    wb = _make_wb()
    format_workbook(wb, _make_tables())
    assert wb.sheetnames[0] == "Index"

def test_index_sheet_ticker_in_a1():
    wb = _make_wb()
    format_workbook(wb, _make_tables(ticker="AAPL"))
    ws = wb["Index"]
    assert "AAPL" in str(ws["A1"].value)

def test_index_lists_data_sheet():
    wb = _make_wb()
    format_workbook(wb, _make_tables())
    ws = wb["Index"]
    col_a_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Data_Financials(Q)" in col_a_values

def test_index_shows_earliest_period():
    wb = _make_wb()
    format_workbook(wb, _make_tables(qs=["FY2010Q1", "FY2024Q4"]))
    ws = wb["Index"]
    all_values = [ws.cell(row=r, column=c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, 5)]
    assert "FY2010Q1" in all_values

def test_index_shows_latest_period():
    wb = _make_wb()
    format_workbook(wb, _make_tables(qs=["FY2010Q1", "FY2024Q4"]))
    ws = wb["Index"]
    all_values = [ws.cell(row=r, column=c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, 5)]
    assert "FY2024Q4" in all_values

def test_index_not_deleted_on_reformat():
    """Index must not be deleted by excel_writer (doesn't start with Data_)."""
    wb = _make_wb()
    format_workbook(wb, _make_tables())
    # Simulate a second write: excel_writer deletes Data_* sheets
    for name in list(wb.sheetnames):
        if name.startswith("Data_"):
            del wb[name]
    assert "Index" in wb.sheetnames


# ── _compute_quality ──────────────────────────────────────────────────────

def _make_q_table(missing=None):
    """StatementTable with all 9 key rows; rows in `missing` have all-None values."""
    missing = set(missing or [])
    values = [
        [None, None, None, None] if c in missing else [100.0, 200.0, 300.0, 400.0]
        for c in _ALL_KEY_ROWS
    ]
    return StatementTable(
        sheet_name="Data_Financials(Q)",
        quarter_labels=["FY2025Q1", "FY2025Q2", "FY2025Q3", "FY2025Q4"],
        filing_dates=["2025-02-01", "2025-05-01", "2025-08-01", "2025-11-01"],
        concepts=_ALL_KEY_ROWS[:],
        values=values,
        ticker="TEST",
    )


def test_compute_quality_all_ok():
    tbl = _make_q_table()
    score, total, missing = _compute_quality([tbl])
    assert total == 9
    assert score == 9
    assert missing == set()


def test_compute_quality_two_missing():
    tbl = _make_q_table(missing=["Operating Income", "Capex"])
    score, total, missing = _compute_quality([tbl])
    assert score == 7
    assert missing == {"Operating Income", "Capex"}


def test_compute_quality_no_q_table():
    result = _compute_quality([])
    assert result is None


def test_compute_quality_non_q_table_ignored():
    tbl = _make_q_table()
    tbl.sheet_name = "Data_Financials(Y)"
    result = _compute_quality([tbl])
    assert result is None


# ── Index quality column ──────────────────────────────────────────────────

def _index_ws(tables=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data_Financials(Q)"
    format_workbook(wb, tables or [])
    return wb["Index"]


def test_index_quality_header_exists():
    ws = _index_ws()
    assert ws.cell(row=4, column=5).value == "完成度"


def test_index_quality_col_all_ok():
    tbl = _make_q_table()
    ws = _index_ws([tbl])
    cell = ws.cell(row=5, column=5)
    assert "9/9" in str(cell.value)
    assert "✓" in str(cell.value)
    assert cell.font.color.rgb == QUALITY_GREEN


def test_index_quality_col_missing():
    tbl = _make_q_table(missing=["Operating Income", "Capex"])
    ws = _index_ws([tbl])
    cell = ws.cell(row=5, column=5)
    assert "7/9" in str(cell.value)
    assert "⚠" in str(cell.value)
    assert cell.font.color.rgb == QUALITY_ORANGE


def test_index_quality_col_no_q_table():
    ws = _index_ws([])
    assert ws.cell(row=4, column=5).value == "完成度"


def test_index_header_merged_to_e():
    ws = _index_ws()
    merged = [str(r) for r in ws.merged_cells.ranges]
    assert any("E1" in r for r in merged), f"A1:E1 not merged, got: {merged}"


# ── Index quality detail section ──────────────────────────────────────────

def _find_detail_header_row(ws) -> int | None:
    """Find the row containing the quality detail section header."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "品質明細" in str(cell.value):
                return cell.row
    return None


def test_index_detail_section_present():
    tbl = _make_q_table()
    ws = _index_ws([tbl])
    assert _find_detail_header_row(ws) is not None, "品質明細 section header not found"


def test_index_detail_section_absent_when_no_q_table():
    ws = _index_ws([])
    assert _find_detail_header_row(ws) is None


def test_index_detail_all_ok_shows_check():
    tbl = _make_q_table()
    ws = _index_ws([tbl])
    hdr_row = _find_detail_header_row(ws)
    b_vals = [ws.cell(row=hdr_row + 1 + i, column=2).value for i in range(9)]
    assert all(v == "✓" for v in b_vals), f"Expected all ✓, got: {b_vals}"


def test_index_detail_missing_row_shows_cross():
    tbl = _make_q_table(missing=["Operating Income", "Capex"])
    ws = _index_ws([tbl])
    hdr_row = _find_detail_header_row(ws)
    rows = {
        ws.cell(row=hdr_row + 1 + i, column=1).value:
        ws.cell(row=hdr_row + 1 + i, column=2).value
        for i in range(9)
    }
    assert "✗" in rows.get("Operating Income", ""), f"Operating Income: {rows.get('Operating Income')}"
    assert "✗" in rows.get("Capex", ""), f"Capex: {rows.get('Capex')}"


def test_index_detail_missing_row_highlighted():
    tbl = _make_q_table(missing=["Capex"])
    ws = _index_ws([tbl])
    hdr_row = _find_detail_header_row(ws)
    for i in range(9):
        a_val = ws.cell(row=hdr_row + 1 + i, column=1).value
        if a_val == "Capex":
            fill_rgb = ws.cell(row=hdr_row + 1 + i, column=1).fill.fgColor.rgb
            assert fill_rgb == QUALITY_MISS_BG, f"Expected orange bg, got: {fill_rgb}"
            return
    pytest.fail("Capex row not found in detail section")


# ═════════════════════════════════════════════════════════════════════════════
# Data_NonGAAP 數值分類（2026-08-01 新增，TODO 第 2 項）
#
# 這張 sheet 的值是 AI 直接從 8-K 新聞稿抓的原始數字：
#   金額 = 絕對數（30400000）、百分比 = 原始小數（20.2）、每股 = 原始小數（0.28）
# 修前只有 EPS / Per Share / per share 三個英文關鍵字能豁免 ÷1M，
# 導致毛利率 37.5 → 3.75e-05、EPS 0.10 → 1e-07。
# 分類關鍵字表在 metric_rules.py。
# ═════════════════════════════════════════════════════════════════════════════

def _make_nongaap_wb():
    """Data_NonGAAP 沒有 B 欄 label，結構與 Data_Financials 不同。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data_NonGAAP"
    ws["A1"] = "ARLO"
    ws["C1"] = "FY2026Q1"
    ws["D1"] = "FY2026Q2"
    ws["C2"] = "2026-02-01"
    ws["D2"] = "2026-05-01"
    rows = [
        ("Non-GAAP Gross Margin",        47.8,       50.1),
        ("Non-GAAP 毛利率",              47.8,       50.1),   # 舊快取殘留的中文名
        ("Non-GAAP Diluted EPS",         0.22,       0.28),
        ("Non-GAAP 每股盈餘",            0.22,       0.28),
        ("Adjusted EBITDA",              23300000.0, 30400000.0),
        ("Adjusted EBITDA Margin",       16.5,       20.2),
        ("Free Cash Flow",               66900000.0, 25400000.0),
        ("自由現金流",                    66900000.0, 25400000.0),
        ("Non-GAAP Effective Tax Rate",  17.0,       17.5),
    ]
    for i, (name, v1, v2) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=3, value=v1)
        ws.cell(row=i, column=4, value=v2)
    return wb


def _cell(name, col="C"):
    """跑完 format_workbook 後，取指定指標那一列的儲存格。"""
    wb = _make_nongaap_wb()
    format_workbook(wb, [])
    ws = wb["Data_NonGAAP"]
    for row_idx in range(3, ws.max_row + 1):
        if str(ws.cell(row=row_idx, column=1).value or "").strip() == name:
            return ws[f"{col}{row_idx}"]
    raise AssertionError(f"找不到指標列：{name}")


# ── 百分比：不除以 1M，套百分比格式 ──────────────────────────────────────────

def test_nongaap_percent_stored_as_excel_ratio():
    """毛利率 47.8 → 0.478（Excel 原生百分比，搭配 0.0% 格式顯示成 47.8%）。
    絕不可變成 4.78e-05——那是被誤當金額除以 1M 的舊 bug。
    要改回存原始數字 47.8，把 excel_formatter.PERCENT_AS_EXCEL_RATIO 設 False。"""
    assert _cell("Non-GAAP Gross Margin").value == pytest.approx(0.478)

def test_nongaap_percent_number_format():
    assert _cell("Non-GAAP Gross Margin").number_format == FMT_PERCENT

def test_nongaap_zh_percent_value_not_divided():
    """中文指標名（舊快取殘留）同樣要走百分比分支，不可被當金額除以 1M。"""
    assert _cell("Non-GAAP 毛利率").value == pytest.approx(0.478)

def test_nongaap_margin_keyword_percent():
    """Adjusted EBITDA Margin 是百分比，不是金額。"""
    assert _cell("Adjusted EBITDA Margin").value == pytest.approx(0.165)

def test_nongaap_tax_rate_is_percent():
    """Rate 結尾也是百分比。"""
    assert _cell("Non-GAAP Effective Tax Rate").number_format == FMT_PERCENT


# ── 每股：不除以 1M，兩位小數 ────────────────────────────────────────────────

def test_nongaap_eps_value_not_divided():
    assert _cell("Non-GAAP Diluted EPS").value == pytest.approx(0.22)

def test_nongaap_zh_eps_value_not_divided():
    """「每股」中文關鍵字要認得——修前 0.10 被除成 1e-07。"""
    assert _cell("Non-GAAP 每股盈餘").value == pytest.approx(0.22)

def test_nongaap_eps_number_format():
    assert _cell("Non-GAAP Diluted EPS").number_format == FMT_EPS


# ── 金額：仍要除以 1M（防過度豁免）───────────────────────────────────────────

def test_nongaap_amount_still_divided():
    """Adjusted EBITDA 23.3M → 23.3。加了百分比／中文關鍵字後最容易誤傷這類。"""
    assert _cell("Adjusted EBITDA").value == pytest.approx(23.3)

def test_nongaap_amount_number_format():
    assert _cell("Adjusted EBITDA").number_format == FMT_FINANCIAL

def test_nongaap_zh_amount_still_divided():
    """中文金額名也要照除——「自由現金流」不含率／每股字樣。"""
    assert _cell("自由現金流").value == pytest.approx(66.9)

def test_nongaap_fcf_divided():
    assert _cell("Free Cash Flow", col="D").value == pytest.approx(25.4)


# ── GAAP 三表不可被影響 ──────────────────────────────────────────────────────

def test_gaap_revenue_still_divided_after_nongaap_rules():
    """新分類規則不得動到 Data_Financials 的行為。"""
    wb = _make_wb()
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"]["C4"].value == pytest.approx(117154.0)

def test_gaap_basic_shares_still_shares_format():
    wb = _make_wb()
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"]["C7"].number_format == FMT_SHARES


# ── 百分比關鍵字不可用裸子字串比對（實查 fetcher_gaap 模板後補的迴歸測試）────
#
# "Operations" 含 "ratio"、"Corporate" 含 "rate"。若關鍵字用裸 `in` 比對，
# XBRL overflow 行（label 常含 Operations）會被誤判成百分比而**不再除以 1M**，
# 三表金額直接錯 6 個數量級。ASCII 關鍵字一律要求詞界。

def _make_wb_with_concept(name, value):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data_Financials(Q)"
    ws["A1"] = "AAPL"
    ws["C1"] = "FY2023Q1"
    ws["A3"] = name
    ws["C3"] = value
    return wb

def test_operations_row_still_divided():
    """'Income from Discontinued Operations' 含 'ratio'，但它是金額。"""
    wb = _make_wb_with_concept("Income from Discontinued Operations", 5000000.0)
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"]["C3"].value == pytest.approx(5.0)

def test_corporate_row_still_divided():
    """'Corporate Expense' 含 'rate'，但它是金額。"""
    wb = _make_wb_with_concept("Corporate Expense", 8000000.0)
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"]["C3"].value == pytest.approx(8.0)

def test_real_tax_rate_still_percent():
    """真的以 Rate 結尾的才算百分比（17.0 → 0.17）。"""
    wb = _make_wb_with_concept("Non-GAAP Effective Tax Rate", 17.0)
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"]["C3"].value == pytest.approx(0.17)

def test_steps_not_eps():
    """'eps' 不可命中 'Steps' 之類的字（同一個裸子字串問題）。"""
    wb = _make_wb_with_concept("Restructuring Steps Charge", 3000000.0)
    format_workbook(wb, [])
    assert wb["Data_Financials(Q)"]["C3"].value == pytest.approx(3.0)


def test_percent_format_matches_storage_mode():
    """存法與格式必須一致——存 0.478 卻套 '#,##0.0"%"' 會顯示成 0.5%。"""
    from excel_formatter import PERCENT_AS_EXCEL_RATIO
    if PERCENT_AS_EXCEL_RATIO:
        assert FMT_PERCENT == "0.0%"
    else:
        assert "%" in FMT_PERCENT and FMT_PERCENT != "0.0%"
