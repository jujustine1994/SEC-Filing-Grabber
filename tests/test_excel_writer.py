import pytest
import openpyxl
from pathlib import Path
from fetcher_gaap import StatementTable
from excel_writer import write_statements


def test_a1_shows_ticker_when_set(tmp_path):
    tbl = StatementTable(
        sheet_name="Data_IS",
        quarter_labels=["FY2023Q1"],
        filing_dates=["2023-02-03"],
        concepts=["Revenue"],
        values=[[1000.0]],
        ticker="AAPL",
    )
    out = tmp_path / "AAPL.xlsx"
    write_statements([tbl], out)
    wb = openpyxl.load_workbook(out)
    assert wb["Data_IS"]["A1"].value == "AAPL"


def test_a1_is_none_when_ticker_empty(tmp_path):
    tbl = StatementTable(
        sheet_name="Data_IS",
        quarter_labels=["FY2023Q1"],
        filing_dates=["2023-02-03"],
        concepts=["Revenue"],
        values=[[1000.0]],
        ticker="",
    )
    out = tmp_path / "AAPL.xlsx"
    write_statements([tbl], out)
    wb = openpyxl.load_workbook(out)
    assert wb["Data_IS"]["A1"].value is None


@pytest.fixture
def sample_tables():
    return [
        StatementTable(
            sheet_name="Data_IS",
            quarter_labels=["FY2023Q1", "FY2023Q2", "FY2023Q3"],
            filing_dates=["2023-02-03", "2023-05-05", "2023-08-04"],
            concepts=["Revenues", "NetIncomeLoss", "EPS"],
            values=[
                [1000.0, 1100.0, 1200.0],
                [200.0,  220.0,  240.0],
                [1.23,   1.35,   1.47],
            ],
            labels=["Total net revenues", "Net income", ""],
        ),
        StatementTable(
            sheet_name="Data_BS",
            quarter_labels=["FY2023Q1", "FY2023Q2"],
            filing_dates=["2023-02-03", "2023-05-05"],
            concepts=["Assets", "Liabilities"],
            values=[[50000.0, 52000.0], [30000.0, 31000.0]],
        ),
    ]


def test_write_creates_file(tmp_path, sample_tables):
    out = tmp_path / "AAPL.xlsx"
    write_statements(sample_tables, out)
    assert out.exists()


def test_write_creates_correct_sheets(tmp_path, sample_tables):
    out = tmp_path / "AAPL.xlsx"
    write_statements(sample_tables, out)
    wb = openpyxl.load_workbook(out)
    assert "Data_IS" in wb.sheetnames
    assert "Data_BS" in wb.sheetnames


def test_col_a_is_concept_name(tmp_path, sample_tables):
    out = tmp_path / "AAPL.xlsx"
    write_statements(sample_tables, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Data_IS"]
    # Col A: row 1 and 2 are empty/ticker; row 3+ = concept names
    assert ws["A3"].value == "Revenues"
    assert ws["A4"].value == "NetIncomeLoss"
    assert ws["A5"].value == "EPS"


def test_col_b_is_original_item(tmp_path, sample_tables):
    out = tmp_path / "AAPL.xlsx"
    write_statements(sample_tables, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Data_IS"]
    assert ws["C3"].value == "Total net revenues"
    assert ws["C4"].value == "Net income"
    assert ws["C5"].value is None   # empty string stored as None in Excel cells


def test_col_b_is_none_when_no_labels(tmp_path):
    """Sheets without a labels list should write None to col B."""
    tbl = StatementTable(
        sheet_name="Data_BS",
        quarter_labels=["FY2023Q1"],
        filing_dates=["2023-02-03"],
        concepts=["Assets"],
        values=[[50000.0]],
    )
    out = tmp_path / "test.xlsx"
    write_statements([tbl], out)
    wb = openpyxl.load_workbook(out)
    assert wb["Data_BS"]["C3"].value is None


def test_row1_is_quarter_labels(tmp_path, sample_tables):
    out = tmp_path / "AAPL.xlsx"
    write_statements(sample_tables, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Data_IS"]
    # Quarter labels now start at col C (index 3)
    assert ws["D1"].value == "FY2023Q1"
    assert ws["E1"].value == "FY2023Q2"
    assert ws["F1"].value == "FY2023Q3"


def test_row1_b_is_empty(tmp_path, sample_tables):
    """B1 must be empty — it is the Original Item column header area."""
    out = tmp_path / "AAPL.xlsx"
    write_statements(sample_tables, out)
    wb = openpyxl.load_workbook(out)
    assert wb["Data_IS"]["C1"].value is None


def test_row2_is_filing_dates(tmp_path, sample_tables):
    out = tmp_path / "AAPL.xlsx"
    write_statements(sample_tables, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Data_IS"]
    assert ws["D2"].value == "2023-02-03"
    assert ws["E2"].value == "2023-05-05"


def test_data_values_correct(tmp_path, sample_tables):
    out = tmp_path / "AAPL.xlsx"
    write_statements(sample_tables, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Data_IS"]
    # Values are divided by 1M during formatting: 1000.0 → 0.001
    assert ws["D3"].value == pytest.approx(1000.0 / 1_000_000)
    assert ws["E3"].value == pytest.approx(1100.0 / 1_000_000)
    assert ws["F3"].value == pytest.approx(1200.0 / 1_000_000)


def test_preserves_non_data_sheets(tmp_path, sample_tables):
    """Python must NOT touch any sheet that doesn't start with Data_."""
    out = tmp_path / "AAPL.xlsx"
    wb = openpyxl.Workbook()
    ws_user = wb.create_sheet("My_IS")
    ws_user["A1"] = "User annotation"
    ws_user["C1"] = "=Data_IS!C3"
    wb.save(out)
    wb.close()

    write_statements(sample_tables, out)

    wb2 = openpyxl.load_workbook(out)
    assert "My_IS" in wb2.sheetnames
    assert wb2["My_IS"]["A1"].value == "User annotation"


def test_rewrite_replaces_old_data(tmp_path, sample_tables):
    """Second write must replace all Data_* content (handles restatements)."""
    out = tmp_path / "AAPL.xlsx"
    write_statements(sample_tables, out)

    updated = [
        StatementTable(
            sheet_name="Data_IS",
            quarter_labels=["FY2023Q1", "FY2023Q2", "FY2023Q3", "FY2023Q4"],
            filing_dates=["2023-02-03", "2023-05-05", "2023-08-04", "2023-11-03"],
            concepts=["Revenues", "NetIncomeLoss", "EPS"],
            values=[
                [1000.0, 1100.0, 1200.0, 1300.0],
                [200.0,  220.0,  240.0,  260.0],
                [1.23,   1.35,   1.47,   1.60],
            ],
        )
    ]
    write_statements(updated, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Data_IS"]
    assert ws["G1"].value == "FY2023Q4"   # 第 4 季在 G 欄（A/B/C 為名稱欄）
    assert "Data_BS" not in wb.sheetnames  # not in updated list → removed


# ═════════════════════════════════════════════════════════════════════════════
# 檔案覆蓋防護（TODO 第 8 項，2026-08-01）
#
# 兩個風險：
#   (a) xlsx 正被 Excel 開著 → Windows 鎖檔 → wb.save() 拋 PermissionError，
#       但這發生在抓取與 AI 呼叫全部跑完的最後一步，白等一分多鐘
#   (b) 整批替換 Data_* 沒有任何備份，第二次抓的年份較窄時舊季度直接消失
# ═════════════════════════════════════════════════════════════════════════════

import pytest
from openpyxl import Workbook, load_workbook
from fetcher_gaap import StatementTable
from excel_writer import write_statements, check_output_writable


def _tbl(name="Data_Financials(Q)", quarters=("FY2024Q1",), value=1.0):
    return StatementTable(
        sheet_name=name,
        quarter_labels=list(quarters),
        filing_dates=["2024-01-01"] * len(quarters),
        concepts=["Revenue"],
        values=[[value] * len(quarters)],
        ticker="TEST",
        labels=[""],
    )


# ── 可寫入預檢 ─────────────────────────────────────────────────────────────

def test_check_writable_ok_for_new_file(tmp_path):
    """檔案不存在時可寫入，回 None。"""
    assert check_output_writable(tmp_path / "new.xlsx") is None


def test_check_writable_ok_for_existing_unlocked_file(tmp_path):
    p = tmp_path / "existing.xlsx"
    Workbook().save(p)
    assert check_output_writable(p) is None


def test_check_writable_detects_locked_file(tmp_path):
    """被別的程式開著（Windows 鎖檔）時要回錯誤訊息，不可回 None。"""
    p = tmp_path / "locked.xlsx"
    Workbook().save(p)
    handle = open(p, "r+b")
    try:
        import os
        if os.name != "nt":
            pytest.skip("檔案鎖定行為僅在 Windows 可測")
        try:
            import msvcrt
            msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
        except OSError:
            pytest.skip("無法取得檔案鎖")
        msg = check_output_writable(p)
        assert msg is not None
        assert "Excel" in msg or "開啟" in msg
    finally:
        handle.close()


def test_check_writable_reports_missing_parent_dir_as_ok(tmp_path):
    """父資料夾不存在不算錯——write_statements 會自己建。"""
    assert check_output_writable(tmp_path / "sub" / "deep" / "x.xlsx") is None


# ── 原子寫入 ───────────────────────────────────────────────────────────────

def test_write_leaves_no_temp_file_behind(tmp_path):
    p = tmp_path / "out.xlsx"
    write_statements([_tbl()], p)
    leftovers = [f.name for f in tmp_path.iterdir() if f.name != "out.xlsx"
                 and not f.name.endswith(".bak.xlsx")]
    assert leftovers == []


def test_existing_file_intact_when_save_fails(tmp_path, monkeypatch):
    """寫到一半失敗時，原檔必須完好——不可留下半個檔。"""
    p = tmp_path / "out.xlsx"
    write_statements([_tbl(value=111.0)], p)

    import excel_writer as ew
    real_save = Workbook.save
    def boom(self, path):
        raise OSError("disk full")
    monkeypatch.setattr(Workbook, "save", boom)

    with pytest.raises(OSError):
        write_statements([_tbl(value=222.0)], p)

    monkeypatch.setattr(Workbook, "save", real_save)
    ws = load_workbook(p)["Data_Financials(Q)"]
    assert ws["D3"].value == pytest.approx(111.0 / 1_000_000)


# ── 覆蓋前備份 ─────────────────────────────────────────────────────────────

def test_backup_created_when_overwriting(tmp_path):
    """覆蓋既有檔前先留一份 .bak.xlsx——年份範圍變窄時舊季度才救得回來。"""
    p = tmp_path / "out.xlsx"
    write_statements([_tbl(quarters=("FY2024Q1", "FY2024Q2"))], p)
    write_statements([_tbl(quarters=("FY2024Q2",))], p)

    bak = tmp_path / "out.bak.xlsx"
    assert bak.exists()
    ws = load_workbook(bak)["Data_Financials(Q)"]
    assert ws["D1"].value == "FY2024Q1"      # 備份裡還有被覆蓋掉的舊季


def test_no_backup_for_new_file(tmp_path):
    """第一次寫入沒有東西可備份，不可產生空的 .bak。"""
    p = tmp_path / "out.xlsx"
    write_statements([_tbl()], p)
    assert not (tmp_path / "out.bak.xlsx").exists()


def test_backup_is_single_rolling_copy(tmp_path):
    """只保留一份備份，不可每次都堆一個新檔。"""
    p = tmp_path / "out.xlsx"
    for i in range(3):
        write_statements([_tbl(value=float(i))], p)
    baks = [f.name for f in tmp_path.iterdir() if ".bak" in f.name]
    assert len(baks) == 1


def test_custom_sheets_still_preserved(tmp_path):
    """既有行為不可壞：My_* 自訂 sheet 必須保留。"""
    p = tmp_path / "out.xlsx"
    wb = Workbook()
    wb.active.title = "My_IS"
    wb["My_IS"]["A1"] = "我的分析"
    wb.save(p)

    write_statements([_tbl()], p)

    wb2 = load_workbook(p)
    assert "My_IS" in wb2.sheetnames
    assert wb2["My_IS"]["A1"].value == "我的分析"


# ═════════════════════════════════════════════════════════════════════════════
# 四欄版面：A 英文名 / B 中文說明 / C 原始標籤 / D 起數據（2026-08-03）
# ═════════════════════════════════════════════════════════════════════════════

def test_data_starts_at_column_d():
    from excel_writer import _DATA_START_COL
    assert _DATA_START_COL == 4


def _four_col_wb(tmp_path):
    tbl = StatementTable(
        sheet_name="Data_Financials(Q)",
        quarter_labels=["FY2025Q1", "FY2025Q2"],
        filing_dates=["2025-05-01", "2025-08-01"],
        concepts=["Revenue", "Cash", "某公司特有科目"],
        values=[[100.0, 110.0], [50.0, 55.0], [7.0, 8.0]],
        ticker="TEST",
        labels=["Net sales", "Cash and cash equivalents", "us-gaap_Weird"],
    )
    p = tmp_path / "four.xlsx"
    write_statements([tbl], p)
    return load_workbook(p)["Data_Financials(Q)"]


def test_english_name_in_column_a(tmp_path):
    ws = _four_col_wb(tmp_path)
    assert ws["A3"].value == "Revenue"


def test_chinese_label_in_column_b(tmp_path):
    ws = _four_col_wb(tmp_path)
    assert ws["B3"].value == "營業收入"
    assert ws["B4"].value == "現金及約當現金"


def test_original_xbrl_label_in_column_c(tmp_path):
    ws = _four_col_wb(tmp_path)
    assert ws["C3"].value == "Net sales"


def test_unknown_concept_has_blank_chinese(tmp_path):
    """overflow 區的公司特有科目沒有中文，留白不可報錯。"""
    ws = _four_col_wb(tmp_path)
    assert ws["B5"].value in (None, "")
    assert ws["C5"].value == "us-gaap_Weird"


def test_values_start_at_column_d(tmp_path):
    ws = _four_col_wb(tmp_path)
    assert ws["D3"].value == pytest.approx(100.0 / 1_000_000)
    assert ws["E3"].value == pytest.approx(110.0 / 1_000_000)


def test_quarter_labels_start_at_column_d(tmp_path):
    ws = _four_col_wb(tmp_path)
    assert ws["D1"].value == "FY2025Q1"
    assert ws["E1"].value == "FY2025Q2"
