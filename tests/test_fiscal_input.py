"""fiscal_input.py — 使用者可改的財年起始月，以及由它驅動的期間公式。"""
import pytest
from openpyxl import Workbook

import fiscal_input as fi


# ── 財年起始月 ──────────────────────────────────────────────────────────────

@pytest.mark.parametrize("fy_end, expected", [
    (12, 1),    # PLTR：12 月結束 → 1 月開始
    (9, 10),    # AAPL
    (1, 2),     # NVDA
    (11, 12),   # AVGO
    (6, 7),     # MSFT
])
def test_fy_start_month(fy_end, expected):
    assert fi.fy_start_month(fy_end) == expected


# ── Python 參考實作（公式的規格，也是抓不到期末日時的退路）─────────────────

@pytest.mark.parametrize("period_end, fy_start, expected", [
    ("2026-04-26", 2, "FY2027Q1"),    # NVDA
    ("2026-01-25", 2, "FY2026Q4"),    # NVDA 財年最後一季
    ("2025-12-27", 10, "FY2026Q1"),   # AAPL
    ("2026-06-27", 10, "FY2026Q3"),   # AAPL
    ("2026-06-30", 1, "FY2026Q2"),    # PLTR
    ("2026-05-03", 12, "FY2026Q2"),   # AVGO
    ("2026-02-01", 12, "FY2026Q1"),   # AVGO
    ("2026-06-30", 7, "FY2026Q4"),    # MSFT 財年最後一季
])
def test_fiscal_quarter_of(period_end, fy_start, expected):
    assert fi.fiscal_quarter_of(period_end, fy_start) == expected


def test_period_ending_in_the_first_days_of_a_month_belongs_to_the_previous_one():
    """52/53 週制：WDC 的 FY2026 Q2 結束在 2026-01-02。

    直接看月份會算成 Q3（整整差一季）。這是調查報告裡量化過的地雷
    （docs/8k-period-off-by-one.md：COST/WDC/PANW 七份對不上都是這個）。
    """
    assert fi.fiscal_quarter_of("2026-01-02", 7) == "FY2026Q2"
    assert fi.calendar_quarter_of("2026-01-02", basis="end") == "2025Q4"


@pytest.mark.parametrize("period_end, expected", [
    ("2026-04-26", "2026Q2"),
    ("2025-12-27", "2025Q4"),
    ("2026-05-03", "2026Q2"),
    ("2026-02-01", "2026Q1"),
])
def test_calendar_quarter_of(period_end, expected):
    assert fi.calendar_quarter_of(period_end, basis="end") == expected


def test_fiscal_year_of_annual_period():
    assert fi.fiscal_year_of("2025-11-02", 12) == "FY2025"   # AVGO
    assert fi.fiscal_year_of("2025-09-27", 10) == "FY2025"   # AAPL
    assert fi.fiscal_year_of("2026-01-25", 2) == "FY2026"    # NVDA


def test_bad_period_end_returns_empty():
    assert fi.fiscal_quarter_of("", 1) == ""
    assert fi.fiscal_quarter_of("不是日期", 1) == ""
    assert fi.calendar_quarter_of(None, basis="end") == ""


# ── 公式字串 ────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("builder", [
    fi.period_label_formula,
    fi.fiscal_quarter_formula,
    fi.calendar_quarter_formula,
])
def test_formula_shape(builder):
    f = builder("D")
    assert f.startswith("=")
    assert f.count("(") == f.count(")"), f
    assert "D5" in f, "必須以第 5 列的真實期末日為錨"


@pytest.mark.parametrize("builder", [fi.period_label_formula, fi.fiscal_quarter_formula])
def test_formulas_reference_the_user_editable_cell(builder):
    assert fi.FY_START_DEFINED_NAME in builder("D")


def test_calendar_quarter_does_not_depend_on_fiscal_year():
    """日曆季只跟期末日有關，改財年起始月不該動到它。"""
    assert fi.FY_START_DEFINED_NAME not in fi.calendar_quarter_formula("D")


def test_annual_label_formula_has_no_quarter():
    assert "Q" not in fi.period_label_formula("D", annual=True).split("&")[-1]


# ── 套用到活頁簿 ────────────────────────────────────────────────────────────

def _workbook_with_headers(labels, period_ends, sheet="Data_Financials(Q)"):
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Index")
    ws = wb.create_sheet(sheet)
    ws.cell(1, 1, "TEST")
    ws.cell(3, 1, "財季 Fiscal Quarter")
    ws.cell(4, 1, "日曆季 Calendar Quarter")
    ws.cell(5, 1, "期末結算日 Period End")
    for i, (label, end) in enumerate(zip(labels, period_ends)):
        col = 4 + i
        ws.cell(1, col, label)
        ws.cell(5, col, end)
    return wb, ws


def test_apply_writes_formulas_into_the_header_rows():
    wb, ws = _workbook_with_headers(["FY2026Q1", "FY2026Q2"],
                                    ["2026-03-29", "2026-06-28"])
    fi.apply_fiscal_year_input(wb, fy_end_month=12)
    for col in ("D", "E"):
        assert str(ws[f"{col}1"].value).startswith("=")
        assert str(ws[f"{col}3"].value).startswith("=")
        assert str(ws[f"{col}4"].value).startswith("=")


def test_period_end_row_stays_a_plain_value():
    """第 5 列是 XBRL 的真實期末日，是整套公式的錨，不可以變成公式。"""
    wb, ws = _workbook_with_headers(["FY2026Q1"], ["2026-03-29"])
    fi.apply_fiscal_year_input(wb, fy_end_month=12)
    assert ws["D5"].value == "2026-03-29"


def test_column_without_period_end_keeps_its_static_label():
    """舊申報沒帶期末日時公式算不出來，維持原值不要變成 #VALUE!。"""
    wb, ws = _workbook_with_headers(["FY2019Q1", "FY2026Q2"], ["", "2026-06-28"])
    fi.apply_fiscal_year_input(wb, fy_end_month=12)
    assert ws["D1"].value == "FY2019Q1"
    assert str(ws["E1"].value).startswith("=")


def test_input_cell_holds_the_start_month_and_is_named():
    wb, _ = _workbook_with_headers(["FY2026Q1"], ["2026-03-29"])
    fi.apply_fiscal_year_input(wb, fy_end_month=9)      # AAPL
    assert wb["Index"][fi.FY_START_CELL].value == 10
    assert fi.FY_START_DEFINED_NAME in wb.defined_names


def test_index_carries_a_visible_instruction():
    wb, _ = _workbook_with_headers(["FY2026Q1"], ["2026-03-29"])
    fi.apply_fiscal_year_input(wb, fy_end_month=9)
    text = " ".join(str(c.value or "") for row in wb["Index"].iter_rows() for c in row)
    assert "財年起始月" in text
    assert "核對" in text, "要明講請使用者自己核對，不能只放一個數字"


def test_annual_sheet_gets_year_labels_not_quarter_labels():
    wb, ws = _workbook_with_headers(["FY2024", "FY2025"],
                                    ["2024-11-03", "2025-11-02"],
                                    sheet="Data_Financials(Y)")
    fi.apply_fiscal_year_input(wb, fy_end_month=11)
    assert "FQ" not in str(ws["D1"].value)
    assert str(ws["D4"].value).startswith("=")


def test_fiscal_span_gives_feedback_for_every_single_month_change():
    """財季是 3 個月一段，B4 改 1 個月常常看不出標籤變化（2、3、4 月開始的
    公司，4 月底結束那季都是 Q1）。沒有一個「改一格就會動」的東西，使用者
    會以為公式壞了——財年區間就是那個回饋。"""
    wb, _ = _workbook_with_headers(["FY2026Q1"], ["2026-03-29"])
    fi.apply_fiscal_year_input(wb, fy_end_month=1)
    span = str(wb["Index"]["C4"].value)
    assert span.startswith("=")
    assert span.count(fi.FY_START_DEFINED_NAME) == 3   # 條件 + 起月 + 迄月
    assert "財年" in span


def test_quarter_labels_are_stable_within_a_three_month_block():
    """釘住上面那條註解講的行為：2/3/4 月開始，4 月底結束的那季都算 Q1。"""
    assert (fi.fiscal_quarter_of("2026-04-26", 2)
            == fi.fiscal_quarter_of("2026-04-26", 3)
            == fi.fiscal_quarter_of("2026-04-26", 4)
            == "FY2027Q1")
    assert fi.fiscal_quarter_of("2026-04-26", 5) == "FY2026Q4"


def test_workbook_is_marked_for_full_recalc():
    """openpyxl 不算公式，沒有快取值。不強制重算，Excel 可能顯示成空白。"""
    wb, _ = _workbook_with_headers(["FY2026Q1"], ["2026-03-29"])
    fi.apply_fiscal_year_input(wb, fy_end_month=12)
    assert wb.calculation.fullCalcOnLoad is True


def test_missing_sheets_are_tolerated():
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Index")
    fi.apply_fiscal_year_input(wb, fy_end_month=12)     # 不該爆


# ── 字型與字級（2026-08-08 CTH 指定，TODO D0b）────────────────────────────
#
# 輸入格這一塊寫在 Index 上，字型／字級必須跟 excel_formatter 建的表格一致，
# 否則同一頁會混兩種字體，比全部都是預設還醜。常數只有一份，在 excel_formatter。

def _input_block(fy_end_month=12):
    wb, _ = _workbook_with_headers(["FY2026Q1"], ["2026-03-29"])
    fi.apply_fiscal_year_input(wb, fy_end_month=fy_end_month)
    return wb["Index"]


def test_input_block_uses_the_shared_font_family():
    from excel_formatter import FONT_NAME
    ws = _input_block()
    # 只看有內容的格：合併範圍的填充格不顯示，樣式取左上角那格。
    names = {c.font.name for row in ws.iter_rows() for c in row if c.value}
    assert names == {FONT_NAME}, f"混到別的字型：{names}"


def test_input_block_sizes_match_the_index_scale():
    """表格 11、輸入格 12、提醒 10——與 excel_formatter 的 Index 同一級。"""
    from excel_formatter import INDEX_TABLE_SIZE, INDEX_INPUT_SIZE, INDEX_NOTE_SIZE
    ws = _input_block()
    assert ws["A4"].font.size == INDEX_TABLE_SIZE == 11   # 「財年起始月（可修改）」
    assert ws["B4"].font.size == INDEX_INPUT_SIZE == 12   # 黃底輸入格
    assert ws["C4"].font.size == INDEX_TABLE_SIZE         # 財年區間即時回饋
    assert ws["A5"].font.size == INDEX_NOTE_SIZE == 10    # 核對提醒


def test_note_row_is_tall_enough_for_the_whole_text():
    """提醒是 wrap 過的長字串，合併儲存格不會自動調列高——只能自己算。

    CTH 2026-08-08 驗收時回報第 5 列會切到文字。實測：文字顯示寬度 508 個半形
    當量、A~E 合併寬度 86，要 6 行；原本寫死的 28（後來 32）只夠 2.4 行。
    """
    ws = _input_block()
    display = sum(2 if ord(ch) > 0x2000 else 1 for ch in fi._note())
    width = sum(ws.column_dimensions[c].width or 8.43 for c in "ABCDE")
    need_lines = display / width
    assert ws.row_dimensions[5].height >= need_lines * 13.5


def test_note_row_height_follows_the_text_length():
    """列高是算出來的不是寫死的——改了提醒文字不必記得回頭改列高。"""
    tall = _input_block().row_dimensions[5].height
    original = fi._note
    try:
        fi._note = lambda: "短"
        short = _input_block().row_dimensions[5].height
    finally:
        fi._note = original
    assert short < tall


# ── 跨公司對齊用的日曆季（2026-08-22）──────────────────────────────────────
#
# calendar_quarter_of() 看的是「期末日落在哪一季」，那是單一公司自己的標籤。
# 跨公司比較要的是「這一季在市場上跟誰是同一期」——NVDA 7 月底結束的那一季，
# 分析師拿它跟 AMD/INTC 6 月底那一季比（兩家同一波財報），不是跟 9 月那一季比。
# 判準是「這一期的多數天數落在哪個日曆季」＝ 期中點（期末日往前推 45 天）。

@pytest.mark.parametrize("period_end, expected", [
    ("2025-06-28", "2025Q2"),   # AMD / INTC 的 6 月季
    ("2025-07-27", "2025Q2"),   # NVDA 同一波財報，期末日晚一個月
    ("2025-09-27", "2025Q3"),   # AMD 9 月季
    ("2025-10-26", "2025Q3"),   # NVDA 同一波
    ("2026-01-25", "2025Q4"),   # NVDA 財年最後一季，主要落在 2025Q4
    ("2025-12-27", "2025Q4"),   # AMD 12 月季
    ("2025-09-27", "2025Q3"),   # AAPL 9 月季（財年最後一季）也是 CQ3
])
def test_calendarized_quarter_of(period_end, expected):
    assert fi.calendarized_quarter_of(period_end) == expected


def test_calendarized_quarter_of_bad_input_returns_empty():
    assert fi.calendarized_quarter_of("") == ""
    assert fi.calendarized_quarter_of(None) == ""
    assert fi.calendarized_quarter_of("不是日期") == ""


@pytest.mark.parametrize("period_end, expected", [
    ("2026-01-25", "2025"),     # NVDA FY2026 → 主要是日曆 2025 年
    ("2025-12-31", "2025"),     # AMD / INTC
    ("2025-09-27", "2025"),     # AAPL FY2025
    ("2025-06-30", "2025"),     # MSFT FY2025（慣例掛在結束年）
])
def test_calendarized_year_of(period_end, expected):
    assert fi.calendarized_year_of(period_end) == expected


# ── G1：日曆季判準統一成「一份實作、兩個具名基準點」（2026-08-22）─────────
#
# 原本同一件事有三套算法散在三個檔案（見 docs/superpowers/
# design-2026-08-22-period-alignment-and-gaps.md 的 G1）。統一的做法不是併成
# 一個數字——「這季結束在哪一季」跟「這季主要落在哪一季」是兩個問題，強行
# 合一必然弄壞一邊。改成同一份實作 + 兩個具名基準點，而且 basis 不給預設值，
# 強迫每個呼叫端表態。

def test_calendar_quarter_of_requires_explicit_basis():
    """不給 basis 要噴 TypeError——這條就是「統一」的實質保障。"""
    with pytest.raises(TypeError):
        fi.calendar_quarter_of("2025-07-27")


@pytest.mark.parametrize("period_end, basis, expected", [
    # NVDA 一月結算：7 月底結束那季，「結束在」Q3、「主要落在」Q2
    ("2025-07-27", "end", "2025Q3"),
    ("2025-07-27", "span", "2025Q2"),
    ("2026-01-25", "end", "2026Q1"),
    ("2026-01-25", "span", "2025Q4"),
    # 12 月結算的公司兩種基準點一致
    ("2025-06-28", "end", "2025Q2"),
    ("2025-06-28", "span", "2025Q2"),
    # 52/53 週制溢出到下個月：兩種基準點都要吃掉漂移
    ("2023-04-01", "end", "2023Q1"),
    ("2023-04-01", "span", "2023Q1"),
])
def test_calendar_quarter_of_two_bases(period_end, basis, expected):
    assert fi.calendar_quarter_of(period_end, basis=basis) == expected


def test_calendar_quarter_of_rejects_unknown_basis():
    with pytest.raises(ValueError):
        fi.calendar_quarter_of("2025-07-27", basis="middle")


def test_calendarized_quarter_of_is_a_thin_wrapper_over_span():
    """跨公司那兩個薄包裝要跟 basis="span" 完全等價，不可以各自演化。"""
    for pe in ["2025-07-27", "2026-01-25", "2025-06-28", "2023-04-01"]:
        assert fi.calendarized_quarter_of(pe) == fi.calendar_quarter_of(pe, basis="span")


# ── 零下載規則：發布日 + EDGAR fiscal_year_end → 財季（B5）──────────────────
#
# 下面每一組都是 SEC EDGAR 實抓的 Item 2.02 8-K（`output/_8k_audit/final_check.json`，
# 200 份）。`expected` 是 production 那條路算出來的 `fiscal_label`
# （＝下載新聞稿抓期末日再換算），也就是這條零下載規則要對齊的基準。

@pytest.mark.parametrize("announce, fye, expected", [
    # WDC：52/53 週制，FY2026 Q2 真實結束在 2026-01-02（跨年）
    ("20260129", "0703", "FY2026Q2"),
    ("20251030", "0703", "FY2026Q1"),
    ("20260805", "0703", "FY2026Q4"),
    # COST：Q3 真實 5/10 就結束、5/28 就發布，名目季末 5/30 落在發布日之後 → 靠 tol
    ("20260528", "0830", "FY2026Q3"),
    ("20250529", "0830", "FY2025Q3"),
    # MU：FYE 0903，名目「月底」法（規則 B）會整家偏一季
    ("20260318", "0903", "FY2026Q2"),
    ("20250625", "0903", "FY2025Q3"),
    # NVDA / TGT：1-2 月結算的財年編號慣例（本專案比公司自稱多一年）
    ("20260520", "0131", "FY2027Q1"),
    ("20260225", "0131", "FY2026Q4"),
    ("20260819", "0201", "FY2027Q2"),
    # AAPL：12 月以外但非 52/53 週制的對照組
    ("20260731", "0926", "FY2026Q3"),
])
def test_quarter_label_from_announcement(announce, fye, expected):
    assert fi.quarter_label_from_announcement(announce, fye) == expected


def test_quarter_label_from_announcement_accepts_iso_dates():
    assert fi.quarter_label_from_announcement("2026-01-29", "0703") == "FY2026Q2"


def test_quarter_label_from_announcement_uses_the_day_not_the_month_end():
    """名目季末用 MMDD 的那一天，不是當月最後一天。

    差別在名目季末落在發布日之後、要靠 tol 吃下去的那幾份：COST FYE 0830 的
    Q3 名目季末是 5/30，比發布日 5/28 晚 2 天，tol=21 吃得下；若改用月底 5/31
    則要 tol>=3。報告實測的規則 B（名目季末一律取財季結束月的月底）只有 79.8%。
    """
    assert fi.quarter_label_from_announcement("20260528", "0830", tol=2) == "FY2026Q3"
    assert fi.quarter_label_from_announcement("20260528", "0831", tol=2) == ""


def test_quarter_label_from_announcement_tol_zero_falls_back_instead_of_mislabelling():
    """tol=0 時 COST 這一份選不到 5/30，退而選到上一季末 2/28（相距 89 天）
    → sanity check 攔下來回空字串，呼叫端退回舊算法，而不是吐一個錯的標籤。"""
    assert fi.quarter_label_from_announcement("20260528", "0830", tol=0) == ""


@pytest.mark.parametrize("announce, fye", [
    ("20260129", ""),          # EDGAR 沒給 fiscalYearEnd
    ("20260129", "07"),        # 只有月份
    ("20260129", "0732"),      # 不合法的日
    ("20260129", "1303"),      # 不合法的月
    ("20260129", "ABCD"),
    ("", "0703"),
    ("2026", "0703"),
    ("20261340", "0703"),      # 不合法的發布日
])
def test_quarter_label_from_announcement_returns_empty_on_bad_input(announce, fye):
    """算不出來一律回空字串，讓呼叫端退回舊算法——不可以整批失敗。"""
    assert fi.quarter_label_from_announcement(announce, fye) == ""


def test_quarter_label_from_announcement_sanity_check_rejects_a_far_away_quarter_end():
    """sanity check：名目季末離發布日太遠就不採用這條規則，回空字串讓呼叫端退回舊算法。

    WDC 這一份的名目季末是 2026-01-03、發布日 2026-01-29，相距 26 天：
    門檻放 70 天（實測最大發布延遲 58 天）過得了，門檻收到 10 天就過不了。
    """
    assert fi.quarter_label_from_announcement("20260129", "0703") == "FY2026Q2"
    assert fi.quarter_label_from_announcement("20260129", "0703", max_lag_days=10) == ""


def test_quarter_label_from_announcement_month_without_that_day():
    """FYE 的「日」在往前推的月份不存在（09-31）→ 退到當月最後一天。"""
    # FYE 1231 → 候選季末 12-31 / 09-30 / 06-30 / 03-31
    assert fi.quarter_label_from_announcement("20260210", "1231") == "FY2025Q4"
    assert fi.quarter_label_from_announcement("20261105", "1231") == "FY2026Q3"
