"""Tests for comparison_writer.py — 跨公司比較 Excel 輸出。"""
import tempfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from comparison import ComparisonResult
from i18n import t
from comparison_writer import (
    write_compare_data_sheet,
    write_snapshot_sheets,
    write_chart_sheets,
    write_comparison_workbook,
)


def _sample_result():
    return ComparisonResult(
        metrics={
            "Revenue": {
                "NVDA": {"FY2024Q1": 100.0, "FY2024Q2": 110.0},
                "AMD": {"FY2024Q1": 80.0, "FY2024Q2": 90.0},
            },
        },
        period_ends={
            "NVDA": {"FY2024Q1": "2024-03-31", "FY2024Q2": "2024-06-30"},
            "AMD": {"FY2024Q1": "2024-03-31", "FY2024Q2": "2024-06-30"},
        },
        failures=[],
    )


# ── Compare_Data ─────────────────────────────────────────────────────────

def test_compare_data_sheet_has_metric_header_and_period_columns():
    """列號是相對的——最上方還有一張 G2 對應表，區塊位置以 block_ranges 為準。"""
    wb = Workbook()
    result = _sample_result()
    ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]
    data_start, _ = ranges["Revenue"]

    assert ws.cell(row=data_start - 3, column=1).value == "Revenue"
    header_row = [c.value for c in ws[data_start - 2]]
    assert "FY2024Q1" in header_row
    assert "FY2024Q2" in header_row


def test_compare_data_sheet_has_static_period_end_row():
    wb = Workbook()
    result = _sample_result()
    ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]
    end_date_row = ranges["Revenue"][0] - 1

    # 原始 period_ends 是 "2024-03-31" 這種帶連字號格式，寫進表裡要轉成
    # 不帶分隔符的 "YYYYMMDD"，跟 Snapshot 輸入格要求的格式一致
    row3 = [c.value for c in ws[end_date_row]]
    assert "20240331" in row3
    for cell in ws[end_date_row]:
        if cell.value:
            assert not str(cell.value).startswith("=")


def test_compare_data_sheet_lists_company_rows_with_values():
    wb = Workbook()
    result = _sample_result()
    write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]

    company_col_a = [c.value for c in ws["A"]]
    assert "NVDA" in company_col_a
    assert "AMD" in company_col_a


def test_compare_data_sheet_returns_block_ranges():
    wb = Workbook()
    result = _sample_result()
    ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    assert "Revenue" in ranges
    start, end = ranges["Revenue"]
    assert end > start


def test_compare_data_sheet_stacks_multiple_metric_blocks():
    wb = Workbook()
    result = _sample_result()
    result.metrics["Gross Margin (%)"] = {"NVDA": {"FY2024Q1": 50.0}}
    ranges = write_compare_data_sheet(wb, result, ["Revenue", "Gross Margin (%)"])
    rev_start, rev_end = ranges["Revenue"]
    gm_start, gm_end = ranges["Gross Margin (%)"]
    assert gm_start > rev_end


# ── Snapshot / Snapshot_Manual ──────────────────────────────────────────

def test_snapshot_sheet_has_yellow_input_cell_and_formulas():
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_snapshot_sheets(wb, result, ["Revenue"], block_ranges, default_date="20240331")

    ws = wb["Snapshot"]
    assert ws["B1"].value == date(2024, 3, 31)
    assert ws["B1"].fill.fgColor.rgb in ("00FFFF00", "FFFFFF00")

    body = [[c.value for c in row] for row in ws.iter_rows(min_row=3)]
    formula_cells = [v for row in body for v in row if isinstance(v, str) and v.startswith("=")]
    assert formula_cells, "Snapshot 應該用公式，不是寫死的值"
    assert any("INDEX" in f and "SUMPRODUCT" in f for f in formula_cells)


def test_snapshot_input_cell_is_a_real_date_not_text():
    """2026-08-21 CTH 回報：B1 原本是純文字要求剛好打中 YYYYMMDD，使用者打
    數字會被 Excel 自動轉成數值型別，跟 Compare_Data 期末結算日列（文字）
    型別對不上，MATCH 抓不到值。B1 改成真正的日期型別，使用者可以打一般
    日期格式，不用湊 8 碼數字，也不會再有型別不匹配的問題。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_snapshot_sheets(wb, result, ["Revenue"], block_ranges, default_date="20240331")

    b1 = wb["Snapshot"]["B1"]
    assert isinstance(b1.value, (date, datetime))
    assert b1.number_format != "General"


def test_snapshot_formula_finds_nearest_period_not_later_than_input_date():
    """2026-08-21 CTH 要求：輸入日期不用剛好對到期末結算日，抓「不晚於這天
    的最近一期」——分析師查某個時間點看得到的數字，不能用未來才公布的財報
    回填過去的時間點。公式要用 <= 比較（不是精確比對），且要排除空白欄位
    （該公司那期沒資料）。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_snapshot_sheets(wb, result, ["Revenue"], block_ranges, default_date="")

    ws = wb["Snapshot"]
    formula = next(
        c.value for row in ws.iter_rows(min_row=3) for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    )
    assert "<=" in formula
    assert "MATCH" not in formula  # 不是精確比對
    assert "IFERROR" in formula    # 找不到任何一期時要顯示空白，不是報錯


def test_snapshot_sheet_hints_that_any_date_is_accepted():
    """CTH 回報 Snapshot 的黃格不知道要填什麼——提示文字要講清楚可以打一般
    日期，不用剛好對到某一期。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_snapshot_sheets(wb, result, ["Revenue"], block_ranges, default_date="")

    ws = wb["Snapshot"]
    all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "2025/12/31" in all_text or "2025" in all_text


def test_snapshot_sheet_lists_available_dates_for_reference():
    """光講格式還不夠，使用者不知道「有哪些日期可以填」——列出 Compare_Data
    裡實際存在的期末結算日，照現有資料範圍給選項，不用自己去翻 Compare_Data。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_snapshot_sheets(wb, result, ["Revenue"], block_ranges, default_date="")

    ws = wb["Snapshot"]
    all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "20240331" in all_text
    assert "20240630" in all_text


def test_snapshot_manual_sheet_is_blank_with_same_headers():
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_snapshot_sheets(wb, result, ["Revenue"], block_ranges, default_date="20240331")

    ws = wb["Snapshot_Manual"]
    header_row = [c.value for c in ws[1]]
    assert "Revenue" in header_row
    company_col = [c.value for c in ws["A"]]
    assert "NVDA" in company_col
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            assert cell.value is None


# ── Chart_<指標> ─────────────────────────────────────────────────────────

def test_chart_sheet_created_per_metric_with_line_chart():
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    assert "Chart_Revenue" in wb.sheetnames
    ws = wb["Chart_Revenue"]
    assert len(ws._charts) == 1


def test_chart_sheet_uses_period_end_date_row_as_categories():
    """X 軸類別要用 Compare_Data 的期末結算日列（絕對日期，如 20240331），
    不是財季標籤列（如 FY2024Q1）——不同公司財年結束月不同，財季標籤字串
    排序無法反映真實時間順序，會讓 Excel 折線圖亂連線。期末結算日列緊接在
    資料列上方（data_start - 1），財季標籤列則在再上一列（data_start - 2）。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    data_start, _ = block_ranges["Revenue"]
    ws = wb["Chart_Revenue"]
    chart = ws._charts[0]
    cat_ref = chart.series[0].cat.numRef.f if chart.series[0].cat.numRef else chart.series[0].cat.strRef.f
    assert f"${data_start - 1}:" in cat_ref or f"${data_start - 1}$" in cat_ref
    assert f"${data_start - 2}" not in cat_ref


def test_chart_sheet_categories_use_str_ref_not_num_ref():
    """2026-08-22 CTH 截圖回報「中間斷線＋圖例被吃」：根因是 openpyxl 的
    `chart.set_categories()` 不管儲存格實際內容永遠寫成 numRef（數值參照），
    但期末結算日是文字（"20240331"），Excel 拿數值參照指向文字儲存格解析
    不出來，類別軸整個讀不到值，連帶把圖例/座標軸擠壓變形。每個 series 的
    `cat` 必須是 strRef，不能是 numRef。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    chart = wb["Chart_Revenue"]._charts[0]
    for series in chart.series:
        assert series.cat.numRef is None, "類別軸不可以是 numRef——期末結算日是文字"
        assert series.cat.strRef is not None
        assert series.cat.strRef.f  # 有指到範圍


def test_chart_sheet_axis_positions_and_tick_labels_are_explicit():
    """2026-08-22 CTH 截圖回報：X 軸完全沒有日期標籤、圖例被擠壓成一行。
    根因是 openpyxl 兩個軸預設都是 axPos="l"（見 openpyxl.chart.axis.
    _BaseAxis），對 Y 軸剛好對、對 X 軸（該在底部）是錯的，Excel 拿到矛盾
    設定會整排 X 軸標籤不畫。兩軸的位置與刻度標籤顯示都要明講，不能依賴
    openpyxl 的預設值。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    chart = wb["Chart_Revenue"]._charts[0]
    assert chart.x_axis.axPos == "b"
    assert chart.y_axis.axPos == "l"
    assert chart.x_axis.tickLblPos == "nextTo"
    assert chart.y_axis.tickLblPos == "nextTo"


def test_chart_sheet_axes_explicitly_not_deleted():
    """2026-08-22 CTH 截圖回報「中間斷線＋圖例被吃＋沒有單位」：實測用 Excel
    COM 比對原生 Excel 建立的圖表 XML 才抓到真正根因——openpyxl 完全不寫
    <c:delete> 元素，Excel 拿到「沒寫」的軸會保守地不畫刻度標籤（跟明講
    delete=False 待遇不同，這是實測結果不是規格文件寫的）。兩軸都要明講
    delete=False，不能依賴 openpyxl 的預設（不寫）。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    chart = wb["Chart_Revenue"]._charts[0]
    assert chart.x_axis.delete is False
    assert chart.y_axis.delete is False


def test_chart_sheet_sets_display_blanks_as_gap():
    """缺值期間不能被 Excel 直接連到下一個有值的點，否則折線會誤導使用者
    以為中間有連續資料。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    chart = wb["Chart_Revenue"]._charts[0]
    assert chart.display_blanks == "gap"


# ── F3 圖表版面調校（2026-08-21，CTH 截圖回報 5 項，CTH 已確認最終方案）──

def test_chart_sheet_is_double_default_size():
    """CTH 要求長寬各拉長一倍，openpyxl 預設 15cm×7.5cm 偏小，改成 30cm×15cm。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    chart = wb["Chart_Revenue"]._charts[0]
    assert chart.width == 30
    assert chart.height == 15


def test_chart_sheet_legend_does_not_overlay():
    """2026-08-22 CTH 截圖回報圖例文字疊在 X 軸日期上看不清楚：實測用 Excel
    COM 比對原生 Excel 建立的圖表 XML 抓到根因——openpyxl 的 `Legend` 沒有
    `overlay` 屬性時完全不寫該欄位，Excel 拿到「沒寫」會讓圖例跟 X 軸標題/
    刻度標籤擠在同一條窄帶、直接疊在一起。原生 Excel 輸出一定帶
    `overlay="0"`（不要疊加，另外保留專屬空間），這裡要明講，不能依賴
    openpyxl 的預設（不寫）。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    chart = wb["Chart_Revenue"]._charts[0]
    assert chart.legend.overlay is False


def test_chart_sheet_legend_at_bottom():
    """圖例放下方橫排，不放右邊——CTH 確認：公司數量不固定，右側直排公司
    一多會佔掉整個右半邊圖表，下方橫排可以隨公司數自動換行撐住。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    chart = wb["Chart_Revenue"]._charts[0]
    assert chart.legend.position == "b"


def test_chart_sheet_y_axis_uses_same_number_format_as_cells():
    """Y 軸數字格式跟 Compare_Data 儲存格同一套規則（excel_formatter.unit_format_for），
    不要另外定義一套會漂移的格式。金額類指標軸標題帶 ($mm) 單位；百分比類
    指標維持指標名稱本身（格式已經看得出是 %，不用重複講）。"""
    wb = Workbook()
    result = _sample_result()
    result.metrics["Gross Margin (%)"] = {"NVDA": {"FY2024Q1": 50.0}}
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue", "Gross Margin (%)"])
    write_chart_sheets(wb, ["Revenue", "Gross Margin (%)"], block_ranges)

    from excel_formatter import FMT_FINANCIAL, FMT_PERCENT

    rev_chart = wb["Chart_Revenue"]._charts[0]
    assert rev_chart.y_axis.numFmt.formatCode == FMT_FINANCIAL
    assert rev_chart.y_axis.title.text.rich.p[0].r[0].t == "Revenue ($mm)"

    margin_chart = wb["Chart_Gross Margin (%)"]._charts[0]
    assert margin_chart.y_axis.numFmt.formatCode == FMT_PERCENT
    assert margin_chart.y_axis.title.text.rich.p[0].r[0].t == "Gross Margin (%)"


def test_chart_sheet_x_axis_skips_labels_when_many_periods():
    """新發現的問題（不在原本 F3 5 項清單裡）：接上 D0-1 Q4 合成後跨公司比較
    的時間跨度可以拉到 60-70 欄，全部日期標籤硬擠在 X 軸上會疊字看不清楚。
    CTH 確認：目標大約 15 個可視標籤，用 tickLblSkip 跳著顯示。"""
    wb = Workbook()
    periods = [f"FY{2009 + i // 4}Q{i % 4 + 1}" for i in range(60)]
    result = _sample_result()
    result.metrics["Revenue"]["NVDA"] = {p: float(i) for i, p in enumerate(periods)}
    result.period_ends["NVDA"] = {
        p: f"{2009 + i // 4}-{i % 4 + 1:02d}-01" for i, p in enumerate(periods)
    }
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    chart = wb["Chart_Revenue"]._charts[0]
    n_periods = len(periods)
    expected_skip = max(1, n_periods // 15)
    assert chart.x_axis.tickLblSkip == expected_skip
    assert chart.x_axis.tickMarkSkip == expected_skip


def test_chart_sheet_x_axis_no_skip_when_few_periods():
    """期間數少的話（原本的 2 期樣本資料）不用跳著顯示，skip 應該是 1。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    chart = wb["Chart_Revenue"]._charts[0]
    assert chart.x_axis.tickLblSkip == 1


def test_chart_sheet_name_truncates_long_metric_names():
    wb = Workbook()
    result = _sample_result()
    long_name = "A Very Long Metric Name That Exceeds Excel Sheet Name Limit (%)"
    result.metrics[long_name] = {"NVDA": {"FY2024Q1": 1.0}}
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue", long_name])
    write_chart_sheets(wb, ["Revenue", long_name], block_ranges)

    assert all(len(name) <= 31 for name in wb.sheetnames)


# ── write_comparison_workbook ────────────────────────────────────────────

def test_write_comparison_workbook_produces_all_expected_sheets():
    result = _sample_result()
    with tempfile.TemporaryDirectory() as tmp:
        out_path = Path(tmp) / "compare_test.xlsx"
        write_comparison_workbook(result, ["Revenue"], out_path, snapshot_date="20240331")

        assert out_path.exists()
        wb = load_workbook(out_path)
        assert wb.sheetnames == ["Compare_Data", "Notes", "Snapshot",
                                 "Snapshot_Manual", "Chart_Revenue"]


# ── 同一日曆季、各公司期末日不同（2026-08-22）───────────────────────────────

def test_period_end_row_takes_the_latest_date_in_the_calendar_quarter():
    """跨公司改用日曆季對齊後，同一欄各公司的期末日不再相同。

    NVDA 那一季結束在 2025-07-27，AMD 是 2025-06-28。期末結算日列只有一格，
    要取**最晚**的那一個：Snapshot 用它做「不晚於 B1」的判斷，取早的那個會
    讓使用者把 B1 設在 7/1 就看到 NVDA 還沒結算完的那一季數字。
    """
    result = ComparisonResult(
        metrics={"Revenue": {"NVDA": {"2025Q2": 46743.0}, "AMD": {"2025Q2": 7685.0}}},
        period_ends={"NVDA": {"2025Q2": "2025-07-27"}, "AMD": {"2025Q2": "2025-06-28"}},
        failures=[],
    )
    wb = Workbook()
    ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    end_date_row = ranges["Revenue"][0] - 1

    assert wb["Compare_Data"].cell(row=end_date_row, column=2).value == "20250727"


def test_compare_data_sheet_drops_periods_with_no_data_at_all():
    """所有公司、所有指標都沒值的期間不要出現在表上。

    合成 Q4 時年報沒有期末日的那幾欄（label 退回 `FY2009Q4` 這種財季標籤）
    值全是空的，改用日曆季當欄位鍵之後它們會被排到最右邊——圖表 X 軸多出
    兩格 2026Q2 之後的空白，看起來像資料抓錯。
    """
    result = ComparisonResult(
        metrics={"Revenue": {
            "NVDA": {"2025Q1": 44062.0, "2025Q2": 46743.0, "FY2009Q4": None},
            "AMD": {"2025Q1": 7438.0, "2025Q2": 7685.0, "FY2009Q4": None},
        }},
        period_ends={
            "NVDA": {"2025Q1": "2025-04-27", "2025Q2": "2025-07-27", "FY2009Q4": ""},
            "AMD": {"2025Q1": "2025-03-29", "2025Q2": "2025-06-28", "FY2009Q4": ""},
        },
        failures=[],
    )
    wb = Workbook()
    write_compare_data_sheet(wb, result, ["Revenue"])

    header_row = [c.value for c in wb["Compare_Data"][2]]
    assert "FY2009Q4" not in header_row
    assert header_row[1:] == ["2025Q1", "2025Q2"]


def test_compare_data_sheet_keeps_period_when_only_one_company_has_data():
    """只有一家有值的期間要留著——那是真資料，不是空欄。"""
    result = ComparisonResult(
        metrics={"Revenue": {
            "NVDA": {"2025Q1": 44062.0, "2025Q2": 46743.0},
            "AMD": {"2025Q1": 7438.0, "2025Q2": None},
        }},
        period_ends={
            "NVDA": {"2025Q1": "2025-04-27", "2025Q2": "2025-07-27"},
            "AMD": {"2025Q1": "2025-03-29", "2025Q2": ""},
        },
        failures=[],
    )
    wb = Workbook()
    write_compare_data_sheet(wb, result, ["Revenue"])

    header_row = [c.value for c in wb["Compare_Data"][2]]
    assert header_row[1:] == ["2025Q1", "2025Q2"]


def test_chart_titles_do_not_overlay_and_use_auto_layout():
    """2026-08-22 CTH 截圖回報 Y 軸標題壓在「50,000.0」刻度上、X 軸標題「期間」
    掉進日期標籤那一排裡面。跟先前圖例那個 bug 同一類：openpyxl 的
    `Title` 沒有 `overlay` / `layout` 屬性時完全不寫這兩個元素，Excel 拿到
    「沒寫」的標題會直接畫在既有內容上面，而不是另外撥一條專屬空間。
    原生 Excel 輸出的每個標題（圖表、兩個軸）一定帶 `<c:layout/>`（明講
    用自動版面）加 `<c:overlay val="0"/>`。三個標題都要明講。"""
    wb = Workbook()
    result = _sample_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    chart = wb["Chart_Revenue"]._charts[0]
    for title in (chart.title, chart.x_axis.title, chart.y_axis.title):
        assert title.overlay is False
        assert title.layout is not None


def test_chart_title_overlay_is_written_into_the_saved_file():
    """光設屬性不夠——Excel 讀的是存檔後的 XML，要在真的 .xlsx 裡確認。"""
    import zipfile

    result = _sample_result()
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "chart_xml.xlsx"
        write_comparison_workbook(result, ["Revenue"], out)
        with zipfile.ZipFile(out) as z:
            name = next(n for n in z.namelist() if "charts/chart" in n)
            xml = z.read(name).decode("utf-8")

    # 圖表標題 + X 軸標題 + Y 軸標題 + 圖例 = 4 個 overlay；
    # layout 只有三個標題有（圖例的版面由 legendPos 決定）
    assert xml.count('<overlay val="0"/>') == 4
    assert xml.count("<layout/>") == 3


# ── G2 日曆季 ↔ 財季對應表（2026-08-25）────────────────────────────────────

def _fiscal_result():
    """NVDA（財年二月起算）與 AMD（日曆年）同一波財報，財季標籤不同。"""
    return ComparisonResult(
        metrics={"Revenue": {
            "NVDA": {"2025Q2": 46743.0, "2025Q3": 57006.0},
            "AMD": {"2025Q2": 7685.0, "2025Q3": 9246.0},
        }},
        period_ends={
            "NVDA": {"2025Q2": "2025-07-27", "2025Q3": "2025-10-26"},
            "AMD": {"2025Q2": "2025-06-28", "2025Q3": "2025-09-27"},
        },
        failures=[],
        fiscal_labels={
            "NVDA": {"2025Q2": "FY2026Q2", "2025Q3": "FY2026Q3"},
            "AMD": {"2025Q2": "FY2025Q2", "2025Q3": "FY2025Q3"},
        },
        synthetic_q4={"NVDA": set(), "AMD": set()},
    )


def test_compare_data_starts_with_the_calendar_to_fiscal_map():
    """整張 Compare_Data 最上方是對應表，在第一個指標區塊之前。一格講完
    「這一欄對這家公司是哪一財季、期末日幾號」。"""
    wb = Workbook()
    write_compare_data_sheet(wb, _fiscal_result(), ["Revenue"])
    ws = wb["Compare_Data"]

    header = [c.value for c in ws[2]]
    assert header[1:] == ["2025Q2", "2025Q3"]

    rows = {ws.cell(row=r, column=1).value: [ws.cell(row=r, column=c).value for c in (2, 3)]
            for r in (3, 4)}
    assert rows["NVDA"] == ["FY2026Q2 (0727)", "FY2026Q3 (1026)"]
    assert rows["AMD"] == ["FY2025Q2 (0628)", "FY2025Q3 (0927)"]


def test_calendar_to_fiscal_map_sits_above_the_first_metric_block():
    wb = Workbook()
    ranges = write_compare_data_sheet(wb, _fiscal_result(), ["Revenue"])
    ws = wb["Compare_Data"]
    data_start, _ = ranges["Revenue"]

    # 指標區塊的標題列（Revenue）在對應表下方，不是第 1 列
    title_row = data_start - 3
    assert ws.cell(row=title_row, column=1).value == "Revenue"
    assert title_row > 4


def test_metric_blocks_still_show_calendar_quarters_only():
    """對應表講完財季之後，下面的指標區塊只給日曆季，不重複財季。"""
    wb = Workbook()
    ranges = write_compare_data_sheet(wb, _fiscal_result(), ["Revenue"])
    ws = wb["Compare_Data"]
    data_start, _ = ranges["Revenue"]

    block_header = [c.value for c in ws[data_start - 2]]
    assert block_header[1:] == ["2025Q2", "2025Q3"]


def test_calendar_to_fiscal_map_leaves_a_blank_cell_when_a_company_lacks_the_period():
    wb = Workbook()
    result = _fiscal_result()
    del result.metrics["Revenue"]["AMD"]["2025Q3"]
    del result.period_ends["AMD"]["2025Q3"]
    del result.fiscal_labels["AMD"]["2025Q3"]
    write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]

    amd_row = next(r for r in (3, 4) if ws.cell(row=r, column=1).value == "AMD")
    assert ws.cell(row=amd_row, column=3).value is None


def test_calendar_to_fiscal_map_reflects_a_mid_stream_fiscal_year_change():
    """公司中途改財年，那一欄自己就會反映出來——每一格都是逐期從實際期末日
    算的，不是從一個「財年開始月份」推的，所以不需要任何例外處理。"""
    result = ComparisonResult(
        metrics={"Revenue": {"XYZ": {"2024Q1": 1.0, "2024Q2": 2.0, "2024Q3": 3.0}}},
        period_ends={"XYZ": {"2024Q1": "2024-03-30", "2024Q2": "2024-06-29",
                             "2024Q3": "2024-09-28"}},
        failures=[],
        fiscal_labels={"XYZ": {"2024Q1": "FY2024Q1", "2024Q2": "FY2024Q2",
                               "2024Q3": "FY2025Q1"}},
    )
    wb = Workbook()
    write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]

    assert [ws.cell(row=3, column=c).value for c in (2, 3, 4)] == [
        "FY2024Q1 (0330)", "FY2024Q2 (0629)", "FY2025Q1 (0928)"]


def test_snapshot_formulas_still_point_at_the_right_rows_after_the_map_is_inserted():
    """最容易壞的地方：對應表把所有列號往下推，Snapshot 公式必須跟著走。
    這裡不比對列號常數，直接回頭查那個列號在 Compare_Data 上是誰。"""
    import re

    wb = Workbook()
    result = _fiscal_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_snapshot_sheets(wb, result, ["Revenue"], block_ranges, default_date="20250727")

    data_ws = wb["Compare_Data"]
    snap = wb["Snapshot"]
    for r in range(3, 5):
        company = snap.cell(row=r, column=1).value
        formula = snap.cell(row=r, column=2).value
        rows = {int(m) for m in re.findall(r"Compare_Data!\$B\$(\d+):", formula)}
        assert rows, formula
        for row_no in rows:
            first = data_ws.cell(row=row_no, column=1).value
            # 公式只會指到兩種列：這家公司的資料列，或期末結算日列
            assert first in (company, t("compare.xls.period_end")), (
                f"{company} 的公式指到第 {row_no} 列，那是 {first!r}")
        # 資料列的確有值（不是指到對應表那幾列）
        data_row = next(row_no for row_no in rows
                        if data_ws.cell(row=row_no, column=1).value == company)
        assert isinstance(data_ws.cell(row=data_row, column=2).value, (int, float))


def test_chart_categories_still_point_at_the_period_end_row_after_the_map():
    wb = Workbook()
    result = _fiscal_result()
    block_ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    write_chart_sheets(wb, ["Revenue"], block_ranges)

    data_start, _ = block_ranges["Revenue"]
    data_ws = wb["Compare_Data"]
    assert data_ws.cell(row=data_start - 1, column=2).value == "20250727"
    chart = wb["Chart_Revenue"]._charts[0]
    assert f"${data_start - 1}" in chart.series[0].cat.strRef.f


# ── G7 說明 sheet（2026-08-25）────────────────────────────────────────────

def _notes_rows(ws):
    """(勾選, 標題, 內文, 實際情況) 的資料列，跳過標題與表頭。"""
    return [tuple(ws.cell(row=r, column=c).value for c in range(1, 5))
            for r in range(3, ws.max_row + 1)
            if ws.cell(row=r, column=2).value]


def test_notes_sheet_sits_right_after_compare_data():
    result = _fiscal_result()
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "notes.xlsx"
        write_comparison_workbook(result, ["Revenue"], out)
        wb = load_workbook(out)
    assert wb.sheetnames[:2] == ["Compare_Data", "Notes"]


def test_notes_sheet_lists_every_item_with_a_checkbox_column():
    from comparison_writer import NOTE_ITEMS, write_notes_sheet

    wb = Workbook()
    wb.active.title = "Compare_Data"
    write_notes_sheet(wb, _fiscal_result(), ["Revenue"])
    rows = _notes_rows(wb["Notes"])

    # 沒有抓取失敗的公司時第 10 條不出現（只在真的踩到時才佔一列）
    assert len(rows) == len(NOTE_ITEMS) - 1
    for check, title, body, _detail in rows:
        assert check in ("✓", None)
        assert title and body


def test_notes_sheet_always_ticks_the_structural_items():
    """「時間軸怎麼定的」「單位」「符號」這幾條是結構性的，任何一份檔案都成立。"""
    from comparison_writer import write_notes_sheet

    wb = Workbook()
    wb.active.title = "Compare_Data"
    write_notes_sheet(wb, _fiscal_result(), ["Revenue"])
    rows = {title: check for check, title, _b, _d in _notes_rows(wb["Notes"])}

    for key in ("compare.xls.notes.timeline", "compare.xls.notes.units",
                "compare.xls.notes.sign", "compare.xls.notes.as_reported"):
        assert rows[t(key)] == "✓", key


def test_notes_sheet_ticks_fiscal_year_difference_only_when_it_exists():
    from comparison_writer import write_notes_sheet

    def _checks(result):
        wb = Workbook()
        wb.active.title = "Compare_Data"
        write_notes_sheet(wb, result, ["Revenue"])
        return {title: check for check, title, _b, _d in _notes_rows(wb["Notes"])}

    differing = _checks(_fiscal_result())          # NVDA 1 月結算、AMD 12 月
    assert differing[t("compare.xls.notes.not_fiscal")] == "✓"
    assert differing[t("compare.xls.notes.not_period_end")] == "✓"

    same = _fiscal_result()
    same.period_ends["NVDA"] = dict(same.period_ends["AMD"])
    same.fiscal_labels["NVDA"] = dict(same.fiscal_labels["AMD"])
    assert _checks(same)[t("compare.xls.notes.not_fiscal")] is None


def test_notes_sheet_reports_the_period_end_spread_within_a_column():
    from comparison_writer import write_notes_sheet

    wb = Workbook()
    wb.active.title = "Compare_Data"
    write_notes_sheet(wb, _fiscal_result(), ["Revenue"])
    rows = {title: (check, detail) for check, title, _b, detail in _notes_rows(wb["Notes"])}

    check, detail = rows[t("compare.xls.notes.period_end_row")]
    assert check == "✓"
    assert "29" in detail and "2025Q2" in detail   # 07-27 與 06-28 差 29 天


def test_notes_sheet_ticks_synthetic_q4_only_when_the_file_has_one():
    from comparison_writer import write_notes_sheet

    def _row(result):
        wb = Workbook()
        wb.active.title = "Compare_Data"
        write_notes_sheet(wb, result, ["Revenue"])
        return {title: (check, detail)
                for check, title, _b, detail in _notes_rows(wb["Notes"])
                }[t("compare.xls.notes.synth_q4")]

    assert _row(_fiscal_result())[0] is None

    with_q4 = _fiscal_result()
    with_q4.synthetic_q4["NVDA"] = {"2025Q3"}
    check, detail = _row(with_q4)
    assert check == "✓"
    assert "2025Q3" in detail


def test_notes_sheet_ticks_blanks_only_when_a_cell_is_empty():
    from comparison_writer import write_notes_sheet

    def _check(result):
        wb = Workbook()
        wb.active.title = "Compare_Data"
        write_notes_sheet(wb, result, ["Revenue"])
        return {title: check for check, title, _b, _d
                in _notes_rows(wb["Notes"])}[t("compare.xls.notes.blanks")]

    assert _check(_fiscal_result()) is None

    holed = _fiscal_result()
    holed.metrics["Revenue"]["AMD"]["2025Q3"] = None
    assert _check(holed) == "✓"


def test_notes_sheet_names_the_companies_that_are_missing_from_this_file():
    """最重要的一條：抓取失敗現在只寫進 GUI log，檔案裡完全看不出來——
    檔名有 TSM、使用者也選了 TSM，但表上只有三家，沒有錯誤訊息也沒有空欄位。"""
    from comparison import CompanyFetchError
    from comparison_writer import write_notes_sheet

    result = _fiscal_result()
    result.failures = [CompanyFetchError(ticker="TSM", error_type="NoDataForFrequency")]
    wb = Workbook()
    wb.active.title = "Compare_Data"
    write_notes_sheet(wb, result, ["Revenue"])
    rows = {title: (check, detail) for check, title, _b, detail in _notes_rows(wb["Notes"])}

    check, detail = rows[t("compare.xls.notes.missing_companies")]
    assert check == "✓"
    assert "TSM" in detail and "NoDataForFrequency" in detail


def test_notes_items_are_data_driven_and_translated_in_every_language():
    """CTH 明講這張表未來會擴充：新增一條只要在 NOTE_ITEMS 加一行 + 四個
    locale 各加兩條，不可以把文字寫死在版面程式裡。"""
    import i18n
    from comparison_writer import NOTE_ITEMS

    try:
        for lang, _, _ in i18n.LANGUAGES:
            table = i18n._strings(lang)
            for item in NOTE_ITEMS:
                assert item.title_key in table, f"{lang} 缺 {item.title_key}"
                assert item.body_key in table, f"{lang} 缺 {item.body_key}"
    finally:
        i18n.set_lang("zh_tw")


# ── G6 抓不到的季度留一整欄空白（2026-08-25）──────────────────────────────

def _gap_result(period_ends_by_company, values_by_company):
    metrics = {"Revenue": {
        c: {p: v for p, v in vals.items()} for c, vals in values_by_company.items()}}
    return ComparisonResult(
        metrics=metrics,
        period_ends=period_ends_by_company,
        failures=[],
        fiscal_labels={c: {p: "" for p in ends} for c, ends in period_ends_by_company.items()},
        synthetic_q4={c: set() for c in period_ends_by_company},
    )


def test_a_quarter_nobody_fetched_still_gets_a_column():
    """現況是「成功抓到什麼就放什麼」，某一季掛掉整欄消失，畫面上 2025Q1 直接
    跳到 2025Q3，使用者與 AI 都看不出中間漏了一季。改成保留欄位、內容全空。"""
    result = _gap_result(
        {"AMD": {"2025Q1": "2025-03-29", "2025Q3": "2025-09-27"}},
        {"AMD": {"2025Q1": 7438.0, "2025Q3": 9246.0}},
    )
    wb = Workbook()
    ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]
    data_start, _ = ranges["Revenue"]

    header = [c.value for c in ws[data_start - 2]]
    assert header[1:] == ["2025Q1", "2025Q2", "2025Q3"]
    # 缺的那一欄整欄空白：資料格與期末結算日都沒有，欄位本身留著
    assert ws.cell(row=data_start, column=3).value is None
    assert ws.cell(row=data_start - 1, column=3).value is None


def test_a_sixteen_week_fourth_quarter_is_not_mistaken_for_a_missing_quarter():
    """COSTCO 的第四季是 16 週（112~119 天）。用固定門檻（例如「>120 天算缺」）
    會把它誤判成缺一季——`round(112/91) = 1` 才是對的。52 家 1,482 對相鄰期間
    實測，111~150 天那 16 筆全部是 COSTCO。"""
    result = _gap_result(
        {"COST": {"2024Q1": "2024-02-18", "2024Q2": "2024-05-12",
                  "2024Q3": "2024-09-01", "2024Q4": "2024-11-24"}},
        {"COST": {"2024Q1": 1.0, "2024Q2": 2.0, "2024Q3": 3.0, "2024Q4": 4.0}},
    )
    wb = Workbook()
    ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]

    header = [c.value for c in ws[ranges["Revenue"][0] - 2]]
    assert header[1:] == ["2024Q1", "2024Q2", "2024Q3", "2024Q4"]


def test_only_one_company_missing_a_quarter_does_not_add_a_column():
    """另一家有抓到，那一欄本來就在——不需要補，也不該補出重複欄。"""
    result = _gap_result(
        {"AMD": {"2025Q1": "2025-03-29", "2025Q3": "2025-09-27"},
         "NVDA": {"2025Q1": "2025-04-27", "2025Q2": "2025-07-27",
                  "2025Q3": "2025-10-26"}},
        {"AMD": {"2025Q1": 7438.0, "2025Q3": 9246.0},
         "NVDA": {"2025Q1": 44062.0, "2025Q2": 46743.0, "2025Q3": 57006.0}},
    )
    wb = Workbook()
    ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]

    header = [c.value for c in ws[ranges["Revenue"][0] - 2]]
    assert header[1:] == ["2025Q1", "2025Q2", "2025Q3"]


def test_a_single_gap_never_generates_more_than_four_columns():
    """實測 52 家沒有任何 >210 天（>2 季）的缺口，真的出現就是資料異常，
    不該讓程式無限生欄。"""
    result = _gap_result(
        {"AMD": {"2016Q1": "2016-03-26", "2025Q3": "2025-09-27"}},
        {"AMD": {"2016Q1": 1.0, "2025Q3": 2.0}},
    )
    wb = Workbook()
    ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]

    header = [c.value for c in ws[ranges["Revenue"][0] - 2]]
    assert header[1:] == ["2016Q1", "2025Q3"]


def test_gap_columns_never_extend_past_the_earliest_or_latest_period():
    """只補在「最早抓到的那一季」與「最新」之間，不往更早補（CTH 已定）。"""
    result = _gap_result(
        {"AMD": {"2025Q1": "2025-03-29", "2025Q3": "2025-09-27"}},
        {"AMD": {"2025Q1": 7438.0, "2025Q3": 9246.0}},
    )
    wb = Workbook()
    ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]

    header = [c.value for c in ws[ranges["Revenue"][0] - 2] if c.value]
    assert header[1] == "2025Q1" and header[-1] == "2025Q3"


def test_annual_output_fills_a_missing_year():
    """年度輸出的欄位是年份，缺的那一年同樣要留一欄空白。"""
    result = _gap_result(
        {"AMD": {"2022": "2022-12-31", "2024": "2024-12-28"}},
        {"AMD": {"2022": 1.0, "2024": 2.0}},
    )
    wb = Workbook()
    ranges = write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]

    header = [c.value for c in ws[ranges["Revenue"][0] - 2]]
    assert header[1:] == ["2022", "2023", "2024"]


def test_the_fiscal_map_shows_the_gap_column_too():
    """對應表跟下面的區塊同一組欄位——缺的那一欄在對應表上也是空的。"""
    result = _gap_result(
        {"AMD": {"2025Q1": "2025-03-29", "2025Q3": "2025-09-27"}},
        {"AMD": {"2025Q1": 7438.0, "2025Q3": 9246.0}},
    )
    result.fiscal_labels["AMD"] = {"2025Q1": "FY2025Q1", "2025Q3": "FY2025Q3"}
    wb = Workbook()
    write_compare_data_sheet(wb, result, ["Revenue"])
    ws = wb["Compare_Data"]

    assert [c.value for c in ws[2]][1:] == ["2025Q1", "2025Q2", "2025Q3"]
    assert ws.cell(row=3, column=3).value is None


def test_the_notes_sheet_counts_the_gap_column_as_blank():
    """補出來的空白欄就是「有缺」，說明 sheet 的「空白代表什麼」要打勾。"""
    from comparison_writer import write_notes_sheet

    result = _gap_result(
        {"AMD": {"2025Q1": "2025-03-29", "2025Q3": "2025-09-27"}},
        {"AMD": {"2025Q1": 7438.0, "2025Q3": 9246.0}},
    )
    wb = Workbook()
    wb.active.title = "Compare_Data"
    write_notes_sheet(wb, result, ["Revenue"])
    rows = {title: check for check, title, _b, _d in _notes_rows(wb["Notes"])}
    assert rows[t("compare.xls.notes.blanks")] == "✓"
