"""spike_derive_mapping.py — 用現行路徑的數字反推 companyfacts 的 concept 對照表。

TODO G11 的第二步。**不改任何現有程式。**

## 為什麼不用手填

模板的 `std_concept` 欄是 edgartools 正規化過的名字（`Revenue`、`NetIncome`、
`ResearchAndDevelopmentExpenses`），不是原始 us-gaap element name
（`Revenues`、`NetIncomeLoss`、`ResearchAndDevelopmentExpense`）。憑印象填 75 列
一定會錯，而且錯了不會有人發現——數字看起來都很像。

## 做法

現行路徑抓到的數字是已知正確的答案（跑過 1000 條測試、跟公開財報對過）。
拿那組數字當答案卷，對 companyfacts 裡的每一個 concept 算「同一個期末日、
數字對得上的比例」，命中率最高的那個就是正確 mapping。

順便偵測兩件事：
  - **正負號相反**：先直接比，不行再全部乘 -1 比一次
  - **跨公司不一致**：同一列在不同公司可能對到不同 concept（早年/近年、
    或產業慣例不同），所以要多跑幾家再取聯集，依命中率排序當 fallback 鏈

用法：
    venv/Scripts/python.exe scripts/spike_derive_mapping.py NVDA AMD INTC AAPL

輸出 `output/_spike/mapping_candidates.json` 與一份人眼可讀的摘要。
"""
from __future__ import annotations

import json
import pickle
import sys
import time
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

import requests

import config
import fetcher_facts as ff
from edgar import Company, set_identity
from fetcher_gaap import BS_TEMPLATE, CF_TEMPLATE, IS_TEMPLATE, fetch_gaap_statements

CACHE = ROOT / "output" / "_spike"
CACHE.mkdir(parents=True, exist_ok=True)

# 一列至少要有這麼多期能比，命中率才有意義。太少的話隨便一個 concept
# 都可能剛好對上一兩格。
_MIN_OVERLAP = 4
# 命中率低於這個就不當候選，避免報一堆雜訊
_MIN_HIT_RATE = 0.6

_KIND_BY_TEMPLATE = {"IS": "quarter", "BS": "instant", "CF": "quarter"}


def _load_facts(ticker: str, cik: int, identity: str) -> dict:
    p = CACHE / f"facts_{ticker}.json"
    if not p.exists():
        url = ff.COMPANYFACTS_URL.format(cik=cik)
        p.write_bytes(requests.get(url, headers={"User-Agent": identity},
                                   timeout=60).content)
    return json.loads(p.read_bytes())


# 答案卷只需要「夠多期能比對」，不需要抓好抓滿。現行路徑每份 filing 要解析
# 1.5 秒且解 4 次，抓 80 份就是 150 秒；抓 16 份約 30 秒。評分門檻是
# `_MIN_OVERLAP = 4`，16 季遠遠超過，證據強度沒有實質損失。
# 這只影響「建 mapping」這件一次性的事，不影響 companyfacts 路徑本身。
_ANSWER_KEY_FILINGS = 16
_ANSWER_KEY_ANNUALS = 5


def _load_gaap(ticker: str, identity: str):
    """現行路徑的結果，當作反推 mapping 的答案卷。

    快取起來——這一步是整個流程唯一慢的地方（companyfacts 那半只要 0.5 秒）。
    """
    p = CACHE / f"gaap_{ticker}.pkl"
    if p.exists():
        return pickle.loads(p.read_bytes())
    t0 = time.time()
    tables = fetch_gaap_statements(ticker, identity,
                                   max_filings=_ANSWER_KEY_FILINGS,
                                   max_annual_filings=_ANSWER_KEY_ANNUALS,
                                   fetch_quarterly=True, fetch_annual=True)
    print(f"  現行路徑 {time.time() - t0:.0f}s", file=sys.stderr)
    q = next(t for t in tables if t.sheet_name == "Data_Financials(Q)")
    # 只留下比對需要的東西，pickle 才小
    out = {"labels": list(q.quarter_labels), "ends": list(q.period_ends or []),
           "concepts": list(q.concepts), "values": [list(r) for r in q.values]}
    p.write_bytes(pickle.dumps(out))
    return out


def _answer_series(gaap: dict) -> dict[str, dict[str, float]]:
    """現行路徑 → {列名: {期末日: 值}}。同名列取第一個（跟 Excel 上看到的一致）。"""
    out: dict[str, dict[str, float]] = {}
    for i, name in enumerate(gaap["concepts"]):
        if not name or name in out:
            continue
        series = {}
        for j, end in enumerate(gaap["ends"]):
            v = gaap["values"][i][j]
            if len(end or "") == 10 and isinstance(v, (int, float)):
                series[end] = float(v)
        if series:
            out[name] = series
    return out


# 掃描時要涵蓋的單位。只掃 USD 的話，EPS（USD/shares）與股數（shares）永遠
# 找不到候選——2026-08-22 第一輪 50 家推導就是這樣，5 個列被誤判成「模板要調整」。
_UNITS = ("USD", "USD/shares", "shares")


# 要掃的 taxonomy。流通股數在 `dei` 不在 `us-gaap`——2026-08-22 第二輪推導
# 「Shares Outstanding 找不到候選」就是因為只掃了 us-gaap。
_TAXONOMIES = ("us-gaap", "dei")


def _all_concept_series(raw: dict, kind: str) -> dict[tuple, dict[str, float]]:
    """companyfacts 的每個 (taxonomy, concept, 單位) → {期末日: 值}。

    key 帶單位與 taxonomy 是因為同一個名字可能出現在不同 taxonomy／不同單位，
    而下游 mapping 必須記下這一列到底該去哪裡、用什麼單位取。
    """
    out = {}
    for taxonomy in _TAXONOMIES:
        for name, node in raw.get("facts", {}).get(taxonomy, {}).items():
            for unit in _UNITS:
                buckets: dict[str, list[dict]] = {}
                for fact in node.get("units", {}).get(unit, []):
                    if ff.classify_period(fact) != kind:
                        continue
                    buckets.setdefault(fact["end"], []).append(fact)
                if buckets:
                    out[(taxonomy, name, unit)] = {
                        e: float(ff.pick_fact(v, prefer="as_reported")["val"])
                        for e, v in buckets.items()}
    return out


def _score(answer: dict[str, float], cand: dict[str, float]) -> tuple[int, int, bool]:
    """回 (命中數, 可比期數, 是否要反號)。"""
    shared = set(answer) & set(cand)
    if len(shared) < _MIN_OVERLAP:
        return 0, len(shared), False
    def hits(sign):
        return sum(1 for e in shared
                   if abs(answer[e] - sign * cand[e]) <= 1e-6 * max(abs(answer[e]), 1.0))
    plus, minus = hits(1), hits(-1)
    return (minus, len(shared), True) if minus > plus else (plus, len(shared), False)


# 大中小型 × 跨產業。刻意放進金融股（JPM/GS/BAC/SCHW）與 REIT（PLD）——它們的
# 報表結構跟製造業差很多，是檢驗「模板列到底通不通用」最有效的一群。小型股
# （ARLO/FORM/ONTO/LITE/COHR）用來看某些 concept 是不是只有大公司才會 tag。
DEFAULT_50 = [
    # 半導體 / 硬體
    "NVDA", "AMD", "INTC", "AVGO", "MRVL", "TXN", "QCOM", "MU", "AMAT", "LRCX",
    "KLAC", "ADI", "NXPI", "ON", "SWKS", "FORM", "ONTO", "COHR", "LITE", "ARLO",
    # 軟體 / 網路
    "MSFT", "AAPL", "GOOGL", "META", "AMZN", "CRM", "ADBE", "ORCL", "NOW", "PANW",
    "SNOW", "DDOG",
    # 金融（報表結構不同，刻意納入）
    "JPM", "GS", "BAC", "SCHW",
    # 醫療 / 消費 / 工業 / 能源 / 公用 / REIT
    "JNJ", "PFE", "UNH", "KO", "PG", "WMT", "COST", "NKE", "MCD",
    "CAT", "GE", "XOM", "CVX", "NEE", "PLD",
]


def derive(tickers, identity):
    """回 (found, answer_counts, row_kind)。

    `answer_counts[列名]` = 現行路徑在幾家公司抓到這一列。它是覆蓋率的分母
    ——沒有答案卷就無從比對，不能算進「這個 mapping 不通用」。
    """
    found = defaultdict(lambda: defaultdict(list))
    answer_counts = defaultdict(int)
    row_kind = {}
    for tag, template in [("IS", IS_TEMPLATE), ("BS", BS_TEMPLATE), ("CF", CF_TEMPLATE)]:
        for row in template:
            row_kind.setdefault(row[0], _KIND_BY_TEMPLATE[tag])

    for n, ticker in enumerate(tickers, 1):
        print(file=sys.stderr)
        print(f"=== [{n}/{len(tickers)}] {ticker} ===", file=sys.stderr)
        try:
            cik = int(Company(ticker).cik)
            raw = _load_facts(ticker, cik, identity)
            answers = _answer_series(_load_gaap(ticker, identity))
        except Exception as e:
            print(f"  跳過：{type(e).__name__}: {e}", file=sys.stderr)
            continue
        by_kind = {k: _all_concept_series(raw, k) for k in ("quarter", "instant")}

        for row_name, answer in answers.items():
            kind = row_kind.get(row_name)
            if kind is None:
                continue          # overflow 列，不在模板裡，不管
            answer_counts[row_name] += 1
            # 模板把 CF 整張表當期間值，但裡面混了時點值（`Ending Cash` 是
            # 期末餘額）。兩種 kind 都試，讓資料自己決定，不要被模板分類綁死。
            for k in ("quarter", "instant"):
                for key, cand in by_kind[k].items():
                    hit, overlap, flipped = _score(answer, cand)
                    if overlap >= _MIN_OVERLAP and hit / overlap >= _MIN_HIT_RATE:
                        found[row_name][(k,) + key].append(
                            (ticker, round(hit / overlap, 3), overlap, flipped))
    return found, answer_counts, row_kind


def rank(found, answer_counts):
    """{列名: [候選, ...]}，依「贏了幾家」排序。

    50 家的規模下「全公司一致」不是合理門檻——金融股的報表結構跟製造業本來
    就不同，小公司也未必 tag 得到所有 concept。改用**覆蓋率加權**：分母是
    「現行路徑在這家有抓到這一列」的家數，分子是「候選對得上」的家數。
    """
    out = {}
    for row_name, cands in sorted(found.items()):
        ranked = []
        for (kind, taxonomy, concept, unit), hits in cands.items():
            rates = [h[1] for h in hits]
            ranked.append({
                "concept": concept,
                "kind": kind,
                "taxonomy": taxonomy,
                "unit": unit,
                "avg_hit_rate": round(sum(rates) / len(rates), 3),
                "companies": len(hits),
                "coverage": round(len(hits) / max(answer_counts.get(row_name, 1), 1), 3),
                "negate": all(h[3] for h in hits),
                "won_in": sorted(h[0] for h in hits),
            })
        ranked.sort(key=lambda r: (-r["companies"], -r["avg_hit_rate"]))
        out[row_name] = ranked
    return out


def emit_mapping(ranked, row_kind, answer_counts,
                 min_companies=3, min_rate=0.9):
    """候選清單 → 可以直接餵給 `fetcher_facts.build_table()` 的 mapping。

    一列可能有多個 concept 各自贏在不同公司（早年/近年換過 tag、產業慣例不同），
    **全部收進來排成 fallback 鏈**，依贏的家數排序。

    `negate` 只在**所有**入選 concept 都需要反號時才設 True——只有部分要反號
    代表這個 mapping 可疑（很可能對到不同科目），寧可留給人看也不要自動決定。
    """
    # 某個 concept 如果已經是**另一列**的首選，就不可以再當這一列的備援。
    # 2026-08-22 實測踩到：`OtherNonoperatingIncomeExpense` 被推成
    # `Operating Income` 的備援——它在某幾家數字剛好對得上，但語意是「營業外
    # 收支」，換一家就會把營業外損益填進營業利益。純看命中率擋不掉這種錯，
    # 要靠「一個 concept 只能是一列的主人」這條結構性規則。
    owner = {}
    for row_name, cands in ranked.items():
        for c in cands:
            if c["companies"] >= min_companies and c["avg_hit_rate"] >= min_rate:
                key = (c["taxonomy"], c["concept"], c["unit"])
                if key not in owner or c["companies"] > owner[key][1]:
                    owner[key] = (row_name, c["companies"])
                break                      # 只看每一列的首選

    out = {}
    for row_name, cands in ranked.items():
        good = [c for c in cands
                if c["companies"] >= min_companies and c["avg_hit_rate"] >= min_rate]
        if not good:
            continue
        # 首選一律保留（它就是這一列的主人）；備援要通過「不是別人的首選」
        good = [good[0]] + [
            c for c in good[1:]
            if owner.get((c["taxonomy"], c["concept"], c["unit"]), (row_name,))[0] == row_name
            # 備援的符號慣例必須跟首選一致，否則單一 negate 表達不了，
            # 落到那個備援時就會靜默生出正負相反的數字
            # 備援的符號慣例必須跟首選一致，否則同一列會時正時負
            and c["negate"] == good[0]["negate"]
        ]
        good = [c for c in good
                if (c["kind"], c["unit"], c["taxonomy"])
                == (good[0]["kind"], good[0]["unit"], good[0]["taxonomy"])]
        out[row_name] = {
            "concepts": [c["concept"] for c in good],
            # kind / unit / taxonomy 一律以「贏最多家」的那個候選為準，
            # 並把不同組合的候選剔掉——混在一起代表 mapping 可疑
            # （很可能對到不同科目），寧可少收也不要收錯
            "kind": good[0]["kind"],
            "unit": good[0]["unit"],
            "taxonomy": good[0]["taxonomy"],
            # **符號一律照公司原始申報，不做正規化**（CTH 2026-08-22 決定：
            # 「尊重公司原始資料，使用者要查找時會自己處理」）。所以這裡永遠
            # 是 False。`negate` 欄位保留是因為 `fetcher_facts.resolve_row()`
            # 支援它、也有測試釘住，日後若改政策不必重寫機制。
            #
            # 附帶說明：反推出來的 `negate` 旗標仍然有診斷價值——它揭露了
            # 現行路徑的符號本身就不一致（同一家公司的 Capex 早年正、近年負），
            # 那份分析留在 docs/superpowers/report-2026-08-22-g11-companyfacts.md
            "negate": False,
            "_evidence": {
                "answer_companies": answer_counts.get(row_name, 0),
                "won": {c["concept"]: c["companies"] for c in good},
                "avg_hit_rate": {c["concept"]: c["avg_hit_rate"] for c in good},
            },
        }
    return out


def export_xlsx(ranked, mapping, answer_counts, row_kind, tickers, path):
    """把結果存成 Excel，方便人眼掃過去（CTH 2026-08-22 要求）。

    三張表：
      Mapping     —— 採用的對照表，含證據欄
      Candidates  —— 所有候選，含贏在哪幾家
      Coverage    —— 每個模板列 × 每家公司：現行路徑有沒有答案卷、facts 有沒有對上
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    bold = Font(bold=True)
    warn = PatternFill("solid", fgColor="FFFFF2CC")

    ws = wb.active
    ws.title = "Mapping"
    hdr = ["模板列", "kind", "單位", "taxonomy", "反號", "主 concept", "備援 concept",
           "答案卷家數", "主 concept 贏幾家", "覆蓋率", "平均命中率"]
    ws.append(hdr)
    for c in ws[1]:
        c.font = bold
    for name in sorted(mapping):
        m = mapping[name]
        ev = m["_evidence"]
        first = m["concepts"][0]
        cov = ev["won"][first] / max(ev["answer_companies"], 1)
        ws.append([name, m["kind"], m.get("unit", "USD"),
                   m.get("taxonomy", "us-gaap"), "Y" if m["negate"] else "",
                   first, ", ".join(m["concepts"][1:]),
                   ev["answer_companies"], ev["won"][first],
                   round(cov, 3), ev["avg_hit_rate"][first]])
        if cov < 0.8:
            for c in ws[ws.max_row]:
                c.fill = warn

    ws2 = wb.create_sheet("Candidates")
    ws2.append(["模板列", "concept", "kind", "單位", "taxonomy", "贏幾家", "覆蓋率", "平均命中率", "反號", "贏在哪幾家"])
    for c in ws2[1]:
        c.font = bold
    for name in sorted(ranked):
        for cand in ranked[name]:
            ws2.append([name, cand["concept"], cand["kind"], cand["unit"],
                        cand["taxonomy"], cand["companies"], cand["coverage"],
                        cand["avg_hit_rate"], "Y" if cand["negate"] else "",
                        ", ".join(cand["won_in"])])

    ws3 = wb.create_sheet("Coverage")
    ws3.append(["模板列", "kind", "答案卷家數"] + list(tickers))
    for c in ws3[1]:
        c.font = bold
    all_rows = sorted({r[0] for T in (IS_TEMPLATE, BS_TEMPLATE, CF_TEMPLATE) for r in T})
    for name in all_rows:
        won = {t for cand in ranked.get(name, [])[:1] for t in cand["won_in"]}
        ws3.append([name, row_kind.get(name, ""), answer_counts.get(name, 0)]
                   + ["O" if t in won else "" for t in tickers])

    for w in (ws, ws2, ws3):
        w.freeze_panes = "B2"
    wb.save(path)


def main(argv):
    tickers = argv[1:] or DEFAULT_50
    identity = config.load_config().get("identity")
    if not identity:
        print("沒有 SEC Identity", file=sys.stderr)
        return 1
    set_identity(identity)

    found, answer_counts, row_kind = derive(tickers, identity)
    ranked = rank(found, answer_counts)
    mapping = emit_mapping(ranked, row_kind, answer_counts)

    (CACHE / "mapping_candidates.json").write_text(
        json.dumps(ranked, indent=2, ensure_ascii=False), encoding="utf-8")
    (CACHE / "mapping.json").write_text(
        json.dumps(mapping, indent=2, ensure_ascii=False), encoding="utf-8")

    all_rows = {r[0] for T in (IS_TEMPLATE, BS_TEMPLATE, CF_TEMPLATE) for r in T}
    weak = sorted(set(ranked) - set(mapping))
    none_ = sorted(all_rows - set(ranked))

    print()
    print(f"{'=' * 92}")
    print(f"模板列共 {len(all_rows)}　採用 {len(mapping)}　"
          f"有候選但證據不足 {len(weak)}　完全找不到 {len(none_)}")
    print(f"{'=' * 92}")
    print()

    print("--- 採用（>=3 家命中、平均命中率 >=0.9）---")
    print(f"{'列名':34}{'concept':46}{'家數':>5}{'覆蓋':>7}{'命中':>7}")
    for r in sorted(mapping):
        m = mapping[r]
        ev = m["_evidence"]
        first = m["concepts"][0]
        neg = " [反號]" if m["negate"] else ""
        extra = f"  +{len(m['concepts']) - 1} 個備援" if len(m["concepts"]) > 1 else ""
        print(f"  {r[:32]:34}{first[:44]:46}{ev['won'][first]:>5}"
              f"{ev['won'][first] / max(ev['answer_companies'], 1):>7.0%}"
              f"{ev['avg_hit_rate'][first]:>7.2f}{neg}{extra}")

    print()
    print("--- 有候選但證據不足（要人工看）---")
    for r in weak:
        top = ranked[r][0]
        print(f"  {r[:32]:34}{top['concept'][:44]:46}"
              f"{top['companies']:>5}{top['coverage']:>7.0%}{top['avg_hit_rate']:>7.2f}")

    print()
    print("--- 完全找不到候選（模板列可能要調整）---")
    for r in none_:
        print(f"  {r}   （答案卷家數 {answer_counts.get(r, 0)}）")

    print()
    xlsx = CACHE / "mapping_evidence.xlsx"
    export_xlsx(ranked, mapping, answer_counts, row_kind, tickers, xlsx)
    print(f"Excel（人眼看這份）：{xlsx}")
    print(f"可用 mapping：{CACHE / 'mapping.json'}")
    print(f"完整候選：    {CACHE / 'mapping_candidates.json'}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
