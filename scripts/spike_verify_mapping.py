"""spike_verify_mapping.py — 用 `facts_mapping` 實跑，逐格比對現行路徑（TODO G11 第四步）。

跟 `spike_derive_mapping.py` 的差別：那支是**推導**（找出 concept 對照），
這支是**驗收**（用推導出來的表實跑，看對得上幾成）。

用 `output/_spike/gaap_*.pkl` 的快取答案卷，**完全不打網路**，幾秒跑完 50 家。

輸出：
  - terminal 摘要：整體命中率、每一列的命中率（低的排前面）
  - `output/_spike/verify_mapping.xlsx`：每列 × 每家的命中率矩陣，方便人眼掃
"""
from __future__ import annotations

import json
import pickle
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

import facts_mapping as fm
import fetcher_facts as ff

CACHE = ROOT / "output" / "_spike"

# 兩邊差多少算對得上。XBRL 是精確值，但公司會改申報精度（千 → 百萬），
# 那不是抓錯，用相對誤差 0.1%。
_TOL = 1e-3


def _answer(gaap):
    out = {}
    for i, name in enumerate(gaap["concepts"]):
        if not name or name in out:
            continue
        series = {}
        for j, end in enumerate(gaap["ends"]):
            v = gaap["values"][i][j]
            if len(end or "") == 10 and isinstance(v, (int, float)):
                series[end] = float(v)
        if series:
            out[name] = series
    return out


def _fy_end_month(raw):
    for c in ("Revenues", "RevenueFromContractWithCustomerExcludingAssessedTax",
              "NetIncomeLoss"):
        s = ff.series_for_concept(raw, c, kind="annual", prefer="as_reported")
        if s:
            return int(sorted(s)[-1][5:7])
    return 12


def main(argv):
    tickers = argv[1:] or sorted(p.stem[5:] for p in CACHE.glob("gaap_*.pkl"))
    per_row = defaultdict(lambda: {"hit": 0, "flip": 0, "total": 0, "companies": set()})
    matrix = defaultdict(dict)
    tot_hit = tot = 0
    only_facts = only_gaap = 0

    for ticker in tickers:
        gaap_p, facts_p = CACHE / f"gaap_{ticker}.pkl", CACHE / f"facts_{ticker}.json"
        if not (gaap_p.exists() and facts_p.exists()):
            continue
        gaap = pickle.loads(gaap_p.read_bytes())
        raw = json.loads(facts_p.read_bytes())
        answers = _answer(gaap)

        tables = ff.build_statement_tables(
            raw, fm.ALL_MAPPINGS, fy_end_month=_fy_end_month(raw),
            ticker=ticker, prefer="as_reported")
        got = {}
        for t in tables:
            for i, name in enumerate(t.concepts):
                got[name] = {e: v for e, v in zip(t.period_ends, t.values[i])
                             if v is not None}

        for name, ans in answers.items():
            if name not in got:
                continue
            f = got[name]
            shared = set(ans) & set(f)
            if not shared:
                only_gaap += len(ans)
                continue
            def close(a, b):
                return abs(a - b) <= _TOL * max(abs(a), abs(b), 1.0)
            hit = sum(1 for e in shared if close(ans[e], f[e]))
            # 「只差正負號」另外算。那是慣例對不齊，不是抓錯數字——兩者的
            # 處理方式完全不同（一個改設定、一個要重查 concept），
            # 混在一起看不出問題到底在哪。
            flip = sum(1 for e in shared
                       if not close(ans[e], f[e]) and close(ans[e], -f[e]))
            per_row[name]["flip"] += flip
            per_row[name]["hit"] += hit
            per_row[name]["total"] += len(shared)
            per_row[name]["companies"].add(ticker)
            matrix[name][ticker] = round(hit / len(shared), 3)
            tot_hit += hit
            tot += len(shared)
            only_facts += len(set(f) - set(ans))
            only_gaap += len(set(ans) - set(f))

    print(f"\n公司 {len({t for r in per_row.values() for t in r['companies']})} 家")
    print(f"兩邊都有的格子 {tot}，其中數字相同 {tot_hit}  → **{tot_hit / max(tot, 1):.2%}**")
    print(f"只有 facts 有 {only_facts} 格（多抓到的），只有現行路徑有 {only_gaap} 格")

    tot_flip = sum(r["flip"] for r in per_row.values())
    print(f"其中「只差正負號」{tot_flip} 格 → 符號對齊後可達 "
          f"**{(tot_hit + tot_flip) / max(tot, 1):.2%}**")
    print()
    print(f"{'列名':36}{'命中率':>9}{'含符號':>9}{'格數':>8}{'家數':>6}")
    rows = sorted(per_row.items(), key=lambda kv: kv[1]["hit"] / max(kv[1]["total"], 1))
    for name, r in rows:
        rate = r["hit"] / max(r["total"], 1)
        with_sign = (r["hit"] + r["flip"]) / max(r["total"], 1)
        flag = ("  <<< 只差符號" if with_sign >= 0.95 > rate
                else "  <<< 要看" if with_sign < 0.95 else "")
        print(f"  {name[:34]:36}{rate:>9.2%}{with_sign:>9.2%}"
              f"{r['total']:>8}{len(r['companies']):>6}{flag}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Verify"
        cols = sorted({t for m in matrix.values() for t in m})
        ws.append(["模板列", "整體命中率", "格數", "家數"] + cols)
        for c in ws[1]:
            c.font = Font(bold=True)
        warn = PatternFill("solid", fgColor="FFFFF2CC")
        bad = PatternFill("solid", fgColor="FFFFD5D5")
        for name, r in rows:
            rate = r["hit"] / max(r["total"], 1)
            ws.append([name, round(rate, 4), r["total"], len(r["companies"])]
                      + [matrix[name].get(t, "") for t in cols])
            if rate < 0.8:
                for c in ws[ws.max_row]:
                    c.fill = bad
            elif rate < 0.95:
                for c in ws[ws.max_row]:
                    c.fill = warn
        ws.freeze_panes = "E2"
        out = CACHE / "verify_mapping.xlsx"
        wb.save(out)
        print(f"\nExcel：{out}")
    except Exception as e:
        print(f"Excel 匯出失敗（不影響上面的結論）：{type(e).__name__}: {e}",
              file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
