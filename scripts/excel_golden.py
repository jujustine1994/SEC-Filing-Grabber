# -*- coding: utf-8 -*-
"""Excel 輸出的逐格回歸驗收（golden file）。

**用途：改動 `excel_writer` / `excel_formatter` / `ratios` / `fiscal_input`
之前先存一份基準，改完再比對。** 單元測試驗的是邏輯，這支驗的是「真的產出來
的那份 xlsx 有沒有變」——2026-08-14 的多語言遷移就是靠它確認 480 條字串搬完
之後繁中輸出逐格不變。

作法：把 `output/_final/*.xlsx` 讀回成 StatementTable，走**真正的**
`write_statements` + `format_workbook` 重產一次，然後 dump 每一格的
值／數字格式／字型／粗體／底色。不打網路，所以可以隨便重跑。

    ./venv/Scripts/python.exe scripts/excel_golden.py make  <基準資料夾>
    # ...改 code...
    ./venv/Scripts/python.exe scripts/excel_golden.py make  <新資料夾>
    ./venv/Scripts/python.exe scripts/excel_golden.py check <基準> <新>

`check` 回傳 exit code 0 = 完全一致。

⚠ 重建 StatementTable 時會跳過 A 欄空白的分隔列，所以產出的**絕對列號**
比真實 pipeline 少幾列。要驗 README 保證的固定列位（Revenue 在第 8 列…）
請直接開 `output/_final/*.xlsx` 看，不要用這支的列號。這支驗的是
「基準與新版之間有沒有差異」，相對比對，不受影響。
"""
from __future__ import annotations

import glob
import json
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import load_workbook           # noqa: E402
from fetcher_gaap import StatementTable      # noqa: E402
from excel_writer import write_statements    # noqa: E402
from output_tables import append_ratio_table  # noqa: E402

SRC_GLOB = str(ROOT / "output" / "_final" / "*.xlsx")
DATA_START = 4          # D 欄


def _tables_from_xlsx(path: str) -> list[StatementTable]:
    """從產出的活頁簿反推 StatementTable。

    第 1 列 = ticker + 期間標籤，第 2 列 = 申報日，第 3 列起 = 資料。
    A=concept、B=譯文（重建時重查，不從檔案讀）、C=公司原文、D 起=數值。
    """
    wb = load_workbook(path)
    tables: list[StatementTable] = []
    ticker = Path(path).stem
    for sn in wb.sheetnames:
        if not sn.startswith("Data_"):
            continue
        # Data_Ratios / Data_Segments 是 output_tables 在組裝時衍生的，
        # 餵回去會變成兩份。
        if sn in ("Data_Ratios", "Data_Segments"):
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        dates_row = rows[1] if len(rows) > 1 else ()
        q_labels = [str(v) for v in header[DATA_START - 1:] if v is not None]
        f_dates = [("" if v is None else str(v))
                   for v in dates_row[DATA_START - 1:DATA_START - 1 + len(q_labels)]]
        concepts, orig, values = [], [], []
        for r in rows[2:]:
            if r[0] is None:
                continue
            concepts.append(str(r[0]))
            orig.append("" if r[2] is None else str(r[2]))
            vals = list(r[DATA_START - 1:DATA_START - 1 + len(q_labels)])
            vals += [None] * (len(q_labels) - len(vals))
            values.append(vals)
        tables.append(StatementTable(
            sheet_name=sn, quarter_labels=q_labels, filing_dates=f_dates,
            concepts=concepts, values=values, ticker=ticker, labels=orig,
        ))
    wb.close()
    return tables


def _dump(path: str) -> dict:
    """每一格：座標／值／數字格式／字型／粗體／底色。

    只比值不夠——「÷1M 沒套到」「百分比格式掉了」「字型混到別的」這幾種
    最常見的排版回歸，值都是對的。
    """
    wb = load_workbook(path)
    out: dict = {}
    for sn in wb.sheetnames:
        cells = []
        for row in wb[sn].iter_rows():
            for c in row:
                if c.value is None and c.fill.fgColor.rgb in (None, "00000000"):
                    continue
                cells.append([c.coordinate,
                              "" if c.value is None else str(c.value),
                              c.number_format, c.font.name,
                              bool(c.font.bold), str(c.fill.fgColor.rgb)])
        out[sn] = cells
    wb.close()
    return out


def make(outdir: str) -> None:
    os.makedirs(outdir, exist_ok=True)
    for src in sorted(glob.glob(SRC_GLOB)):
        if ".bak" in src:
            continue
        t = Path(src).stem
        tables = _tables_from_xlsx(src)
        append_ratio_table(tables)
        dst = os.path.join(outdir, f"{t}.xlsx")
        if os.path.exists(dst):
            os.remove(dst)
        write_statements(tables, dst)
        json.dump(_dump(dst),
                  open(os.path.join(outdir, f"{t}.json"), "w", encoding="utf-8"),
                  ensure_ascii=False, indent=0)
        print("made", t)


def check(base: str, new: str) -> int:
    bad = 0
    for bf in sorted(glob.glob(os.path.join(base, "*.json"))):
        t = Path(bf).stem
        nf = os.path.join(new, f"{t}.json")
        if not os.path.exists(nf):
            print(f"MISSING {t}")
            bad += 1
            continue
        b = json.load(open(bf, encoding="utf-8"))
        n = json.load(open(nf, encoding="utf-8"))
        if set(b) != set(n):
            print(f"{t}: sheet 清單不同 {set(b) ^ set(n)}")
            bad += 1
        for sn in b:
            if sn not in n:
                continue
            bm = {c[0]: c for c in b[sn]}
            nm = {c[0]: c for c in n[sn]}
            diffs = [k for k in set(bm) | set(nm) if bm.get(k) != nm.get(k)]
            if diffs:
                bad += 1
                print(f"{t} / {sn}: {len(diffs)} 格不同")
                for k in sorted(diffs)[:8]:
                    print(f"    {k}: {bm.get(k)}  ->  {nm.get(k)}")
    # Index!A2 帶當天日期，跨日重跑必定有一格差異——那不是回歸
    print("IDENTICAL" if bad == 0 else f"{bad} 組差異")
    return bad


if __name__ == "__main__":
    if len(sys.argv) < 2 or sys.argv[1] not in ("make", "check"):
        sys.exit(__doc__)
    if sys.argv[1] == "make":
        make(sys.argv[2])
    else:
        sys.exit(1 if check(sys.argv[2], sys.argv[3]) else 0)
